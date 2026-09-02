#!/usr/bin/env python3
"""
Guitar Center Used Inventory Tracker — Web App
------------------------------------------------
Run with:  python3 gc_tracker_app.py
Then open: http://localhost:5050
"""

import html as _html
import hmac
import json, os, re, sys, time, threading, queue, webbrowser, random, sqlite3
import gc as _gc, ctypes as _ctypes
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as _FutureTimeoutError
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path


def _sleep(base: float, jitter: float = 0.5):
    """Sleep for base ± jitter seconds to avoid looking like a bot."""
    time.sleep(max(0.1, base + random.uniform(-jitter, jitter)))

try:
    from flask import (Flask, request, jsonify, Response, stream_with_context,
                       session, redirect, send_file)
    from werkzeug.security import generate_password_hash, check_password_hash
except ImportError:
    sys.exit("Missing Flask. Run:  pip3 install flask requests openpyxl")

try:
    import requests as http
except ImportError:
    sys.exit("Missing requests. Run:  pip3 install flask requests openpyxl")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing openpyxl. Run:  pip3 install openpyxl")

try:
    from authlib.integrations.flask_client import OAuth as _AuthlibOAuth
    _AUTHLIB_AVAILABLE = True
except ImportError:
    _AUTHLIB_AVAILABLE = False

try:
    import psycopg2
    import psycopg2.extras
    _PSYCOPG2_AVAILABLE = True
except ImportError:
    _PSYCOPG2_AVAILABLE = False

# ── Paths & config ────────────────────────────────────────────────────────────
SCRIPT_DIR     = Path(__file__).parent
DATA_DIR       = Path(os.environ.get("DATA_DIR", SCRIPT_DIR))
DATA_DIR.mkdir(parents=True, exist_ok=True)

STATE_FILE     = DATA_DIR / "gc_state.json"
OUTPUT_FILE    = DATA_DIR / "gc_new_inventory.xlsx"
STORES_CACHE   = DATA_DIR / "gc_stores_cache.json"
FAVORITES_FILE = DATA_DIR / "gc_favorites.json"
CAT_CACHE_FILE = DATA_DIR / "gc_category_cache.json"
WATCHLIST_FILE   = DATA_DIR / "gc_watchlist.json"
KEYWORDS_FILE    = DATA_DIR / "gc_keywords.json"
STORE_COORDS_FILE    = DATA_DIR / "gc_store_coords.json"
NEW_DEALS_CACHE_FILE = DATA_DIR / "gc_new_deals_cache.json"


PORT              = int(os.environ.get("PORT", 5050))
APP_PASSWORD      = (os.environ.get("APP_PASSWORD") or "").strip()
GA_MEASUREMENT_ID = os.environ.get("GA_MEASUREMENT_ID", "").strip()
ADMIN_EMAIL       = (os.environ.get("ADMIN_EMAIL") or "").strip().lower()
PG_DATABASE_URL   = (os.environ.get("DATABASE_URL") or "").strip()

# ── User accounts (SQLite) ────────────────────────────────────────────────────
USER_DB = DATA_DIR / "gc_users.db"

def _user_db():
    """Open a connection to the user database."""
    conn = sqlite3.connect(str(USER_DB))
    conn.row_factory = sqlite3.Row
    return conn


# ── Daily user-DB backup (v2.15.3, 2026-07 audit "do now" #1) ─────────────────
# gc_users.db is the only non-regenerable data on the volume. VACUUM INTO makes an
# atomic snapshot from a live DB; we keep the last 7 dailies in DATA_DIR/backups.
# Piggybacks on the after_request hook — first request of each UTC day pays ~ms.
_BACKUP_DIR  = DATA_DIR / "backups"
_BACKUP_KEEP = 7
_backup_lock = threading.Lock()
_last_backup_day = None

def _maybe_backup_users_db():
    global _last_backup_day
    today = datetime.utcnow().strftime("%Y%m%d")
    if _last_backup_day == today:
        return
    if not _backup_lock.acquire(blocking=False):
        return  # another request thread is on it
    try:
        if _last_backup_day == today:
            return
        dest = _BACKUP_DIR / f"gc_users_{today}.db"
        if not dest.exists() and USER_DB.exists():
            _BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            tmp = _BACKUP_DIR / f".gc_users_{today}.db.tmp"
            if tmp.exists():
                tmp.unlink()
            conn = sqlite3.connect(str(USER_DB))
            try:
                conn.execute(f"VACUUM INTO '{tmp}'")  # path is fully server-controlled
            finally:
                conn.close()
            os.replace(tmp, dest)
            for old in sorted(_BACKUP_DIR.glob("gc_users_*.db"))[:-_BACKUP_KEEP]:
                old.unlink()
        _last_backup_day = today  # set even on skip; retry next UTC day
    except Exception as e:
        print(f"[backup] users-db backup failed: {type(e).__name__}: {e}")
        _last_backup_day = today  # don't hammer a persistently failing backup
    finally:
        _backup_lock.release()

def _init_user_db():
    with _user_db() as conn:
        # WAL: better concurrent read/write behavior on the threaded server; the
        # setting is persistent in the DB file, so once is enough. (2026-07 audit)
        try:
            conn.execute("PRAGMA journal_mode=WAL")
        except sqlite3.OperationalError:
            pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT    UNIQUE NOT NULL COLLATE NOCASE,
                email         TEXT    UNIQUE COLLATE NOCASE,
                password_hash TEXT,
                google_id     TEXT    UNIQUE,
                created_at    TEXT    NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_data (
                user_id        INTEGER PRIMARY KEY REFERENCES users(id),
                watchlist      TEXT    DEFAULT '{}',
                keywords       TEXT    DEFAULT '[]',
                favorites      TEXT    DEFAULT '[]',
                last_run       TEXT    DEFAULT '',
                new_ids        TEXT    DEFAULT '[]',
                saved_searches TEXT    DEFAULT '[]',
                last_anchor    TEXT    DEFAULT '',
                updated_at     TEXT    DEFAULT ''
            )
        """)
        # Migration: add saved_searches column for existing databases
        try:
            conn.execute("ALTER TABLE user_data ADD COLUMN saved_searches TEXT DEFAULT '[]'")
        except Exception:
            pass  # Column already exists
        # Migration: add last_anchor column for existing databases (v2.10.18)
        # Per-user anchor for NEW detection — replaces the buggy global-cache anchor
        # which was contaminated by other users' scans.
        try:
            conn.execute("ALTER TABLE user_data ADD COLUMN last_anchor TEXT DEFAULT ''")
        except Exception:
            pass  # Column already exists
        # Migration: add google_id column for existing databases
        # NOTE: SQLite ALTER TABLE ADD COLUMN cannot include UNIQUE — add column
        # first, then create the index separately.
        try:
            conn.execute("ALTER TABLE users ADD COLUMN google_id TEXT")
        except Exception:
            pass  # Column already exists
        try:
            conn.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_users_google_id
                ON users(google_id) WHERE google_id IS NOT NULL
            """)
        except Exception:
            pass
        # Migration: add deleted_at column for soft-delete / scheduled deletion (v2.11.2)
        try:
            conn.execute("ALTER TABLE users ADD COLUMN deleted_at TEXT")
        except Exception:
            pass  # Column already exists
        # Migration: add last_login column (v2.12.4)
        try:
            conn.execute("ALTER TABLE users ADD COLUMN last_login TEXT")
        except Exception:
            pass  # Column already exists
        conn.commit()

def _user_by_username(username: str) -> dict | None:
    with _user_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE username=?", (username.strip(),)).fetchone()
        return dict(row) if row else None

def _touch_last_login(user_id: int) -> None:
    """Stamp the current UTC time as this user's last login."""
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    with _user_db() as conn:
        conn.execute("UPDATE users SET last_login=? WHERE id=?", (now, user_id))
        conn.commit()

def _user_by_id(user_id: int) -> dict | None:
    with _user_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
        return dict(row) if row else None

def _user_by_email(email: str) -> dict | None:
    with _user_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE email=?", (email.strip().lower(),)).fetchone()
        return dict(row) if row else None

def _user_by_google_id(google_id: str) -> dict | None:
    with _user_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE google_id=?", (google_id,)).fetchone()
        return dict(row) if row else None

def _gen_google_username(display_name: str) -> str:
    """Generate a unique username from a Google display name."""
    base = re.sub(r'[^A-Za-z0-9_\-]', '', display_name.replace(' ', '_'))[:25]
    if len(base) < 3:
        base = 'user'
    candidate = base
    i = 1
    with _user_db() as conn:
        while conn.execute("SELECT id FROM users WHERE username=?", (candidate,)).fetchone():
            candidate = f"{base}_{i}"
            i += 1
    return candidate

def _get_user_data(user_id: int) -> dict:
    with _user_db() as conn:
        row = conn.execute("SELECT * FROM user_data WHERE user_id=?", (user_id,)).fetchone()
    if not row:
        return {"watchlist": {}, "keywords": [], "favorites": [], "last_run": "", "new_ids": [], "saved_searches": [], "last_anchor": ""}
    try:
        ss = json.loads(row["saved_searches"] or "[]")
    except Exception:
        ss = []
    # last_anchor column may not exist on rows from before the migration ran in
    # this process; sqlite3.Row raises IndexError for missing keys, so guard it.
    try:
        last_anchor = row["last_anchor"] or ""
    except (KeyError, IndexError):
        last_anchor = ""
    return {
        "watchlist":      json.loads(row["watchlist"] or "{}"),
        "keywords":       json.loads(row["keywords"]  or "[]"),
        "favorites":      json.loads(row["favorites"] or "[]"),
        "last_run":       row["last_run"] or "",
        "new_ids":        json.loads(row["new_ids"]   or "[]"),
        "saved_searches": ss,
        "last_anchor":    last_anchor,
    }

def _set_user_data(user_id: int, **kwargs):
    """Update one or more user_data fields. Valid keys: watchlist, keywords, favorites, last_run, new_ids, saved_searches, last_anchor"""
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    with _user_db() as conn:
        conn.execute(
            "INSERT OR IGNORE INTO user_data (user_id, updated_at) VALUES (?,?)",
            (user_id, now)
        )
        for field, value in kwargs.items():
            if field in ("watchlist", "keywords", "favorites", "new_ids", "saved_searches"):
                conn.execute(
                    f"UPDATE user_data SET {field}=?, updated_at=? WHERE user_id=?",
                    (json.dumps(value), now, user_id)
                )
            elif field in ("last_run", "last_anchor"):
                # Plain TEXT fields — stored as-is (not JSON-encoded)
                try:
                    conn.execute(
                        f"UPDATE user_data SET {field}=?, updated_at=? WHERE user_id=?",
                        (value, now, user_id)
                    )
                except sqlite3.OperationalError:
                    # last_anchor column missing on this connection (very rare —
                    # migration runs at startup, but tolerate it anyway)
                    pass
        conn.commit()

_init_user_db()

# ── HTTP session ──────────────────────────────────────────────────────────────
_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2.1 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]

_HEADERS = {
    "User-Agent":                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept":                    "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language":           "en-US,en;q=0.9",
    "Accept-Encoding":           "gzip, deflate, br",
    "Connection":                "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest":            "document",
    "Sec-Fetch-Mode":            "navigate",
    "Sec-Fetch-Site":            "none",
    "Sec-Fetch-User":            "?1",
    "Cache-Control":             "max-age=0",
}

_http = http.Session()
_http.headers.update(_HEADERS)

# Load persisted cookies if available
COOKIE_FILE = DATA_DIR / "gc_cookies.json"

def _load_cookies():
    if COOKIE_FILE.exists():
        try:
            cookies = json.loads(COOKIE_FILE.read_text())
            _http.cookies.update(cookies)
        except Exception:
            pass

def _save_cookies():
    try:
        COOKIE_FILE.write_text(json.dumps(dict(_http.cookies)))
    except Exception:
        pass

def _rotate_ua():
    """Pick a random User-Agent for the next request."""
    _http.headers["User-Agent"] = random.choice(_USER_AGENTS)

# ── Category cache ────────────────────────────────────────────────────────────
_cat_cache: dict = {}
_cat_cache_mtime = None   # mtime of the file at last parse; skip re-parse when unchanged

def _load_cat_cache():
    # The category cache is a ~50MB JSON of ~92K items. It used to be re-read and
    # re-parsed from disk on EVERY call — and /api/browse calls it on every keystroke,
    # filter, sort, and page flip. That parse is ~400ms and holds the GIL, so on the
    # threaded dev server it serialized all request threads behind it. Memoize by file
    # mtime: only re-parse when the file actually changed (i.e. after a scan saves).
    # A stat() guard costs ~1µs. (v2.13.0)
    global _cat_cache, _cat_cache_mtime
    try:
        mtime = CAT_CACHE_FILE.stat().st_mtime
    except OSError:
        # File missing (fresh deploy, or an admin reset deleted it) — leave whatever is
        # in memory. Reset sets _cat_cache = {} itself, so this won't resurrect data.
        return
    if mtime == _cat_cache_mtime and _cat_cache:
        return
    try:
        _cat_cache = json.loads(CAT_CACHE_FILE.read_text())
        _cat_cache_mtime = mtime
    except Exception:
        # Corrupt/partial file (e.g. a crash during a non-atomic write on an older
        # build). Do NOT blank the catalog for every user — keep the last good
        # in-memory cache. If we have data, advance mtime so we stop re-reading the
        # bad file on every browse; a later good write bumps mtime and we reload.
        # If we have nothing yet, leave mtime unset so we keep retrying.
        if not _cat_cache:
            _cat_cache_mtime = None
        else:
            _cat_cache_mtime = mtime

def _save_cat_cache():
    # Atomic write: a crash/redeploy partway through writing the ~53MB file used to
    # leave it truncated, after which json.loads raised and the in-memory catalog was
    # reset to {} — an empty site for every user until the next full scan. Write to a
    # temp file then os.replace() (atomic on POSIX) so readers only ever see a complete
    # file.
    try:
        tmp = CAT_CACHE_FILE.parent / (CAT_CACHE_FILE.name + ".tmp")
        tmp.write_text(json.dumps(_cat_cache))
        os.replace(tmp, CAT_CACHE_FILE)
    except Exception:
        pass

# ── Postgres catalog store (Phase A, v2.16.14) ────────────────────────────────
# Migration target for the flat-JSON _cat_cache above (~92K items, ~53MB) — see
# POSTGRES_MIGRATION_PLAN.md for the full plan and rationale. This phase ONLY
# creates the schema at startup; nothing on the request path reads from or
# writes to Postgres yet, and _cat_cache / gc_category_cache.json remains the
# sole source of truth. Guarded end-to-end so a missing DATABASE_URL (local
# dev, or before the Railway variable reference is actually deployed), a
# missing psycopg2 install, or an unreachable DB can never block app startup —
# same defensive posture as _load_cat_cache() / _maybe_backup_users_db().
#
# Schema lives in pg_schema.sql (repo root) rather than inline here so the
# one-off backfill script (migrate_cat_cache_to_pg.py) can apply the exact
# same DDL without duplicating it — see that script's docstring.
PG_SCHEMA_FILE = SCRIPT_DIR / "pg_schema.sql"

def _init_pg_schema():
    if not (_PSYCOPG2_AVAILABLE and PG_DATABASE_URL):
        return
    try:
        schema_sql = PG_SCHEMA_FILE.read_text()
    except OSError as e:
        print(f"[pg] schema init skipped — can't read {PG_SCHEMA_FILE.name}: {e}")
        return
    try:
        conn = psycopg2.connect(PG_DATABASE_URL, connect_timeout=10)
        try:
            with conn.cursor() as cur:
                cur.execute(schema_sql)
            conn.commit()
        finally:
            conn.close()
        print("[pg] items table ready")
    except Exception as e:
        # Never let a Postgres hiccup block startup — same tolerance as every
        # other best-effort persistence path in this app.
        print(f"[pg] schema init skipped: {type(e).__name__}: {e}")

_init_pg_schema()

# ── Postgres connection pool (Phase C, v2.16.20) ──────────────────────────────
# The scan-triggered/admin-only Postgres paths above (_init_pg_schema,
# _pg_sync_scan, _pg_full_backfill, _pg_parity_check) each open one ad-hoc
# psycopg2.connect() per call — fine for rare, background-thread-triggered
# work, but not for something hit on every /api/browse request once Phase D
# cuts over. See POSTGRES_MIGRATION_PLAN.md §5. Sized for this app's
# `--workers=1 --worker-class=gthread --threads=8` Procfile: up to 8 concurrent
# request threads + the scan thread + headroom for the admin-only paths above
# (which still open their own ad-hoc connections — not migrated to the pool,
# since they're rare and not on any hot path; migrating them is a follow-up,
# not required for Phase C). Created once at module load, torn down never
# (lives for the process lifetime, same as _cat_cache).
_PG_POOL = None
if _PSYCOPG2_AVAILABLE and PG_DATABASE_URL:
    try:
        import psycopg2.pool
        _PG_POOL = psycopg2.pool.ThreadedConnectionPool(
            2, 12, PG_DATABASE_URL, connect_timeout=10)
        print("[pg] connection pool ready (minconn=2, maxconn=12)")
    except Exception as e:
        # Same tolerance as _init_pg_schema() — a pool that fails to init just
        # means Tier 1 shadow reads no-op (see _pg_tier1_eligible's caller in
        # api_browse); it can never block startup or any other Postgres path.
        print(f"[pg] connection pool init skipped: {type(e).__name__}: {e}")
        _PG_POOL = None

import contextlib as _contextlib

@_contextlib.contextmanager
def _pg_conn():
    """Yield a pooled Postgres connection, committing on a clean exit and
    rolling back + discarding the connection back to the pool on error —
    mirrors _user_db()'s `with`-based pattern for SQLite, just pooled since
    Postgres connections are more expensive to establish than SQLite's."""
    if _PG_POOL is None:
        raise RuntimeError("Postgres connection pool not available")
    conn = _PG_POOL.getconn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        _PG_POOL.putconn(conn)

# ── Postgres dual-write (Phase B, v2.16.15) ───────────────────────────────────
# Mirrors this scan's results into Postgres, best-effort, AFTER _cat_cache/JSON
# have already been fully updated by the normal scan logic in _run() below —
# this function only writes what the JSON path already decided; it is never a
# second source of scan logic, and a failure here can never affect scan
# completion, the JSON cache, or anything the user sees. Purpose: let Postgres
# accumulate real production scan data (Phase B of POSTGRES_MIGRATION_PLAN.md)
# so it can be diffed against the JSON-based system before anything ever reads
# from it. Silently no-ops if DATABASE_URL isn't set or psycopg2 isn't
# installed — same guard as _init_pg_schema().
#
# Runs in its own background thread (see the call site in _run()) so a slow
# Postgres write never delays the scan's "done" SSE message — dual-write is
# purely for building confidence, not something anything currently depends on.
_PG_UPSERT_COLS = ["sku", "name", "brand", "category", "subcategory", "condition",
                   "condition_note", "price", "list_price", "has_price_drop", "price_drop",
                   "price_drop_since", "store", "location", "url", "image_id", "is_vintage",
                   "available", "date_listed", "first_seen"]
# Keep in sync with migrate_cat_cache_to_pg.py's COLS if the schema ever changes.
_PG_UPSERT_SQL = (
    f"INSERT INTO items ({', '.join(_PG_UPSERT_COLS)}) VALUES %s "
    f"ON CONFLICT (sku) DO UPDATE SET "
    f"{', '.join(f'{c}=EXCLUDED.{c}' for c in _PG_UPSERT_COLS if c != 'sku')}"
)

def _pg_row_for(sku: str, it: dict) -> tuple:
    return (
        sku, it.get("name", "") or "", it.get("brand", "") or "", it.get("category", "") or "",
        it.get("subcategory", "") or "", it.get("condition", "") or "",
        it.get("condition_note", "") or "", it.get("price", 0) or 0, it.get("list_price", 0) or 0,
        bool(it.get("has_price_drop", False)), it.get("price_drop", 0) or 0,
        it.get("price_drop_since", "") or "", it.get("store", "") or "",
        it.get("location", "") or it.get("store", "") or "", it.get("url", "") or "",
        it.get("image_id", "") or "", bool(it.get("is_vintage", False)),
        it.get("available", True), it.get("date_listed", "") or "", it.get("first_seen", "") or "",
    )

def _pg_sync_scan(ids_this_run: set, nationwide: bool, scan_incomplete: bool,
                   incomplete_stores: set, stores_to_scan: list):
    """Best-effort Postgres mirror of one scan. See module docstring above.
    Reads _cat_cache (read-only) and the same run-scoped values _run() already
    computed — never recomputes scan logic, only replays the JSON path's own
    two effects: (1) upsert every item this run touched, (2) mark unavailable,
    in Postgres, whatever the JSON sold-marking block just marked unavailable
    in _cat_cache, under the IDENTICAL coverage-gap-safe condition and store
    scope (v2.16.11) — see _run()'s own sold-marking block for that reasoning."""
    if not (_PSYCOPG2_AVAILABLE and PG_DATABASE_URL):
        return
    try:
        conn = psycopg2.connect(PG_DATABASE_URL, connect_timeout=10)
        try:
            with conn.cursor() as cur:
                if ids_this_run:
                    rows = [_pg_row_for(sku, _cat_cache[sku]) for sku in ids_this_run if sku in _cat_cache]
                    if rows:
                        psycopg2.extras.execute_values(cur, _PG_UPSERT_SQL, rows, page_size=2000)
                # Mirrors _run()'s "Mark sold items" block's exact condition and scope.
                if not _stop_event.is_set() and (not nationwide or not scan_incomplete):
                    cur.execute("CREATE TEMP TABLE _run_skus (sku TEXT PRIMARY KEY) ON COMMIT DROP")
                    if ids_this_run:
                        psycopg2.extras.execute_values(
                            cur, "INSERT INTO _run_skus (sku) VALUES %s",
                            [(s,) for s in ids_this_run], page_size=5000,
                        )
                    if nationwide:
                        cur.execute("""
                            UPDATE items SET available=false
                            WHERE available
                            AND NOT EXISTS (SELECT 1 FROM _run_skus WHERE _run_skus.sku = items.sku)
                        """)
                    else:
                        scanned_store_set = list(set(stores_to_scan) - incomplete_stores)
                        if scanned_store_set:
                            cur.execute("""
                                UPDATE items SET available=false
                                WHERE available
                                AND NOT EXISTS (SELECT 1 FROM _run_skus WHERE _run_skus.sku = items.sku)
                                AND store = ANY(%s)
                            """, (scanned_store_set,))
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        # Never let a Postgres hiccup surface anywhere a user or the scan logic
        # would see it — this is purely a background confidence-building mirror.
        print(f"[pg] scan dual-write skipped: {type(e).__name__}: {e}")

# ── Postgres full backfill (Phase A fix, v2.16.18) ────────────────────────────
# One-time, admin-triggered corrective re-sync of the Postgres `items` table
# from the LIVE in-process _cat_cache. Fixes Phase A's original backfill
# (migrate_cat_cache_to_pg.py, run against a stale local gc_category_cache.json
# snapshot from 2026-04-29) — /api/pg-parity-check confirmed on 2026-09-02 that
# Postgres was missing 255,886 of 436,325 live SKUs, plus stale field values on
# the ~180K rows it did have. See POSTGRES_MIGRATION_PLAN.md for the full plan.
#
# Deliberately full-catalog, not active-only: Chuck decided 2026-09-02 to keep
# _cat_cache's full sold/delisted history on purpose (future price-over-time /
# average-used-price analytics), so this backfills every sku currently in
# _cat_cache regardless of `available`, matching what the JSON has always had.
#
# Upsert-only — never deletes or marks-unavailable anything already in
# Postgres that isn't in _cat_cache. The 2026-09-02 parity check found
# extra_in_pg was already 0, so there's nothing to clean up; if that ever
# changes, treat it as a separate investigation, not something this backfill
# should silently paper over.
#
# Reuses the exact _PG_UPSERT_SQL / _pg_row_for already used by _pg_sync_scan —
# same upsert shape, just applied to the full cache instead of one scan's
# ids_this_run. Runs in a background thread via the same _lock/_q pattern as
# _validate_stores / _fill_gaps, so the ~440K-row write — measured at ~15K
# rows/sec against a local scratch Postgres (436,240 synthetic rows in ~25-29s,
# including 180K seeded-stale rows needing correction); real timing against
# Railway's network/DB will vary — never blocks the request or risks a proxy
# timeout. Commits per chunk, so an interrupted run (stop button, deploy,
# crash) leaves whatever it finished already correct in Postgres; re-running
# it is idempotent (verified in the scratch-Postgres test: an
# interrupted-then-resumed run produces the same end state as one
# uninterrupted run — see /tmp/pgtest/test_backfill.py from the session that
# built this).
def _pg_full_backfill():
    """Full upsert of every _cat_cache item into Postgres. See module comment above."""
    def send(msg): _q.put(msg)
    total = 0
    done = 0
    try:
        if not (_PSYCOPG2_AVAILABLE and PG_DATABASE_URL):
            send({"type": "progress", "msg": "Postgres not configured — aborting."})
            return
        _load_cat_cache()
        skus = list(_cat_cache.keys())
        total = len(skus)
        send({"type": "progress", "msg": f"Starting full Postgres backfill from live _cat_cache: "
                                          f"{total:,} items (all statuses — available and "
                                          f"sold/delisted history both included)."})
        CHUNK = 5000
        conn = psycopg2.connect(PG_DATABASE_URL, connect_timeout=10)
        conn.autocommit = False
        t0 = time.time()
        last_report = 0
        try:
            with conn.cursor() as cur:
                for i in range(0, total, CHUNK):
                    if _stop_event.is_set():
                        send({"type": "progress",
                              "msg": f"⏹ Stopped by user after {done:,}/{total:,} "
                                     "(already-upserted rows remain correct; safe to re-run)."})
                        break
                    batch = skus[i:i + CHUNK]
                    rows = [_pg_row_for(sku, _cat_cache[sku]) for sku in batch if sku in _cat_cache]
                    if rows:
                        psycopg2.extras.execute_values(cur, _PG_UPSERT_SQL, rows, page_size=2000)
                        conn.commit()
                    done += len(batch)
                    if done - last_report >= 50000 or done == total:
                        elapsed = time.time() - t0
                        rate = done / elapsed if elapsed > 0 else 0
                        send({"type": "progress",
                              "msg": f"  upserted {done:,}/{total:,} "
                                     f"({elapsed:.0f}s elapsed, ~{rate:,.0f} rows/sec)"})
                        last_report = done
        finally:
            conn.close()
        elapsed = time.time() - t0
        send({"type": "progress",
              "msg": f"\n✓ Full backfill done: {done:,} rows upserted in {elapsed:.0f}s. "
                     f"Run /api/pg-parity-check next to confirm pg_total == json_total."})
        send({"type": "done", "baseline": False, "stopped": _stop_event.is_set(),
              "new_ids": [], "items": []})
    except Exception as e:
        # Don't leak exception text over the (public) SSE stream — log it server-side.
        # (same posture as _validate_stores / 2026-07 audit L3)
        print(f"[pg] full backfill failed after {done:,}/{total:,}: {type(e).__name__}: {e}")
        send({"type": "done", "error": "Backfill failed — see server logs.",
              "new_ids": [], "items": []})
    finally:
        _lock.release()

# ── Postgres parity check (Phase B verification, v2.16.16) ───────────────────
# Admin-only, on-demand, read-only diagnostic: compares the live in-memory
# _cat_cache (the JSON path's source of truth) against the live Postgres
# `items` table, to check whether Phase B's dual-write has actually kept them
# in sync across real production scans. Not on any user-facing path and
# doesn't touch either data source — pure comparison. See
# POSTGRES_MIGRATION_PLAN.md Phase C and NEXT_SESSION_PROMPT.md.
#
# Verified against a scratch Postgres before shipping: seeded deliberate
# price/available/date_listed drift plus missing/extra rows on each side and
# confirmed the diff catches exactly those and nothing else (a $0.01 price
# rounding difference is intentionally NOT flagged); also timed at
# production scale (~92K rows, synthetic) at well under 1s.
def _pg_parity_check() -> dict:
    _load_cat_cache()
    conn = psycopg2.connect(PG_DATABASE_URL, connect_timeout=10)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT sku, available, price, date_listed, store FROM items")
            pg_rows = {
                r[0]: {"available": r[1], "price": float(r[2]) if r[2] is not None else None,
                       "date_listed": r[3], "store": r[4]}
                for r in cur.fetchall()
            }
    finally:
        conn.close()

    json_skus = set(_cat_cache.keys())
    pg_skus   = set(pg_rows.keys())
    missing_in_pg = json_skus - pg_skus   # in JSON, never made it to Postgres
    extra_in_pg   = pg_skus - json_skus   # in Postgres, not in JSON (shouldn't happen)

    mismatches = []
    for sku in json_skus & pg_skus:
        j, p = _cat_cache[sku], pg_rows[sku]
        diffs = {}
        if bool(j.get("available", True)) != bool(p["available"]):
            diffs["available"] = [j.get("available"), p["available"]]
        jp = j.get("price")
        if jp is not None and p["price"] is not None and abs(float(jp) - p["price"]) > 0.01:
            diffs["price"] = [jp, p["price"]]
        if (j.get("date_listed") or "") != (p["date_listed"] or ""):
            diffs["date_listed"] = [j.get("date_listed"), p["date_listed"]]
        if diffs:
            mismatches.append({"sku": sku, "diffs": diffs})

    return {
        "json_total":             len(json_skus),
        "pg_total":                len(pg_skus),
        "missing_in_pg":           len(missing_in_pg),
        "missing_in_pg_sample":    sorted(missing_in_pg)[:20],
        "extra_in_pg":             len(extra_in_pg),
        "extra_in_pg_sample":      sorted(extra_in_pg)[:20],
        "field_mismatches":        len(mismatches),
        "field_mismatches_sample": mismatches[:20],
        "checked_at":              datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    }

# ── Memoized /api/browse base item list (E2, v2.16.4) ────────────────────────
# /api/browse used to rebuild this ~92K-item lightweight-dict list from scratch
# on EVERY call — price/date formatting + name/brand lowercasing for every item
# — even though none of that depends on who's asking or what they're filtering
# by. That's ~110-230ms held under the GIL per call, on the threaded dev server,
# on an endpoint hit by every filter click, keystroke, and page flip. Memoize
# the expensive request-independent projection once per cache generation (keyed
# off _cat_cache_mtime — the same trigger _load_cat_cache already uses) so each
# request only pays for a cheap second pass: store-selection filter, per-user
# scan-date gating, and the per-user annotations (watched/isNew/kwMatch/isFav)
# that genuinely can't be cached.
#
# Deliberately memoizes ONE unfiltered list, not one per store-selection — with
# up to 298 stores selectable in any combination, caching per-selection would
# chase a combinatorial key space and blow memory. Store filtering stays a
# cheap per-request pass over the single cached list instead.
#
# No lock: same unsynchronized memoize-by-mtime pattern as _load_cat_cache
# (v2.13.0) — under Flask's threaded dev server + the GIL, two threads racing
# to rebuild after a scan just do redundant identical work, not incorrect work
# (both read the same _cat_cache snapshot via list(...)).
_base_item_list: list = []
_base_item_list_mtime = None

def _build_base_item_list():
    """The memoized half of the old all_items loop: every field that's the
    same no matter who's asking, for every AVAILABLE item (the availability
    filter is static per item, so items failing it are dropped here rather
    than carried through every request). Does NOT include store filtering,
    scan-date gating, or the per-user watched/isNew/kwMatch/isFav flags —
    those stay a per-request pass in api_browse(). `first_seen`/`name_lower`/
    `brand_lower` are internal-only (used for gating/matching), not part of
    the JSON shape ultimately returned to the client."""
    global _base_item_list, _base_item_list_mtime
    if _base_item_list_mtime == _cat_cache_mtime and _base_item_list:
        return _base_item_list
    items = []
    for sku, cached in list(_cat_cache.items()):
        if not cached.get("available", True):
            continue
        price_raw = cached.get("price", 0) or 0
        name      = cached.get("name", "")
        brand     = cached.get("brand", "")
        date_raw  = cached.get("date_listed", "")
        items.append({
            "id":               sku,
            "name":             name,
            "brand":            brand,
            "name_lower":       name.lower(),
            "brand_lower":      brand.lower(),
            "price":            f"${price_raw:,.2f}" if price_raw else "",
            "price_raw":        price_raw,
            "list_price_raw":   cached.get("list_price", 0) or 0,
            "price_drop":       cached.get("price_drop", 0) or 0,
            "price_drop_since": cached.get("price_drop_since", "") or "",
            "store":            cached.get("store", ""),
            "location":         cached.get("location") or cached.get("store", ""),
            "url":              cached.get("url", ""),
            "category":         cached.get("category", ""),
            "subcategory":      cached.get("subcategory", ""),
            "condition":        cached.get("condition", ""),
            "date":             _fmt_date(date_raw),
            "date_raw":         date_raw,
            "first_seen":       cached.get("first_seen", ""),
            "image_id":         cached.get("image_id", ""),
            "is_vintage":       bool(cached.get("is_vintage")),
            "condition_note":   cached.get("condition_note", "") or "",
        })
    _base_item_list = items
    _base_item_list_mtime = _cat_cache_mtime
    return _base_item_list

# ── Store list ────────────────────────────────────────────────────────────────

FALLBACK_STORES: list[str] = []  # Populated from GC live data via Validate Stores


def get_store_list() -> list[str]:
    cached = []
    if STORES_CACHE.exists():
        try:
            cached = json.loads(STORES_CACHE.read_text()).get("stores", [])
        except Exception:
            pass
    blocklist = _get_blocklist()
    return sorted(set(cached) - blocklist)


_US_STATES = [
    "al","ak","az","ar","ca","co","ct","de","fl","ga","hi","id","il","in",
    "ia","ks","ky","la","me","md","ma","mi","mn","ms","mo","mt","ne","nv",
    "nh","nj","nm","ny","nc","nd","oh","ok","or","pa","ri","sc","sd","tn",
    "tx","ut","vt","va","wa","wv","wi","wy","dc",
]

def _fetch_state_stores(state: str) -> list[str]:
    """Fetch store city names for a single state from stores.guitarcenter.com.
    Only extracts names from confirmed store-page URLs matching /{state}/{city}/{id},
    which is the only reliable signal — headings and link text pick up nav garbage."""
    return [name for name, _ in _fetch_state_stores_with_state(state)]


def _fetch_state_stores_with_state(state: str) -> list[tuple]:
    """Like _fetch_state_stores but returns (name, state_abbr) tuples."""
    try:
        r = _http.get(f"https://stores.guitarcenter.com/{state}/", timeout=10)
        if r.status_code != 200:
            return []
        html = r.text
        # Only trust URLs in the form /state/city-slug/numeric-id
        slug_to_name = {}
        for slug in re.findall(
            rf'href="/{re.escape(state)}/([a-z][a-z0-9\-]+)/(\d+)(?:/[^"]*)?"',
            html
        ):
            city_slug, store_id = slug
            name = " ".join(w.capitalize() for w in city_slug.split("-"))
            slug_to_name[city_slug] = name
        return [(name, state.upper()) for name in slug_to_name.values()]
    except Exception:
        return []


def _build_store_coords(send_progress=None, force: bool = False):
    """Geocode all known stores using Algolia's 'storeName' field as the query.

    For each store, we first pull ONE hit from Algolia to discover the human-readable
    storeName (e.g. 'South Austin, TX' — a real neighborhood Nominatim recognises),
    then geocode that string directly via Nominatim. This is vastly more reliable than
    querying 'Guitar Center {store}' because GC stores aren't POIs in Nominatim's DB,
    but their storeName strings are real geographic places.

    Stores with zero live used inventory (e.g. closed stores still in the cache) can't
    be resolved via Algolia and are reported as 'no items'. They fall back to a
    last-ditch '{store}, {state}' Nominatim query using state context from GC's state
    pages, so permanently-closed stores still have SOME chance of getting coords.

    If force=True, re-geocodes all stores even if already in the coords file.
    Writes gc_store_coords.json. Returns {store: {lat, lng, source}}.
    """
    def _send(msg):
        if send_progress:
            send_progress(msg)

    known_stores = get_store_list()
    total = len(known_stores)
    _send(f"Step 1: Found {total} stores in cache.")

    # Load existing coords so we can skip already-geocoded stores (unless force=True).
    existing: dict = {}
    if STORE_COORDS_FILE.exists():
        try:
            existing = json.loads(STORE_COORDS_FILE.read_text())
        except Exception:
            pass
    coords: dict = dict(existing) if not force else {}

    # Load pre-seeded coords from CSV-derived seed file (committed alongside the app).
    # These use real ZIP-code centroids so they're more accurate than city geocoding.
    # Seed entries are used as-is and skip the Algolia+Nominatim pipeline entirely.
    seed_file = Path(__file__).parent / "gc_store_coords_seed.json"
    if seed_file.exists():
        try:
            seed = json.loads(seed_file.read_text())
            loaded = 0
            for store, data in seed.items():
                if force or store not in coords:
                    coords[store] = data
                    loaded += 1
            _send(f"  Loaded {loaded} pre-seeded coords from gc_store_coords_seed.json.")
        except Exception as e:
            _send(f"  Warning: could not load seed file: {e}")

    # Collect state context as a last-ditch fallback for dead stores
    # (stores with no Algolia items — can't determine storeName from the API).
    _send("Step 2: Scraping state pages for fallback state context…")
    name_to_state: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_fetch_state_stores_with_state, st): st for st in _US_STATES}
        for future in as_completed(futures):
            try:
                for name, state in future.result():
                    if name not in name_to_state:
                        name_to_state[name] = state
            except Exception:
                pass

    # Step 3: Hit Algolia once per store to get storeName.
    # Algolia has no rate limit for this volume — ~235 req completes in ~30 sec.
    _send("Step 3: Fetching storeName from Algolia for each active store…")
    store_to_location: dict[str, str] = {}
    no_items: list[str] = []
    # Skip stores already resolved (existing file OR seed). Seed always wins —
    # it comes from real street addresses so we never need to re-geocode it.
    todo = [s for s in known_stores if s not in coords]
    for i, store in enumerate(todo, 1):
        try:
            data = fetch_page(store, 1)
            hits = data.get("results", [{}])[0].get("hits", [])
            if not hits:
                no_items.append(store)
                continue
            sn = (hits[0].get("storeName") or "").strip()
            if sn:
                store_to_location[store] = sn
            else:
                no_items.append(store)
        except Exception as e:
            _send(f"  Algolia error for {store}: {type(e).__name__}")
            no_items.append(store)
        if i % 50 == 0:
            _send(f"  [{i}/{len(todo)}] {len(store_to_location)} storeNames, {len(no_items)} no-items so far…")

    _send(f"  Got storeName for {len(store_to_location)} stores; {len(no_items)} had no items.")
    if no_items:
        sample = ", ".join(no_items[:15])
        more = f" (+{len(no_items)-15} more)" if len(no_items) > 15 else ""
        _send(f"  No-items stores (likely closed): {sample}{more}")

    # Step 4: Geocode each storeName via Nominatim (1 req/sec per ToS).
    # Fresh session with clean API headers — NOT the shared _http session which
    # carries browser-impersonation headers that Nominatim rejects. (v2.1.3 fix)
    nom_session = http.Session()
    nom_session.headers.update({
        "User-Agent": "GCTracker/2.2 (personal tool; non-commercial)",
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9",
    })

    def _nom(query: str):
        url = (
            "https://nominatim.openstreetmap.org/search"
            f"?q={http.utils.quote(query)}&format=json&limit=1&countrycodes=us"
        )
        r = nom_session.get(url, timeout=10)
        if r.status_code != 200:
            return None, f"HTTP {r.status_code}"
        data = r.json()
        return (data[0] if data else None), None

    _send(f"Step 4: Geocoding {len(store_to_location)} storeNames via Nominatim (1/sec)…")
    failed: list[str] = []
    succeeded = 0

    # Primary pass: stores with storeName from Algolia.
    loc_items = sorted(store_to_location.items())
    for i, (store, location) in enumerate(loc_items, 1):
        try:
            result, err = _nom(location)
            if not result:
                # Fallback: strip "at <venue>" suffix from storeName.
                # e.g. "Yonkers at Ridge Hill, NY" → "Yonkers, NY"
                # Handles GC stores named after the shopping center they're in.
                stripped = re.sub(r'\s+at\s+[^,]+', '', location, flags=re.IGNORECASE).strip()
                if stripped != location:
                    time.sleep(1.0)
                    result, err = _nom(stripped)
                    if result:
                        location = stripped  # record the simpler query as source
            if result:
                coords[store] = {
                    "lat": float(result["lat"]),
                    "lng": float(result["lon"]),
                    "source": location,
                }
                succeeded += 1
            else:
                failed.append(f"{store} (storeName={store_to_location[store]}{', '+err if err else ''})")
        except Exception as e:
            failed.append(f"{store} ({type(e).__name__})")

        if i % 25 == 0:
            _send(f"  [{i}/{len(loc_items)}] {succeeded} geocoded so far…")
            STORE_COORDS_FILE.write_text(json.dumps(coords, indent=2))

        time.sleep(1.0)

    # Last-ditch pass: no-items stores, try "{store}, {state}" with state context.
    # These are likely closed, but we give them one shot so the coords file is complete.
    dead_attempted = 0
    dead_ok = 0
    for store in no_items:
        state = name_to_state.get(store, "")
        if not state:
            continue  # nothing we can do without state
        dead_attempted += 1
        query = f"{store}, {state}"
        try:
            result, _err = _nom(query)
            if result:
                coords[store] = {
                    "lat": float(result["lat"]),
                    "lng": float(result["lon"]),
                    "source": f"fallback-no-items: {query}",
                }
                dead_ok += 1
        except Exception:
            pass
        time.sleep(1.0)

    if dead_attempted:
        _send(f"  Last-ditch no-items pass: {dead_ok}/{dead_attempted} resolved via state context.")

    STORE_COORDS_FILE.write_text(json.dumps(coords, indent=2))
    skipped = total - len(todo)
    _send(f"\n✓ Done — {succeeded + dead_ok} newly geocoded, {skipped} skipped (cached), "
          f"{len(no_items)} no-items ({dead_ok} recovered), {len(failed)} failed. "
          f"Total coords: {len(coords)}/{total}.")
    if failed:
        _send(f"  Failed: {', '.join(failed[:20])}{'…' if len(failed)>20 else ''}")
    return coords


def _extract_stores_from_used_page(html: str) -> list[str]:
    """Extract all store names from GC's used inventory page filter facets.
    The __NEXT_DATA__ blob or page HTML contains the complete list of valid store
    names exactly as the filters=stores: parameter expects them."""
    stores = []

    # Strategy 1: __NEXT_DATA__ — find facet values for the 'stores' facet
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            nd = json.loads(m.group(1))
            # Walk looking for arrays of facet values near a 'stores' key
            def find_store_facets(obj, depth=0):
                if depth > 12: return
                if isinstance(obj, dict):
                    # Look for facet arrays keyed by 'stores' or containing store-like values
                    for k, v in obj.items():
                        if k.lower() in ('stores', 'store') and isinstance(v, list):
                            for item in v:
                                if isinstance(item, dict):
                                    val = item.get('displayValue') or item.get('value') or item.get('name') or ''
                                elif isinstance(item, str):
                                    val = item
                                else:
                                    continue
                                if isinstance(val, str) and 2 < len(val) < 60:
                                    stores.append(val)
                        elif isinstance(v, (dict, list)):
                            find_store_facets(v, depth + 1)
                elif isinstance(obj, list):
                    for item in obj:
                        find_store_facets(item, depth + 1)
            find_store_facets(nd)
        except Exception:
            pass

    # Strategy 2: look for displayValue patterns near "stores" in the raw JSON
    if len(stores) < 10:
        # Find JSON arrays that look like store facets
        for m2 in re.finditer(r'"(?:stores|store)"[^[]*(\[[^\]]{100,}\])', html, re.DOTALL):
            try:
                arr = json.loads(m2.group(1))
                for item in arr:
                    if isinstance(item, dict):
                        val = item.get('displayValue') or item.get('value') or ''
                        if isinstance(val, str) and 2 < len(val) < 60:
                            stores.append(val)
            except Exception:
                pass

    return stores


def refresh_store_list(send_progress=None) -> list[str]:
    """Fetch authoritative store list from GC's used inventory page filter facets,
    then fall back to state-by-state scraping. Removes blocklisted stores."""
    live_names = []

    # Strategy 1: fetch GC's used inventory page and extract store names from filter facets
    # This is the gold standard — these are the exact names the filter system accepts
    try:
        r = _http.get("https://www.guitarcenter.com/Used/", timeout=20)
        if r.status_code == 200:
            live_names = _extract_stores_from_used_page(r.text)
    except Exception:
        pass

    # Strategy 2: scrape stores.guitarcenter.com state by state in parallel
    if len(live_names) < 50:
        try:
            with ThreadPoolExecutor(max_workers=10) as pool:
                futures = {pool.submit(_fetch_state_stores, st): st for st in _US_STATES}
                for future in as_completed(futures):
                    try:
                        live_names.extend(future.result())
                    except Exception:
                        pass
        except Exception:
            pass

    # Strategy 3: main stores page URL pattern
    if len(live_names) < 20:
        try:
            r = _http.get("https://www.guitarcenter.com/Stores/", timeout=15)
            r.raise_for_status()
            html = r.text
            for slug in re.findall(r'href="https?://stores\.guitarcenter\.com/([a-z]{2})/([a-z][a-z0-9\-]+)/(\d+)"', html):
                _, city_slug, _ = slug
                name = " ".join(w.capitalize() for w in city_slug.split("-"))
                live_names.append(name)
        except Exception:
            pass

    # Strip nav garbage
    _NAV_GARBAGE = {
        "find your local guitar center store", "my account", "sign in", "track order",
        "returns", "faqs", "store locator", "guitar center lessons", "guitar center",
        "home", "shop all", "new arrivals", "top sellers", "on sale", "price drop",
        "used", "vintage", "sell your gear", "financing", "outlet", "deals",
        "daily pick", "gc pro", "lessons", "repairs", "rentals", "riffs blog",
        "accessibility statement", "privacy policy", "terms of use", "site map",
        "careers", "about", "contact us", "press room", "service", "support",
        "all rights reserved", "california transparency", "do not sell",
    }
    live_names = [
        n for n in live_names
        if n.strip().lower() not in _NAV_GARBAGE
        and len(n.strip()) >= 3
        and not any(bad in n.lower() for bad in ("guitar center", "my account", "sign in",
                                                   "track order", "©", "all rights"))
    ]

    blocklist = _get_blocklist()
    merged = sorted(set(live_names) - blocklist)
    STORES_CACHE.write_text(json.dumps({
        "stores":      merged,
        "live_count":  len(set(live_names)),
        "updated":     datetime.now().isoformat(),
    }))
    return merged


def get_store_info() -> dict:
    """Return metadata about the store list (count, last updated, live vs fallback)."""
    if STORES_CACHE.exists():
        try:
            d = json.loads(STORES_CACHE.read_text())
            return {
                "count":      len(d.get("stores", [])),
                "live_count": d.get("live_count", 0),
                "updated":    d.get("updated", ""),
            }
        except Exception:
            pass
    return {"count": len(FALLBACK_STORES), "live_count": 0, "updated": ""}


# ── Favorites ─────────────────────────────────────────────────────────────────
# (load_favorites/save_favorites and load_keywords/save_keywords removed in v2.14.5
#  along with their dead API routes — 2026-07 audit E3. The FAVORITES_FILE /
#  KEYWORDS_FILE constants stay: admin export/import/reset reference the paths.)

def load_watchlist() -> dict:
    """Returns {sku: {name, price, store, url, condition, category, date_added, sold}}"""
    if WATCHLIST_FILE.exists():
        try:
            return json.loads(WATCHLIST_FILE.read_text())
        except Exception:
            pass
    return {}


def save_watchlist(wl: dict):
    WATCHLIST_FILE.write_text(json.dumps(wl, indent=2))


# ── GC scraping ───────────────────────────────────────────────────────────────

PAGE_SIZE = 240

def _fmt_date(d: str) -> str:
    """Convert YYYY-MM-DD to M/D/YY."""
    try:
        from datetime import date
        dt = date.fromisoformat(d[:10])
        return f"{dt.month}/{dt.day}/{str(dt.year)[2:]}"
    except Exception:
        return d



def _clean_name(name: str) -> str:
    """Strip redundant 'Used ' prefix from item names."""
    name = name.strip()
    if name.lower().startswith("used "):
        name = name[5:].strip()
    return name


ALGOLIA_APP_ID  = os.environ.get("ALGOLIA_APP_ID", "")
ALGOLIA_API_KEY = os.environ.get("ALGOLIA_API_KEY", "")
ALGOLIA_INDEX   = "cD-guitarcenter"
ALGOLIA_URL     = f"https://{ALGOLIA_APP_ID.lower()}-dsn.algolia.net/1/indexes/*/queries"
ALGOLIA_HEADERS = {
    "x-algolia-application-id": ALGOLIA_APP_ID,
    "x-algolia-api-key":        ALGOLIA_API_KEY,
    "Content-Type":             "application/json",
}

def fetch_page(store_name: str = None, page: int = 1) -> dict:
    """Fetch one page of used inventory via Algolia API.
    If store_name is provided, filters to that store.
    If store_name is None, fetches ALL used inventory nationwide."""
    import time as _time
    ts = int(_time.time())
    facet_filters = [
        "categoryPageIds:Used",
        "condition.lvl0:Used",
    ]
    if store_name:
        facet_filters.append([f"stores:{store_name}"])
    payload = {"requests": [{
        "indexName":     ALGOLIA_INDEX,
        "analyticsTags": ["Did Not Search"],
        "facetFilters":  facet_filters,
        "facets":        ["*"],
        "hitsPerPage":   240,
        "maxValuesPerFacet": 10,
        "numericFilters": [f"startDate<={ts}"],
        "page":          page - 1,
        "query":         "",
        "ruleContexts":  ["used-page", "primary_itemtype", "extension_itemtype"],
        "attributesToRetrieve": ["*"],
    }]}
    r = _http.post(ALGOLIA_URL, headers=ALGOLIA_HEADERS, json=payload, timeout=20)
    r.raise_for_status()
    return r.json()


# ── New Deals helpers ──────────────────────────────────────────────────────────
_new_deals_cache: dict | None = None

_SOFTWARE_KEYWORDS = {
    "software", "plug-in", "plug in", "plugin", "virtual instrument",
    "digital download", "pro audio software", "ilok", "(download)",
    "sample pack", "sample library", "expansion pack", "loop library",
}

def _is_software_item(name: str, category: str) -> bool:
    """Return True if the item appears to be software/a plugin (by name or category)."""
    text = ((name or "") + " " + (category or "")).lower()
    return any(kw in text for kw in _SOFTWARE_KEYWORDS)

def _fetch_new_page(page: int):
    """Fetch one page of new GC inventory from Algolia. Returns (hits, nb_pages)."""
    import time as _time
    ts = int(_time.time())
    payload = {"requests": [{
        "indexName":     ALGOLIA_INDEX,
        "analyticsTags": ["Did Not Search"],
        "facetFilters":  ["condition.lvl0:New"],
        "facets":        ["*"],
        "hitsPerPage":   240,
        "numericFilters": [f"startDate<={ts}"],
        "page":          page,
    }]}
    r = _http.post(ALGOLIA_URL, headers=ALGOLIA_HEADERS, json=payload, timeout=30)
    r.raise_for_status()
    res = r.json()["results"][0]
    return res["hits"], res.get("nbPages", 1)

def _load_new_deals_cache() -> dict | None:
    global _new_deals_cache
    if _new_deals_cache is not None:
        return _new_deals_cache
    if NEW_DEALS_CACHE_FILE.exists():
        try:
            _new_deals_cache = json.loads(NEW_DEALS_CACHE_FILE.read_text())
            return _new_deals_cache
        except Exception:
            pass
    return None

def _save_new_deals_cache(items: dict, last_updated: str):
    global _new_deals_cache
    data = {"last_updated": last_updated, "items": items}
    # Atomic write (temp + os.replace) — same truncation-on-crash class the category
    # cache had before v2.13.1; a redeploy mid-write left a corrupt file. (2026-07 audit E6)
    tmp = NEW_DEALS_CACHE_FILE.parent / (NEW_DEALS_CACHE_FILE.name + ".tmp")
    tmp.write_text(json.dumps(data, separators=(',', ':')))
    os.replace(tmp, NEW_DEALS_CACHE_FILE)
    _new_deals_cache = data



_CONDITION_MAP = {
    "new":          "New",
    "likenew":      "Like New",
    "excellent":    "Excellent",
    "great":        "Great",
    "verygood":     "Very Good",
    "good":         "Good",
    "fair":         "Fair",
    "poor":         "Poor",
    "usedcondition":"Used",
    "refurbished":  "Refurbished",
    "blemished":    "Blemished",
}

def _parse_condition(raw: str) -> str:
    """Normalise a schema.org itemCondition URL or plain text to a readable label."""
    if not raw:
        return ""
    # Strip schema.org URL prefix, e.g. "https://schema.org/GoodCondition" → "GoodCondition"
    key = raw.split("/")[-1].lower().replace("condition", "").replace(" ", "").replace("-", "")
    return _CONDITION_MAP.get(key, raw.split("/")[-1])  # fall back to raw tail if unknown




# ── Condition & Details note extraction (v2.16.6) ────────────────────────────
# GC's `longDescription` field (already pulled on every scan via
# attributesToRetrieve:["*"] — zero extra Algolia calls) sometimes ends with a
# staff-written freeform note, HTML-stripped-to-plaintext by Algolia's own
# indexing: "...marketing paragraph...  Condition &amp; Details   Includes
# Hardshell Case". Confirmed present on real live items (probe_condition_
# details.py, gitignored); confirmed NOT always present (many items have no
# such section) and NOT always about accessories (one sampled item's note was
# about cosmetic scuffs and country of manufacture instead). So: extract just
# the short suffix and store it — do NOT store the full longDescription in the
# cache (it's a full marketing paragraph per item; at ~90-100K items that's
# real bytes for text nobody would read twice).
_COND_DETAILS_RE = re.compile(r'condition\s*&(?:amp;)?\s*details\s*', re.IGNORECASE)

def _extract_condition_note(long_description: str) -> str:
    """Return the freeform staff note after a 'Condition & Details' marker in
    longDescription, or '' if that section isn't present. Collapses the extra
    whitespace left behind where Algolia stripped the source HTML tags."""
    if not long_description:
        return ""
    m = _COND_DETAILS_RE.search(long_description)
    if not m:
        return ""
    note = long_description[m.end():].strip()
    note = re.sub(r'\s+', ' ', note)
    return note[:300]  # sound cap — these are short staff notes, not essays

def parse_products(data, store_name: str = None) -> list[dict]:
    """Parse products from Algolia API response. store_name can be None for all-stores queries."""
    if isinstance(data, dict):
        products = []
        try:
            results = data.get("results", [])
            if not results:
                return []
            hits = results[0].get("hits", [])
            for hit in hits:
                sku   = str(hit.get("objectID") or "").strip()
                name  = _clean_name(hit.get("displayName") or hit.get("name") or "")
                if not sku or not name:
                    continue
                price_raw = hit.get("price") or 0
                list_price_raw = hit.get("listPrice") or 0
                condition_note = _extract_condition_note(hit.get("longDescription") or "")
                # Fall back to listPrice if price is absent (listPrice is the original/regular price)
                if not price_raw and list_price_raw:
                    price_raw = list_price_raw
                try:    price = float(price_raw) if price_raw else None
                except: price = None
                try:    list_price = float(list_price_raw) if list_price_raw else 0.0
                except: list_price = 0.0
                has_price_drop = bool(hit.get("priceDrop", False))
                seo_url = hit.get("seoUrl") or ""
                url = ("https://www.guitarcenter.com" + seo_url) if seo_url else ""
                # Brand
                brand = hit.get("brand") or ""
                # Condition: "Used > Great" → "Great"
                condition = hit.get("condition") or {}
                if isinstance(condition, dict):
                    lvl1 = condition.get("lvl1") or condition.get("lvl0") or ""
                    condition = lvl1.split(">")[-1].strip() if ">" in lvl1 else lvl1
                elif isinstance(condition, str):
                    condition = condition.split(">")[-1].strip()
                condition = _parse_condition(condition) if condition else ""
                # Category from categories array: [{lvl0: "Guitars", lvl1: "Guitars > Electric Guitars", ...}]
                cats = hit.get("categories") or []
                cats_slug = hit.get("categoriesSlug") or {}
                if cats and isinstance(cats, list) and isinstance(cats[0], dict):
                    category    = cats[0].get("lvl0") or ""
                    subcategory = cats_slug.get("lvl1") or ""
                    # Fallback: parse lvl1 from full hierarchy if slug not available
                    if not subcategory:
                        lvl1_full = cats[0].get("lvl1") or ""
                        subcategory = lvl1_full.split(">")[-1].strip() if ">" in lvl1_full else ""
                else:
                    category, subcategory = "", ""
                # Date listed from startDate (seconds timestamp) — when
                # the item was published to the storefront.  Falls back to
                # creationDate (milliseconds) if startDate is missing.
                start_ts   = hit.get("startDate") or 0
                creation_ts = hit.get("creationDate") or 0
                try:
                    if start_ts:
                        date_str = datetime.utcfromtimestamp(float(start_ts)).strftime("%Y-%m-%dT%H:%M:%SZ")
                    elif creation_ts:
                        date_str = datetime.utcfromtimestamp(float(creation_ts) / 1000).strftime("%Y-%m-%dT%H:%M:%SZ")
                    else:
                        date_str = ""
                except Exception:
                    date_str = ""
                # Location: storeName gives "Austin, TX" format
                location = hit.get("storeName") or store_name or ""
                # Store: from hit's stores array when querying all stores
                hit_stores = hit.get("stores") or []
                store = store_name or (hit_stores[0] if hit_stores else "")
                # Image ID for thumbnail hover
                image_id = hit.get("imageId") or ""
                # GC's own vintage classification (premiumGear == "Vintage") — the
                # signal behind the public /Vintage dept. List-tolerant just in case.
                _pg = hit.get("premiumGear")
                is_vintage = (_pg == "Vintage") or (isinstance(_pg, list) and "Vintage" in _pg)
                products.append({
                    "id":             sku,
                    "name":           name,
                    "brand":          brand,
                    "price":          price,
                    "list_price":     list_price,
                    "has_price_drop": has_price_drop,
                    "store":          store,
                    "location":       location,
                    "url":            url,
                    "condition":      condition,
                    "category":       category,
                    "subcategory":    subcategory,
                    "date_listed":    date_str,
                    "image_id":       image_id,
                    "is_vintage":     is_vintage,
                    "condition_note": condition_note,
                })
        except Exception:
            pass
        return products
    return []


def _clean_gc_cat(s: str) -> str:
    """Strip 'Used ' prefix from GC category breadcrumb names."""
    s = s.strip()
    if s.lower().startswith("used "):
        s = s[5:].strip()
    return s


def _find_breadcrumbs_in_json(data, depth: int = 0):
    """Recursively search for a breadcrumb array inside __NEXT_DATA__ JSON."""
    if depth > 8:
        return None
    if isinstance(data, dict):
        for key in ("breadcrumbs", "breadcrumb", "breadCrumbs", "Breadcrumbs",
                    "crumbs", "navCrumbs", "categoryPath", "categories"):
            val = data.get(key)
            if isinstance(val, list) and len(val) >= 2:
                name_keys = ("name", "displayName", "label", "text", "title")
                if all(isinstance(v, dict) and any(k in v for k in name_keys) for v in val):
                    return val
        for v in data.values():
            if isinstance(v, (dict, list)):
                result = _find_breadcrumbs_in_json(v, depth + 1)
                if result:
                    return result
    elif isinstance(data, list):
        for item in data:
            if isinstance(item, (dict, list)):
                result = _find_breadcrumbs_in_json(item, depth + 1)
                if result:
                    return result
    return None


def _extract_condition_from_html(html: str) -> str:
    """Extract condition label from a GC page (listing or product page).
    Prioritises the visible 'Condition: X' text that appears on both page types."""

    _VALID = {"new", "like new", "excellent", "great", "very good", "good", "fair", "poor",
              "blemished", "refurbished", "used"}

    # Strategy A: plain visible text — "Condition: Good" / "Condition: Very Good"
    # This is the most reliable; it's what the shopper sees on the page.
    m = re.search(r'[Cc]ondition\s*[:\-–]\s*([A-Za-z][A-Za-z\s]{1,20}?)(?:\s*[<\n\r,]|$)', html)
    if m:
        val = m.group(1).strip().rstrip(".,;")
        if val.lower() in _VALID:
            return val.title()

    # Strategy B: JSON-LD itemCondition on a Product page
    for block in re.findall(
        r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>', html, re.DOTALL
    ):
        try:
            d = json.loads(block)
            offers = None
            if d.get("@type") == "Product":
                offers = d.get("offers", {})
            elif d.get("@type") == "CollectionPage":
                items = d.get("mainEntity", {}).get("itemListElement", [])
                if items:
                    offers = items[0].get("item", {}).get("offers", {})
            if offers:
                raw = offers.get("itemCondition", "")
                if raw:
                    parsed = _parse_condition(raw)
                    if parsed.lower() not in ("used", "usedcondition") and parsed:
                        return parsed
        except Exception:
            pass

    # Strategy C: __NEXT_DATA__ JSON — look for condition keys
    m2 = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m2:
        try:
            nd = json.loads(m2.group(1))
            cond = _find_key_in_json(nd, ("conditionDisplayName", "usedCondition",
                                          "productCondition", "itemCondition", "condition"))
            if cond and str(cond).lower() in _VALID:
                return str(cond).strip().title()
        except Exception:
            pass

    # Strategy D: data attributes / inline JSON strings
    for pat in [
        r'data-condition="([^"]+)"',
        r'"conditionDisplayName"\s*:\s*"([^"]+)"',
        r'"usedCondition"\s*:\s*"([^"]+)"',
    ]:
        m3 = re.search(pat, html)
        if m3:
            val = m3.group(1).strip()
            if val.lower() in _VALID:
                return val.title()

    return ""


def _find_key_in_json(data, keys: tuple, depth: int = 0):
    """Recursively search a JSON structure for any of the given keys."""
    if depth > 10:
        return None
    if isinstance(data, dict):
        for k in keys:
            if k in data and isinstance(data[k], str) and data[k]:
                return data[k]
        for v in data.values():
            result = _find_key_in_json(v, keys, depth + 1)
            if result:
                return result
    elif isinstance(data, list):
        for item in data:
            result = _find_key_in_json(item, keys, depth + 1)
            if result:
                return result
    return None


def fetch_page_data(url: str, name: str) -> tuple[str, str, str]:
    """Fetch (category, subcategory, condition) from a GC product page URL.
    Tries JSON-LD BreadcrumbList first, then __NEXT_DATA__, then keyword fallback."""
    try:
        r = _http.get(url, timeout=15)
        if r.status_code != 200:
            cat, subcat = classify_by_name(name)
            return cat, subcat, ""
        html = r.text

        condition = _extract_condition_from_html(html)

        # Strategy 1: JSON-LD BreadcrumbList
        for block in re.findall(
            r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>',
            html, re.DOTALL
        ):
            try:
                d = json.loads(block)
                if d.get("@type") == "BreadcrumbList":
                    els = sorted(d.get("itemListElement", []),
                                 key=lambda x: x.get("position", 0))
                    names = []
                    for el in els:
                        n = ((el.get("item") or {}).get("name") or el.get("name") or "").strip()
                        if n and n.lower() not in ("home", "used & vintage", "used"):
                            names.append(_clean_gc_cat(n))
                    if names:
                        cat    = names[0]
                        subcat = names[2] if len(names) >= 3 else (names[1] if len(names) >= 2 else "")
                        return cat, subcat, condition
            except Exception:
                pass

        # Strategy 2: __NEXT_DATA__ JSON blob (Next.js server-side props)
        m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
        if m:
            try:
                nd = json.loads(m.group(1))
                crumbs = _find_breadcrumbs_in_json(nd)
                if crumbs:
                    names = []
                    for c in crumbs:
                        n = (c.get("displayName") or c.get("name") or
                             c.get("label") or c.get("text") or "").strip()
                        if n and n.lower() not in ("home", "used & vintage", "used"):
                            names.append(_clean_gc_cat(n))
                    if names:
                        cat    = names[0]
                        subcat = names[2] if len(names) >= 3 else (names[1] if len(names) >= 2 else "")
                        return cat, subcat, condition
            except Exception:
                pass

        cat, subcat = classify_by_name(name)
        return cat, subcat, condition

    except Exception:
        pass

    # Final fallback: keyword classification
    cat, subcat = classify_by_name(name)
    return cat, subcat, ""


# Keep old name as alias for any callers
def classify_by_name(name: str) -> tuple[str, str]:
    """Infer category and subcategory from product name using keyword matching.
    Returns (category, subcategory). Fast — no HTTP requests required."""
    n = name.lower()

    # ── Wireless Systems (before mic/recording so 'wireless' routes here) ────
    if re.search(r'wireless system|wireless mic|wireless guitar|wireless transmitter'
                 r'|in.ear wireless|iem wireless|\bqlxd\b|\bulgx\b|\bglxd\b|\bgldx\b'
                 r'|\bslxd\b|\bpgxd\b|\bbgxd\b|\bew\d|\batwr\b', n):
        return ("Microphones & Wireless", "Wireless Systems")

    # ── Amplifiers & Cabinets — check BEFORE guitars ──────────────────────────
    # "Guitar Combo Amp", "Guitar Cabinet", "Guitar Amp Head" all contain 'guitar'
    # so we must catch amp-type gear first.
    _amp_kw = re.search(
        r'combo amp|amp combo|amp head|guitar amp|tube amp|solid.state amp|valve amp'
        r'|practice amp|\bcabinet\b|\bcab\b|speaker cab|speaker cabinet'
        r'|\d+\s*[wW]\s*(combo|head|amp)\b|(combo|head)\s*\d+\s*[wW]'
        r'|\bx\d+\b.*amp|\bamp\b.*\bhead\b', n)
    if _amp_kw:
        if re.search(r'\bbass\b', n) and not re.search(r'drum|snare|cymbal', n):
            return ("Amplifiers & Effects", "Bass Amplifiers")
        if re.search(r'keyboard|piano', n):
            return ("Amplifiers & Effects", "Keyboard Amplifiers")
        if re.search(r'acoustic', n):
            return ("Amplifiers & Effects", "Acoustic Amplifiers")
        return ("Amplifiers & Effects", "Guitar Amplifiers")

    # ── Powered Monitors / Studio Monitors / PA Speakers ─────────────────────
    if re.search(r'powered monitor|studio monitor|reference monitor|nearfield|'
                 r'pair.*monitor|monitor.*pair|powered speaker|pa speaker|'
                 r'\blp-\d|kali audio|yamaha hs\d|adam a\d|krk\b|rokit\b|'
                 r'genelec|focal alpha|jbl.*(lsr|305|306|308|310|series3)', n):
        if re.search(r'studio|reference|nearfield|kali|krk|rokit|genelec|focal|adam\b', n):
            return ("Recording", "Studio Monitors")
        return ("Live Sound", "PA Speakers")

    # ── Bass (before guitar) ──────────────────────────────────────────────────
    if re.search(r'\bbass\b', n) and not re.search(r'drum|cymbal|hi.hat|snare|bassoon', n):
        if re.search(r'acoustic|upright|stand.?up|arco|double bass', n):
            return ("Bass", "Acoustic Bass Guitars")
        if re.search(r'amp|amplifier|cabinet|combo|head\b|cab\b', n):
            return ("Amplifiers & Effects", "Bass Amplifiers")
        if re.search(r'pedal|effect|pre.?amp|di\b|direct box', n):
            return ("Amplifiers & Effects", "Bass Effects")
        return ("Bass", "Electric Bass Guitars")

    # ── Guitars ───────────────────────────────────────────────────────────────
    guitar_kw = (r'guitar|stratocaster|strat\b|telecaster|tele\b|les paul|sg\b'
                 r'|flying.?v|explorer\b|jazzmaster|jaguar\b|mustang\b'
                 r'|semi.hollow|hollow.body|archtop|resonator|dobro'
                 r'|banjo|mandolin|ukulele|squier|epiphone|prs\b|gretsch'
                 r'|rickenbacker|es.?[0-9]')
    if re.search(guitar_kw, n):
        if re.search(r'banjo', n):
            return ("Folk & Traditional Instruments", "Banjos")
        if re.search(r'mandolin', n):
            return ("Folk & Traditional Instruments", "Mandolins")
        if re.search(r'ukulele', n):
            return ("Folk & Traditional Instruments", "Ukuleles")
        if re.search(r'acoustic|classical|nylon|parlor|dreadnought|folk|fingerstyle|12.string', n):
            return ("Guitars", "Acoustic Guitars")
        if re.search(r'classical|nylon|spanish', n):
            return ("Guitars", "Classical & Nylon Guitars")
        return ("Guitars", "Electric Guitars")

    # ── Effects & Pedals ──────────────────────────────────────────────────────
    if re.search(r'pedal|effect\b|reverb\b|delay\b|distortion|overdrive|fuzz\b|wah\b'
                 r'|chorus\b|flanger|phaser|compressor|tremolo|boost\b|looper|tuner\b'
                 r'|pedalboard|multi.effect|octave\b|harmonizer|pitch shift', n):
        return ("Amplifiers & Effects", "Effects Pedals & Processors")

    # ── Amplifiers (broader — standalone \bamp\b not caught above) ────────────
    if re.search(r'\bamp\b|amplifier', n):
        if re.search(r'\bbass\b', n):
            return ("Amplifiers & Effects", "Bass Amplifiers")
        if re.search(r'keyboard|piano', n):
            return ("Amplifiers & Effects", "Keyboard Amplifiers")
        return ("Amplifiers & Effects", "Guitar Amplifiers")

    # ── Drums & Percussion ────────────────────────────────────────────────────
    if re.search(r'drum|snare|cymbal|hi.?hat|bass drum|\btom\b|drum kit|drum set'
                 r'|drum throne|djembe|cajon|bongo|conga|percussion|cowbell'
                 r'|tambourine|marimba|xylophone|vibraphone|timpani|electronic drum'
                 r'|volca beats|volca drum|tr.?\d{2,3}|drum machine|beat.*machine', n):
        if re.search(r'electronic|digital|e.?drum|drum machine|volca|tr.?\d', n):
            return ("Drums & Percussion", "Electronic Drums")
        if re.search(r'cymbal|hi.?hat', n):
            return ("Drums & Percussion", "Cymbals")
        if re.search(r'snare', n):
            return ("Drums & Percussion", "Snare Drums")
        if re.search(r'djembe|bongo|conga|cajon|hand drum', n):
            return ("Drums & Percussion", "Hand Drums")
        return ("Drums & Percussion", "Drum Sets")

    # ── Keyboards & MIDI ──────────────────────────────────────────────────────
    if re.search(r'keyboard|piano|organ\b|synth|synthesizer|workstation\b'
                 r'|midi controller|electric piano|stage piano|arranger|clav'
                 r'|wurlitzer|rhodes\b|nord\b|sound module|volca\b|groovebox'
                 r'|roland\b.*\b(jd|juno|jupiter|fa|rd|fp|gaia)'
                 r'|korg\b|yamaha\b.*\b(psr|cp|ck|np|p-\d|montage|motif)', n):
        if re.search(r'midi|controller\b', n):
            return ("Keyboards & MIDI", "MIDI Controllers")
        if re.search(r'synth|synthesizer|volca|groovebox|sound module', n):
            return ("Keyboards & MIDI", "Synthesizers & Sound Modules")
        if re.search(r'organ', n):
            return ("Keyboards & MIDI", "Organs")
        if re.search(r'digital piano|stage piano|acoustic piano', n):
            return ("Keyboards & MIDI", "Digital Pianos")
        return ("Keyboards & MIDI", "Keyboards")

    # ── Recording & Studio ────────────────────────────────────────────────────
    if re.search(r'audio interface|recording interface|usb interface|thunderbolt interface', n):
        return ("Recording", "Audio Interfaces")
    if re.search(r'microphone|condenser mic|dynamic mic|ribbon mic|vocal mic\b', n):
        return ("Recording", "Microphones")
    if re.search(r'\bmic\b', n) and not re.search(r'microphone stand', n):
        return ("Recording", "Microphones")
    if re.search(r'preamp|pre.?amplifier|channel strip|outboard', n):
        return ("Recording", "Preamps & Channel Strips")
    if re.search(r'mixer|mixing console|mixing board|analog mixer|digital mixer', n):
        return ("Recording", "Mixers")
    if re.search(r'headphone|headset|earphone|in.ear monitor|iem\b', n):
        return ("Recording", "Headphones & Monitoring")
    if re.search(r'audio recorder|field recorder|multitrack|interface\b', n):
        return ("Recording", "Audio Interfaces")

    # ── DJ Equipment ─────────────────────────────────────────────────────────
    if re.search(r'\bdj\b|turntable|cdj\b|serato|traktor|rekordbox|dj mixer|dj controller', n):
        return ("DJ Equipment & Lighting", "DJ Equipment")

    # ── Live Sound ────────────────────────────────────────────────────────────
    if re.search(r'\bpa\b|powered speaker|live sound|subwoofer|stage monitor'
                 r'|line array|public address', n):
        return ("Live Sound", "PA Systems")

    # ── Accessories ───────────────────────────────────────────────────────────
    if re.search(r'\bstrap\b|guitar strap|instrument strap', n):
        return ("Accessories", "Straps")
    if re.search(r'\bstring\b|guitar string|bass string', n):
        return ("Accessories", "Strings")
    if re.search(r'\bcase\b|gig bag|hardshell|soft case', n):
        return ("Accessories", "Cases & Bags")
    if re.search(r'\bstand\b|guitar stand|amp stand|keyboard stand', n):
        return ("Accessories", "Stands & Racks")
    if re.search(r'\bcable\b|instrument cable|patch cable|speaker cable', n):
        return ("Accessories", "Cables")
    if re.search(r'\bpick\b|plectrum', n):
        return ("Accessories", "Picks")

    return ("", "")


def scrape_store(store_name: str, send, stop_event: threading.Event) -> tuple[list[dict], set, bool]:
    """Returns (all_products_found, ids_seen_this_store, complete).

    `complete` is False when this store's fetch was cut short by an error or a
    user-initiated stop, meaning what we got back is a PARTIAL, possibly-stale
    view of this store's inventory — not "this store genuinely has fewer items
    than before." Callers use this to avoid two mistakes: (1) marking this
    store's previously-cached items as sold just because they didn't show up
    in an incomplete fetch, and (2) letting this run's max date_listed (which
    may be missing whatever this store's freshest listings are) advance the
    user's NEW-detection anchor past items we never actually got a chance to
    see. A natural "ran out of pages" or "store genuinely has 0 items right
    now" finish is still `complete = True` — a confirmed 404 (store gone) is
    also `complete = True` since that's a real, not partial, answer. (v2.16.11)
    """
    all_products, ids_seen = [], set()
    complete = True
    page = 1
    while page <= 50:
        if stop_event.is_set():
            send({"type": "progress", "msg": f"  [{store_name}] stopped."})
            complete = False
            break
        try:
            data = fetch_page(store_name, page)
        except Exception as e:
            if "404" in str(e):
                send({"type": "progress", "msg": f"  [{store_name}] not found — removing from store list."})
                _remove_invalid_store(store_name)
                # Confirmed gone is a complete, real answer — not a partial fetch.
            elif page == 1:
                # One quick retry on page 1 only — a fanout of ~298 near-simultaneous
                # per-store requests routinely trips a transient timeout/connection
                # error on a handful of stores; a single immediate retry clears most
                # of those without meaningfully slowing the scan down. (v2.16.11)
                send({"type": "progress", "msg": f"  [{store_name}] error: {e} — retrying once…"})
                time.sleep(0.75)
                try:
                    data = fetch_page(store_name, page)
                except Exception as e2:
                    send({"type": "progress", "msg": f"  [{store_name}] retry failed: {e2}"})
                    complete = False
                    break
            else:
                send({"type": "progress", "msg": f"  [{store_name}] error: {e}"})
                complete = False
                break
        products = parse_products(data, store_name)
        if not products:
            break
        if all(p["id"] in ids_seen for p in products):
            break
        for p in products:
            if p["id"] not in ids_seen:
                all_products.append(p)
                ids_seen.add(p["id"])
        # Algolia tells us total pages via nbPages
        try:
            nb_pages = data.get("results", [{}])[0].get("nbPages", 1)
            if page >= nb_pages:
                break
        except Exception:
            if len(products) < PAGE_SIZE:
                break
        page += 1
        # No sleep needed between Algolia API pages
    send({"type": "progress", "msg": f"  [{store_name}] {len(all_products)} items"})
    return all_products, ids_seen, complete


def _remove_invalid_store(store_name: str):
    """Remove a store that returned 404 from the stores cache.
    Also saves to a blocklist so it stays removed after refreshes."""
    # Remove from cache
    if STORES_CACHE.exists():
        try:
            d = json.loads(STORES_CACHE.read_text())
            stores = d.get("stores", [])
            if store_name in stores:
                stores.remove(store_name)
                d["stores"] = stores
                STORES_CACHE.write_text(json.dumps(d))
        except Exception:
            pass
    # Add to persistent blocklist
    blocklist_file = DATA_DIR / "gc_invalid_stores.json"
    try:
        blocklist = json.loads(blocklist_file.read_text()) if blocklist_file.exists() else []
        if store_name not in blocklist:
            blocklist.append(store_name)
            blocklist_file.write_text(json.dumps(sorted(blocklist)))
    except Exception:
        pass


def _get_blocklist() -> set:
    """Return the set of stores confirmed invalid (404'd)."""
    blocklist_file = DATA_DIR / "gc_invalid_stores.json"
    try:
        if blocklist_file.exists():
            return set(json.loads(blocklist_file.read_text()))
    except Exception:
        pass
    return set()


# ── State ─────────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {"last_run": None, "seen_ids": [], "item_dates": {}}



# ── Excel ─────────────────────────────────────────────────────────────────────

_COLS    = ["Status", "Date Listed", "Item Name", "Brand", "Condition", "Category", "Subcategory", "Price", "Location", "Link"]
_WIDTHS  = [8, 14, 50, 16, 14, 22, 22, 12, 18, 70]
_HDR_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
_HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_ROW_FONT = Font(name="Arial", size=10)
_NEW_FONT = Font(name="Arial", bold=True, size=10)
_ALT_FILL = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")

def _fmt_row(ws, r):
    fill = _ALT_FILL if r % 2 == 0 else None
    for col in range(1, len(_COLS) + 1):
        c = ws.cell(r, col)
        c.font = _ROW_FONT
        if fill: c.fill = fill

def write_excel(new_items: list[dict]):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    n  = len(new_items)

    # If existing file has old column count, back it up and start fresh
    if OUTPUT_FILE.exists():
        try:
            wb_check = load_workbook(OUTPUT_FILE)
            if wb_check.active.max_column != len(_COLS):
                backup = OUTPUT_FILE.with_name(OUTPUT_FILE.stem + "_backup" + OUTPUT_FILE.suffix)
                OUTPUT_FILE.rename(backup)
        except Exception:
            pass

    if OUTPUT_FILE.exists():
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        ws.insert_rows(2, amount=n)
        for i, item in enumerate(new_items):
            r = 2 + i
            date_listed = item.get("date_listed") or ""
            ws.cell(r, 1, "New"); ws.cell(r, 2, _fmt_date(date_listed) if date_listed else ts)
            ws.cell(r, 3, item["name"]); ws.cell(r, 4, item.get("brand", ""))
            ws.cell(r, 5, item.get("condition", ""))
            ws.cell(r, 6, item.get("category", "")); ws.cell(r, 7, item.get("subcategory", ""))
            pc = ws.cell(r, 8, item["price"]); pc.number_format = '$#,##0.00'
            ws.cell(r, 9, item.get("location") or item.get("store", ""))
            lc = ws.cell(r, 10, item["url"] or "")
            if item["url"]: lc.hyperlink = item["url"]; lc.style = "Hyperlink"
            _fmt_row(ws, r); ws.cell(r, 1).font = _NEW_FONT
        for r in range(2 + n, ws.max_row + 1):
            _fmt_row(ws, r)
    else:
        wb = Workbook(); ws = wb.active
        ws.title = "New Inventory"; ws.freeze_panes = "A2"
        ws.append(_COLS)
        for ci in range(1, len(_COLS) + 1):
            c = ws.cell(1, ci); c.fill = _HDR_FILL; c.font = _HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        for ci, w in enumerate(_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        for i, item in enumerate(new_items):
            r = 2 + i
            date_listed = item.get("date_listed") or ""
            ws.cell(r, 1, "New"); ws.cell(r, 2, _fmt_date(date_listed) if date_listed else ts)
            ws.cell(r, 3, item["name"]); ws.cell(r, 4, item.get("brand", ""))
            ws.cell(r, 5, item.get("condition", ""))
            ws.cell(r, 6, item.get("category", "")); ws.cell(r, 7, item.get("subcategory", ""))
            pc = ws.cell(r, 8, item["price"]); pc.number_format = '$#,##0.00'
            ws.cell(r, 9, item.get("location") or item.get("store", ""))
            lc = ws.cell(r, 10, item["url"] or "")
            if item["url"]: lc.hyperlink = item["url"]; lc.style = "Hyperlink"
            _fmt_row(ws, r); ws.cell(r, 1).font = _NEW_FONT
    wb.save(OUTPUT_FILE)


# ── Flask ─────────────────────────────────────────────────────────────────────

app             = Flask(__name__)
_secret = os.environ.get("SECRET_KEY", "").strip()
if not _secret:
    raise RuntimeError("SECRET_KEY env var is required — refusing to start with no secret")
app.secret_key  = _secret
# Secure session cookie settings
# SESSION_COOKIE_SECURE=True means the cookie is only sent over HTTPS.
# We enable it when running on Railway (RAILWAY_ENVIRONMENT is set); local dev
# is HTTP so we leave it off there to avoid breaking local testing.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"]   = os.environ.get("RAILWAY_ENVIRONMENT") is not None
# Cap request bodies (413 before parsing). /api/browse is unauthenticated and parses
# request.json in full before any per-field cap applies; without this a huge JSON body
# is a memory DoS. Largest legit payload (240-store array + 750-keyword want list +
# admin import) is well under 1 MB. (2026-07 audit M1)
app.config["MAX_CONTENT_LENGTH"] = 1 * 1024 * 1024  # 1 MB
# Static assets are safe to cache long-term because their URLs carry ?v=APP_VERSION
# (see the template replacements at the bottom of the file) — a deploy changes the
# URL, so stale caches can't survive a version bump. (2026-07 audit S7)
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 31536000  # 1 year
# Railway does NOT gzip responses (verified 2026-07-06: no Content-Encoding on
# static/gc.js) — 194KB gc.js + 53KB gc.css went over the wire uncompressed, and so
# did every /api/browse JSON page. Flask-Compress gzips html/css/js/json (default
# mimetype list; text/event-stream is untouched, so SSE is safe). (2026-07 audit S7)
try:
    from flask_compress import Compress
    # Newer Flask serves .js as text/javascript, which is NOT in flask-compress's
    # default mimetype list — without this line gc.js stays uncompressed (the whole
    # point). svg included for og-image/favicon.
    app.config["COMPRESS_MIMETYPES"] = [
        "text/html", "text/css", "text/xml",
        "text/javascript", "application/javascript",
        "application/json", "image/svg+xml",
    ]
    # Static files are streamed responses, and flask-compress's streaming algorithm
    # list excludes gzip by default (["zstd","br","deflate"]) — a gzip-only client
    # (curl -H 'Accept-Encoding: gzip', older agents) would get static uncompressed.
    app.config["COMPRESS_ALGORITHM_STREAMING"] = ["zstd", "br", "gzip", "deflate"]
    Compress(app)
except ImportError:
    print("[warn] flask-compress not installed — responses will be uncompressed")

# ProxyFix: Railway terminates TLS so Flask sees HTTP; this makes url_for() produce https://
if os.environ.get("RAILWAY_ENVIRONMENT"):
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# ── Google OAuth setup ────────────────────────────────────────────────────────
# Uses direct HTTP requests (no authlib session dependency) to avoid Railway
# proxy state-mismatch issues. State is stored server-side in _oauth_pending.
_GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID", "")
_GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
_GOOGLE_OAUTH_ENABLED = bool(_GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET)

_oauth_pending: dict = {}   # state_token → {next_url, expires}

# ── Login rate-limiting (in-memory, per IP) ────────────────────────────────────
# Keyed by IP → list of attempt timestamps. Pruned on each check.
_login_attempts: dict = {}
_LOGIN_WINDOW   = 300   # seconds (5 min rolling window)
_LOGIN_MAX      = 10    # max failed attempts before lockout

# ── Scan rate-limiting (in-memory, per IP) ─────────────────────────────────────
# Prevents unauthenticated bots from hammering /api/run and exhausting Algolia quota.
_scan_last: dict = {}   # IP → last scan start timestamp
_SCAN_COOLDOWN = 60     # seconds between scans per IP (logged-in users are exempt)

def _check_login_rate(ip: str) -> bool:
    """Return True (allowed) or False (rate-limited). Only counts failed attempts."""
    now      = time.time()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < _LOGIN_WINDOW]
    _login_attempts[ip] = attempts
    return len(attempts) < _LOGIN_MAX

def _record_login_failure(ip: str):
    now = time.time()
    bucket = [t for t in _login_attempts.get(ip, []) if now - t < _LOGIN_WINDOW]
    bucket.append(now)
    _login_attempts[ip] = bucket

def _client_ip() -> str:
    """Canonical client IP for rate-limiting and logging.
    ProxyFix(x_for=1) is applied on Railway, so request.remote_addr is
    already normalized to the real client IP — no need to read the
    X-Forwarded-For header directly (which clients can spoof)."""
    return request.remote_addr or "unknown"

def _safe_next(raw: str, default: str) -> str:
    """Validate a ?next= redirect target so it can only point back at this site.
    Rejects anything that isn't a single-slash relative path. In particular this
    blocks '//host' (protocol-relative) AND the backslash trick '/\\host', which
    browsers normalize to '//host' → an off-site open redirect. Also rejects any
    embedded CR/LF/TAB to avoid header/redirect smuggling."""
    if (not raw
            or not raw.startswith("/")
            or raw.startswith("//")
            or "\\" in raw
            or any(c in raw for c in ("\r", "\n", "\t"))):
        return default
    return raw

_q              = queue.Queue()        # legacy fallback (kept for non-run endpoints)
_run_queues: dict[str, list[queue.Queue]] = {}  # run_id → list of subscriber queues (fan-out)
_run_queues_lock = threading.Lock()
_lock           = threading.Lock()
_stop_event     = threading.Event()
_current_run_id: str   = ""   # run_id of the scan currently in progress (empty if none)
_current_run_time: str = ""   # run_time of the current scan

import uuid as _uuid

# ── Device access tracking ─────────────────────────────────────────────────────
_DEVICE_LOG       = DATA_DIR / "gc_device_log.jsonl"
_device_log_lock  = threading.Lock()
_seen_today: set  = set()   # (device_id, date) pairs already written today

def _log_device(device_id: str):
    """Append one line to gc_device_log.jsonl the first time a device is seen each day."""
    today = datetime.utcnow().strftime("%Y-%m-%d")
    key   = (device_id, today)
    if key in _seen_today:
        return
    _seen_today.add(key)
    entry = json.dumps({
        "date":       today,
        "time":       datetime.utcnow().strftime("%H:%M:%SZ"),
        "device_id":  device_id,
        "ua":         request.headers.get("User-Agent", "")[:120],
        "ip":         _client_ip(),
    })
    with _device_log_lock:
        with open(_DEVICE_LOG, "a") as f:
            f.write(entry + "\n")

# ── Old-domain redirect (301 to gcgeartracker.com) ───────────────────────────
@app.before_request
def _redirect_old_domain():
    """301-redirect any request arriving on the old hostname to gcgeartracker.com."""
    host = request.host.split(":")[0].lower()
    if host == "gctracker.animalsintrees.com":
        target = "https://gcgeartracker.com" + request.full_path.rstrip("?")
        return redirect(target, code=301)

# ── CSRF protection (Origin check on state-changing requests) ─────────────────
@app.before_request
def _csrf_check():
    """Block cross-origin POST/PUT/DELETE/PATCH requests.
    JSON APIs already get implicit protection (browsers won't send
    Content-Type: application/json cross-origin without CORS preflight),
    but this adds an explicit Origin/Referer check as defense-in-depth."""
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    origin  = request.headers.get("Origin", "")
    referer = request.headers.get("Referer", "")
    # Allow requests with no Origin (e.g. same-origin, curl, server-to-server)
    if not origin and not referer:
        return None
    host = request.host  # e.g. "gctracker.animalsintrees.com" or "localhost:5050"
    # Check Origin header first
    if origin:
        from urllib.parse import urlparse
        parsed = urlparse(origin)
        if parsed.netloc == host:
            return None
        return jsonify({"error": "Cross-origin request blocked."}), 403
    # Fallback: check Referer
    if referer:
        from urllib.parse import urlparse
        parsed = urlparse(referer)
        if parsed.netloc == host:
            return None
        return jsonify({"error": "Cross-origin request blocked."}), 403
    return None

@app.after_request
def _track_device(response):
    """Set a long-lived device cookie, log first visit of each day, and add security headers."""
    # Security headers — applied to every response
    response.headers.setdefault("X-Frame-Options",        "SAMEORIGIN")
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy",        "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy",     "camera=(), microphone=(), geolocation=()")
    # Cross-origin isolation for the top-level document. OAuth here is redirect-based
    # (no popup relies on window.opener), so 'same-origin-allow-popups' is safe and
    # earns the security-scanner credit. NOTE: we deliberately do NOT set
    # Cross-Origin-Resource-Policy globally — it would stop social crawlers (Twitter/
    # Slack/Facebook) from fetching the cross-origin OG image.
    response.headers.setdefault("Cross-Origin-Opener-Policy", "same-origin-allow-popups")
    # HSTS — only on Railway (HTTPS); tells browsers to always use HTTPS for this domain
    if os.environ.get("RAILWAY_ENVIRONMENT"):
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    # CSP — inline scripts/styles needed (single-file app), but block everything else.
    # default-src 'none' forces explicit allowlists for every resource type.
    # frame-ancestors 'none' prevents clickjacking (stronger than X-Frame-Options alone).
    response.headers.setdefault("Content-Security-Policy",
        "default-src 'none'; "
        "script-src 'self' https://accounts.google.com https://apis.google.com https://www.googletagmanager.com; "
        "style-src 'self' 'unsafe-inline' https://accounts.google.com; "
        "img-src 'self' data: https://media.guitarcenter.com https://*.googleusercontent.com; "
        "connect-src 'self' https://accounts.google.com https://oauth2.googleapis.com https://www.googleapis.com https://api.zippopotam.us https://www.google-analytics.com; "
        "frame-src https://accounts.google.com; "
        "font-src 'self'; "
        "object-src 'none'; "
        "base-uri 'self'; "
        "form-action 'self' https://accounts.google.com; "
        "frame-ancestors 'none'"
    )
    # Skip SSE streams for device tracking (cookie/logging only, headers already set above)
    if request.path.startswith("/api/progress"):
        return response
    _maybe_backup_users_db()
    device_id = request.cookies.get("gt_device_id")
    if not device_id:
        device_id = str(_uuid.uuid4())
        # 2-year cookie — survives browser restarts; Secure only on Railway (HTTPS)
        response.set_cookie("gt_device_id", device_id,
                            max_age=60*60*24*730, httponly=True, samesite="Lax",
                            secure=bool(os.environ.get("RAILWAY_ENVIRONMENT")),
                            path="/")
    _log_device(device_id)
    return response

def _create_run_queue() -> tuple[str, queue.Queue]:
    """Start a new run: create a fan-out entry and return (run_id, first_subscriber_queue)."""
    global _current_run_id
    run_id = _uuid.uuid4().hex[:12]
    q = queue.Queue()
    with _run_queues_lock:
        _run_queues[run_id] = [q]
        _current_run_id = run_id
    return run_id, q

def _subscribe_to_run(run_id: str) -> queue.Queue | None:
    """Join an in-progress run. Returns a new subscriber queue, or None if run is gone."""
    q = queue.Queue()
    with _run_queues_lock:
        if run_id not in _run_queues:
            return None
        _run_queues[run_id].append(q)
    return q

def _broadcast(run_id: str, msg):
    """Send a message to all subscriber queues for a run."""
    with _run_queues_lock:
        subscribers = list(_run_queues.get(run_id, []))
    for q in subscribers:
        q.put(msg)

def _get_run_queue(run_id: str) -> queue.Queue | None:
    """Return the first subscriber queue (legacy helper, unused by _run directly)."""
    with _run_queues_lock:
        subs = _run_queues.get(run_id)
        return subs[0] if subs else None

def _cleanup_subscriber(run_id: str, q: queue.Queue):
    """Remove one subscriber queue. If it's the last, remove the whole run."""
    global _current_run_id
    with _run_queues_lock:
        subs = _run_queues.get(run_id)
        if subs and q in subs:
            subs.remove(q)
        if not subs:
            _run_queues.pop(run_id, None)
            if _current_run_id == run_id:
                _current_run_id = ""

def _cleanup_run_queue(run_id: str):
    """Remove all subscribers for a run (called when scan finishes)."""
    global _current_run_id
    with _run_queues_lock:
        _run_queues.pop(run_id, None)
        if _current_run_id == run_id:
            _current_run_id = ""


def optional_user_context(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        # Site access is open — no login required.
        # Individual sensitive endpoints (e.g. /api/reset) enforce their own password.
        return f(*args, **kwargs)
    return decorated

# ── Admin session auth ────────────────────────────────────────────────────────
# Replaces the old ?pw= query-string pattern so the admin password never
# appears in URLs, browser history, or server/proxy logs.

_ADMIN_LOGIN_HTML = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Admin Login</title>
<style>
body{background:#111;color:#eee;font-family:monospace;display:flex;align-items:center;
     justify-content:center;height:100vh;margin:0}
.box{background:#1a1a1a;border:1px solid #2e2e2e;border-radius:10px;padding:40px;width:320px}
h2{text-align:center;margin-bottom:16px;color:#fff}
input{width:100%;padding:10px;background:#222;border:1px solid #444;color:#eee;
      border-radius:4px;margin-bottom:12px;box-sizing:border-box;font-size:1rem}
button{width:100%;padding:10px;background:#c00;color:#fff;border:none;
       border-radius:4px;cursor:pointer;font-size:1rem}
.err{color:#f88;text-align:center;margin-bottom:12px;font-size:.85rem}
</style></head><body>
<div class="box"><h2>Admin Login</h2>
<form method="POST">
  {err}
  <input type="hidden" name="_csrf" value="{csrf}">
  <input name="pw" type="password" placeholder="Admin password" autofocus>
  <button type="submit">Enter</button>
</form>
</div></body></html>"""

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    import secrets as _secrets
    if request.method == "GET":
        token = _secrets.token_hex(32)
        session["_admin_csrf"] = token
        html = _ADMIN_LOGIN_HTML.replace('{err}','').replace('{csrf}', token)
        return Response(html, content_type="text/html")
    # Validate CSRF token before anything else
    submitted = request.form.get("_csrf") or ""
    expected  = session.get("_admin_csrf") or ""
    if not expected or not hmac.compare_digest(submitted, expected):
        token = _secrets.token_hex(32)
        session["_admin_csrf"] = token
        return Response(
            _ADMIN_LOGIN_HTML.replace('{err}','<div class="err">Invalid request. Please reload and try again.</div>').replace('{csrf}', token),
            status=403, content_type="text/html")
    ip = _client_ip()
    if not _check_login_rate(ip):
        token = _secrets.token_hex(32)
        session["_admin_csrf"] = token
        return Response(
            _ADMIN_LOGIN_HTML.replace('{err}','<div class="err">Too many attempts. Please wait a few minutes.</div>').replace('{csrf}', token),
            status=429, content_type="text/html")
    pw = (request.form.get("pw") or "").strip()
    admin_pw = APP_PASSWORD
    if not admin_pw or not hmac.compare_digest(pw, admin_pw):
        _record_login_failure(ip)
        print(f"[Admin] Failed login attempt from {ip}")
        token = _secrets.token_hex(32)
        session["_admin_csrf"] = token
        return Response(
            _ADMIN_LOGIN_HTML.replace('{err}','<div class="err">Incorrect password.</div>').replace('{csrf}', token),
            status=401, content_type="text/html")
    session["admin"] = True
    session.pop("_admin_csrf", None)
    next_url = _safe_next(request.args.get("next", "/admin/users"), "/admin/users")
    return redirect(next_url)

@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin", None)
    return redirect("/admin/login")

def _is_admin() -> bool:
    """Check if the current request has admin access.
    Two paths: (1) explicit admin session from /admin/login password form,
    (2) logged-in Google user whose email matches ADMIN_EMAIL env var.

    IMPORTANT: the email check requires google_id to be set (i.e. the account
    must have authenticated via Google, not just claimed the email at registration).
    Without this guard, any user who registers with ADMIN_EMAIL as their
    self-reported email would pass the check."""
    if bool(session.get("admin")) and bool(APP_PASSWORD):
        return True
    if ADMIN_EMAIL:
        user_id = session.get("user_id")
        if user_id:
            user = _user_by_id(user_id)
            if (user
                    and user.get("google_id")   # must have authenticated via Google
                    and (user.get("email") or "").strip().lower() == ADMIN_EMAIL):
                return True
    return False

def _require_admin():
    """Return a 403 or redirect if not admin, else None.
    Normal path: log in via Google on the main app → admin footer link appears.
    Break-glass: /admin/login still exists for password-based access if Google auth breaks."""
    if _is_admin():
        return None
    # If not logged in at all, send to main app login
    if not session.get("user_id"):
        return redirect("/")
    # Logged in but not admin — show a plain 403
    return Response(
        "<!DOCTYPE html><html><head><meta charset='UTF-8'><title>403</title>"
        "<style>body{background:#111;color:#888;font-family:monospace;display:flex;"
        "align-items:center;justify-content:center;height:100vh;margin:0}"
        ".box{text-align:center}.box h1{color:#fff;font-size:1.4rem;margin-bottom:8px}"
        ".box a{color:#666;font-size:.85rem}</style></head>"
        "<body><div class='box'><h1>403 — Not authorized</h1>"
        "<a href='/'>← Back to app</a></div></body></html>",
        status=403, content_type="text/html"
    )

def _require_admin_api():
    """For POST API endpoints — return a JSON 401 if not admin, else None."""
    if _is_admin():
        return None
    return jsonify({"error": "Unauthorized"}), 401

def _admin_page_csrf() -> str:
    """Return (creating if needed) a CSRF token for admin POST actions.
    The login flow pops the token on success, so post-login pages call this
    to ensure one always exists in the session."""
    import secrets as _secrets
    if not session.get("_admin_csrf"):
        session["_admin_csrf"] = _secrets.token_hex(32)
    return session["_admin_csrf"]

def _check_admin_csrf_header() -> bool:
    """Validate the X-CSRF-Token header against the session token."""
    submitted = request.headers.get("X-CSRF-Token", "")
    expected  = session.get("_admin_csrf") or ""
    if not expected or not submitted:
        return False
    return hmac.compare_digest(submitted, expected)


# (Dead sitewide-password login page + /login + GET /logout removed in v2.14.5 —
#  session["logged_in"] was written there and read nowhere; GET /logout was also a
#  CSRF force-logout. Admin auth lives at /admin/login; user auth is /api/login +
#  Google OAuth. 2026-07 audit L3/E3.)

# ── User account API ──────────────────────────────────────────────────────────

@app.route("/api/register", methods=["POST"])
def api_register():
    ip = _client_ip()
    if not _check_login_rate(ip):
        return jsonify({"error": "Too many attempts from this device. Please wait a few minutes."}), 429
    _record_login_failure(ip)   # count every register attempt to limit spam
    data     = request.json or {}
    username = re.sub(r'[^A-Za-z0-9_\-]', '', (data.get("username") or "").strip())
    password = (data.get("password") or "").strip()
    if not username or len(username) < 3:
        return jsonify({"error": "Username must be at least 3 characters (letters, numbers, _ -)"}), 400
    if len(username) > 30:
        return jsonify({"error": "Username must be 30 characters or fewer."}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters."}), 400
    if _user_by_username(username):
        return jsonify({"error": "That username is already taken."}), 409
    email = (data.get("email") or "").strip().lower() or None
    if email and ("@" not in email or "." not in email.split("@")[-1]):
        return jsonify({"error": "Please enter a valid email address, or leave it blank."}), 400
    if email:
        with _user_db() as conn:
            if conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone():
                return jsonify({"error": "That email is already linked to another account."}), 409
    pw_hash = generate_password_hash(password)
    now     = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    with _user_db() as conn:
        cur     = conn.execute(
            "INSERT INTO users (username, email, password_hash, created_at) VALUES (?,?,?,?)",
            (username, email, pw_hash, now)
        )
        user_id = cur.lastrowid
        conn.commit()
    session.permanent = True
    session["user_id"]       = user_id
    session["user_username"] = username
    return jsonify({"status": "registered", "username": username, "google_linked": False, "data": _get_user_data(user_id)})

@app.route("/api/login", methods=["POST"])
def api_login():
    ip = _client_ip()
    if not _check_login_rate(ip):
        return jsonify({"error": "Too many login attempts. Please wait a few minutes and try again."}), 429
    data     = request.json or {}
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    user     = _user_by_username(username)
    if not user or not user.get("password_hash") or not check_password_hash(user["password_hash"], password):
        _record_login_failure(ip)
        if user and user.get("google_id") and not user.get("password_hash"):
            return jsonify({"error": "This account uses Google sign-in. Please use the 'Sign in with Google' button."}), 401
        return jsonify({"error": "Incorrect username or password."}), 401
    session.permanent = True
    session["user_id"]       = user["id"]
    session["user_username"] = username
    _touch_last_login(user["id"])
    return jsonify({"status": "ok", "username": username, "google_linked": bool(user.get("google_id")), "data": _get_user_data(user["id"])})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.pop("user_id",       None)
    session.pop("user_username", None)
    return jsonify({"status": "logged_out"})

# ── Google OAuth routes ───────────────────────────────────────────────────────

@app.route("/api/auth/google")
def auth_google():
    if not _GOOGLE_OAUTH_ENABLED:
        return "Google Sign-In is not configured.", 501
    import secrets as _sec, urllib.parse as _up
    from flask import url_for
    # Prevent open redirect — only allow same-site relative paths (also blocks the
    # '/\\host' backslash trick that browsers turn into a protocol-relative '//host').
    next_url     = _safe_next(request.args.get("next", "/"), "/")
    redirect_uri = url_for("auth_google_callback", _external=True)
    # Store state server-side — avoids session cookie issues on Railway proxy
    state = _sec.token_urlsafe(32)
    _oauth_pending[state] = {"next_url": next_url, "expires": time.time() + 600}
    # Purge expired states
    now_t = time.time()
    for k in [k for k, v in list(_oauth_pending.items()) if v["expires"] < now_t]:
        _oauth_pending.pop(k, None)
    params = {
        "client_id":     _GOOGLE_CLIENT_ID,
        "redirect_uri":  redirect_uri,
        "response_type": "code",
        "scope":         "openid email profile",
        "state":         state,
    }
    return redirect("https://accounts.google.com/o/oauth2/v2/auth?" + _up.urlencode(params))

@app.route("/api/auth/google/callback")
def auth_google_callback():
    import urllib.parse as _up, traceback as _tb
    try:
        return _auth_google_callback_inner()
    except Exception as exc:
        print(f"[Google OAuth] UNHANDLED:\n{_tb.format_exc()}")
        return redirect("/?google_error=1")

def _auth_google_callback_inner():
    if not _GOOGLE_OAUTH_ENABLED:
        return "Google Sign-In is not configured.", 501
    import urllib.parse as _up
    from flask import url_for

    state = request.args.get("state", "")
    code  = request.args.get("code",  "")

    # Validate state against server-side dict (no session needed)
    pending = _oauth_pending.pop(state, None)
    if not pending or pending["expires"] < time.time():
        return redirect("/?google_error=1")
    next_url     = pending["next_url"]
    redirect_uri = url_for("auth_google_callback", _external=True)

    try:
        # Exchange code for token directly — no authlib session dependency
        token_resp = http.post("https://oauth2.googleapis.com/token", data={
            "code":          code,
            "client_id":     _GOOGLE_CLIENT_ID,
            "client_secret": _GOOGLE_CLIENT_SECRET,
            "redirect_uri":  redirect_uri,
            "grant_type":    "authorization_code",
        }, timeout=10)
        token_data   = token_resp.json()
        access_token = token_data.get("access_token", "")
        if not access_token:
            raise ValueError(f"No access_token: {token_data}")
        # Fetch user info
        ui_resp  = http.get("https://www.googleapis.com/oauth2/v3/userinfo",
                            headers={"Authorization": f"Bearer {access_token}"}, timeout=10)
        userinfo       = ui_resp.json()
        google_id      = userinfo.get("sub", "")
        email          = (userinfo.get("email") or "").strip().lower()
        email_verified = bool(userinfo.get("email_verified", False))
        name           = (userinfo.get("name") or "").strip()
    except Exception as exc:
        print(f"[Google OAuth] token/userinfo error: {exc}")
        return redirect("/?google_error=1")

    if not google_id:
        return redirect("/?google_error=1")

    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    # 1) Already linked by google_id → just log in
    user = _user_by_google_id(google_id)
    if user:
        session.permanent        = True
        session["user_id"]       = user["id"]
        session["user_username"] = user["username"]
        _touch_last_login(user["id"])
        return redirect(next_url)

    # 2) Email matches an existing account → link & log in
    # Require email_verified to prevent account takeover via unverified Google emails
    if email and email_verified:
        user = _user_by_email(email)
        if user:
            try:
                with _user_db() as conn:
                    conn.execute("UPDATE users SET google_id=? WHERE id=?", (google_id, user["id"]))
                    conn.commit()
            except Exception as exc:
                print(f"[Google OAuth] DB link error: {exc}")
            session.permanent        = True
            session["user_id"]       = user["id"]
            session["user_username"] = user["username"]
            _touch_last_login(user["id"])
            return redirect(next_url)

    # 3) New Google user → create account
    # password_hash='' (not None) so existing DBs with NOT NULL constraint don't error
    username = _gen_google_username(name or (email.split("@")[0] if email else "user"))
    try:
        with _user_db() as conn:
            cur = conn.execute(
                "INSERT INTO users (username, email, password_hash, google_id, created_at) VALUES (?,?,?,?,?)",
                (username, email or None, "", google_id, now)
            )
            user_id = cur.lastrowid
            conn.commit()
    except Exception as exc:
        print(f"[Google OAuth] DB insert error: {exc}")
        return redirect("/?google_error=1")
    session.permanent        = True
    session["user_id"]       = user_id
    session["user_username"] = username
    _touch_last_login(user_id)
    # Flag new Google users so the frontend shows the welcome/setup modal
    sep = "&" if "?" in next_url else "?"
    return redirect(next_url + sep + "google_new=1")

@app.route("/api/auth/config")
def auth_config():
    return jsonify({"google_oauth": _GOOGLE_OAUTH_ENABLED})

@app.route("/api/me")
def api_me():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"logged_in": False})
    user = _user_by_id(user_id)
    return jsonify({
        "logged_in":    True,
        "username":     session.get("user_username", ""),
        "google_linked": bool(user and user.get("google_id")),
        "has_email":    bool(user and user.get("email")),
        "is_admin":     _is_admin(),
        "data":         _get_user_data(user_id),
    })

@app.route("/api/sync", methods=["POST"])
def api_sync():
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "Not logged in."}), 401
    data   = request.json or {}
    kwargs = {}
    for field in ("watchlist", "keywords", "favorites", "new_ids", "saved_searches"):
        if field in data:
            kwargs[field] = data[field]
    if "last_run" in data:
        kwargs["last_run"] = data["last_run"]
    if "last_anchor" in data:
        kwargs["last_anchor"] = data["last_anchor"] or ""
    if kwargs:
        _set_user_data(user_id, **kwargs)
    return jsonify({"status": "synced"})

@app.route("/api/setup-google-account", methods=["POST"])
def api_setup_google_account():
    """Set/change username after Google sign-in, optionally importing an existing account."""
    user_id = session.get("user_id")
    if not user_id:
        return jsonify({"error": "Not logged in."}), 401
    ip = _client_ip()
    if not _check_login_rate(ip):
        return jsonify({"error": "Too many attempts. Please wait a few minutes."}), 429
    data         = request.json or {}
    new_username = re.sub(r'[^A-Za-z0-9_\-]', '', (data.get("username") or "").strip())
    import_pw    = (data.get("import_password") or "").strip()
    if not new_username or len(new_username) < 3:
        return jsonify({"error": "Username must be at least 3 characters (letters, numbers, _ -)"}), 400
    if len(new_username) > 30:
        return jsonify({"error": "Username must be 30 characters or fewer."}), 400
    existing = _user_by_username(new_username)
    if existing and existing["id"] != user_id:
        # Username belongs to another account — need password to import it
        if not import_pw:
            return jsonify({"error": "taken"}), 409
        if not existing.get("password_hash") or not check_password_hash(existing["password_hash"], import_pw):
            _record_login_failure(ip)
            return jsonify({"error": "wrong_password"}), 401
        # Credentials verified — merge old account into current Google account
        old_data = _get_user_data(existing["id"])
        new_data = _get_user_data(user_id)
        merged = {
            "watchlist":      {**new_data["watchlist"], **old_data["watchlist"]},
            "keywords":       old_data["keywords"]       if old_data["keywords"]       else new_data["keywords"],
            "favorites":      list(set(new_data["favorites"] + old_data["favorites"])),
            "saved_searches": old_data["saved_searches"] if old_data["saved_searches"] else new_data["saved_searches"],
            "last_run":       old_data["last_run"]       or new_data["last_run"],
            "new_ids":        old_data["new_ids"]        if old_data["new_ids"]        else new_data["new_ids"],
            "last_anchor":    max(old_data.get("last_anchor", ""), new_data.get("last_anchor", "")),
        }
        _set_user_data(user_id, **merged)
        with _user_db() as conn:
            conn.execute("DELETE FROM user_data WHERE user_id=?", (existing["id"],))
            conn.execute("DELETE FROM users WHERE id=?", (existing["id"],))
            conn.execute("UPDATE users SET username=? WHERE id=?", (new_username, user_id))
            conn.commit()
        session["user_username"] = new_username
        return jsonify({"status": "imported", "username": new_username, "data": _get_user_data(user_id)})
    # Username available (or unchanged) — just update it
    with _user_db() as conn:
        conn.execute("UPDATE users SET username=? WHERE id=?", (new_username, user_id))
        conn.commit()
    session["user_username"] = new_username
    return jsonify({"status": "ok", "username": new_username, "data": _get_user_data(user_id)})

_ADMIN_NAV_LINKS = [
    ("/admin/users",            "👤 Users"),
    ("/admin/devices",          "📡 Devices"),
    ("/admin/listing-patterns", "📊 Listing Patterns"),
    ("/admin/build-coords",     "🗺 Build Coords"),
    ("/admin/validate-stores",  "✓ Validate Stores"),
    ("/admin/pg-backfill",      "🗄 PG Backfill"),
]

def _admin_nav(current: str) -> str:
    """Render a top nav bar for admin pages. `current` is the active path."""
    links = []
    for path, label in _ADMIN_NAV_LINKS:
        if path == current:
            links.append(f'<span style="color:#fff;font-weight:700">{label}</span>')
        else:
            links.append(f'<a href="{path}" style="color:#888;text-decoration:none">{label}</a>')
    links_html = ' &nbsp;·&nbsp; '.join(links)
    return (
        '<nav style="background:#1a1a1a;border-bottom:1px solid #2e2e2e;padding:10px 24px;'
        'margin:-24px -24px 28px -24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap">'
        f'<a href="/" style="color:#c00;font-weight:700;text-decoration:none;margin-right:8px;font-size:.85rem">← App</a>'
        f'<span style="color:#333">|</span>'
        f'<span style="font-size:.82rem">{links_html}</span>'
        '</nav>'
    )


@app.route("/admin/devices")
def admin_devices():
    """Session-protected device access summary page."""
    denied = _require_admin()
    if denied:
        return denied

    # Parse log
    entries = []
    if _DEVICE_LOG.exists():
        for line in _DEVICE_LOG.read_text().splitlines():
            line = line.strip()
            if line:
                try: entries.append(json.loads(line))
                except: pass

    # Aggregate
    from collections import defaultdict
    unique_devices  = {e["device_id"] for e in entries}
    by_device       = defaultdict(list)
    by_date         = defaultdict(set)
    for e in entries:
        by_device[e["device_id"]].append(e)
        by_date[e["date"]].add(e["device_id"])

    rows = []
    for did, evts in sorted(by_device.items(), key=lambda x: x[1][-1]["date"], reverse=True):
        last  = evts[-1]
        first = evts[0]
        ua    = last.get("ua", "")
        # Guess platform
        if "iPhone" in ua or "iPad" in ua:    platform = "📱 iOS"
        elif "Android" in ua:                  platform = "📱 Android"
        elif "Macintosh" in ua:                platform = "💻 Mac"
        elif "Windows" in ua:                  platform = "🖥 Windows"
        elif "Linux" in ua:                    platform = "🖥 Linux"
        else:                                  platform = "❓ Unknown"
        rows.append({
            "id":       did[:8] + "…",
            "platform": platform,
            "first":    first["date"],
            "last":     last["date"] + " " + last["time"],
            "days":     len(evts),
            "ip":       last.get("ip", ""),
        })

    # Daily active table
    daily = sorted(by_date.items(), reverse=True)[:30]

    html  = ['<!DOCTYPE html><html><head><meta charset="UTF-8">']
    html += ['<title>Device Log</title>']
    html += ['<style>body{background:#111;color:#ddd;font-family:monospace;padding:24px;font-size:.88rem}']
    html += ['h1{color:#fff;margin-bottom:4px}h2{color:#aaa;font-size:1rem;margin:24px 0 8px}']
    html += ['table{border-collapse:collapse;width:100%;max-width:900px}']
    html += ['th{background:#1e1e1e;padding:8px 12px;text-align:left;border-bottom:2px solid #333;color:#aaa}']
    html += ['td{padding:6px 12px;border-bottom:1px solid #222}tr:hover td{background:#1a1a1a}']
    html += ['</style></head><body>']
    html += [_admin_nav('/admin/devices')]
    html += [f'<h1>📊 Device Tracker</h1>']
    html += [f'<p style="color:#666">{len(unique_devices)} unique devices &nbsp;·&nbsp; {len(entries)} total day-visits &nbsp;·&nbsp; {len(entries) and entries[-1]["date"]} last activity</p>']

    html += ['<h2>All Devices</h2><table>']
    html += ['<tr><th>ID</th><th>Platform</th><th>First seen</th><th>Last seen</th><th>Days active</th><th>IP</th></tr>']
    for r in rows:
        html += [f'<tr><td>{_html.escape(str(r["id"]))}</td><td>{_html.escape(str(r["platform"]))}</td><td>{_html.escape(str(r["first"]))}</td>'
                 f'<td>{_html.escape(str(r["last"]))}</td><td>{int(r["days"])}</td><td>{_html.escape(str(r["ip"]))}</td></tr>']
    html += ['</table>']

    html += ['<h2>Daily Active Devices (last 30 days)</h2><table>']
    html += ['<tr><th>Date</th><th>Unique devices</th></tr>']
    for date, devs in daily:
        html += [f'<tr><td>{date}</td><td>{len(devs)}</td></tr>']
    html += ['</table></body></html>']

    return Response("".join(html), content_type="text/html")


def _admin_task_page(title: str, api_path: str, description: str,
                     options_html: str = "", nav_current: str = "") -> str:
    """Shared HTML template for long-running admin task pages (build-coords,
    validate-stores, pg-backfill).

    CSP note (v2.16.19): this template used to have its Run/SSE-progress logic as an
    inline <script> with onclick="run()" baked directly into the returned HTML. CSP's
    script-src is 'self' only (no 'unsafe-inline', no nonce) -- the browser silently
    refused to run that inline script AND the inline onclick handler, so "Run Now" did
    nothing on every page built from this template, since well before this was
    noticed. No console error appears for a blocked inline *handler* failing to
    attach (only for a blocked <script> tag's content), which is how it went
    undetected. Fixed by moving the logic to static/admin-task.js (same-origin, so
    'self' allows it) and passing config via data-api-path on #run-btn instead of
    templating raw JS -- matches the data-*-attribute + external-JS pattern every
    other onclick="..." in this app was already converted to back in v2.10.18.

    options_html: optional HTML snippet inserted above the Run button (e.g.
    checkboxes). The one existing use -- Build Coords' "force re-geocode" checkbox --
    must use id="force-cb": static/admin-task.js looks for that specific id and
    includes {force: <checked>} in the POST body if present. There's no generic
    extra-field mechanism; add one in admin-task.js if a second page ever needs a
    different field.
    Auth is handled by the caller via _require_admin().
    """
    safe_api  = api_path.replace('"', '&quot;')
    safe_title = title.replace('<', '').replace('>', '')
    safe_desc  = description.replace('<', '').replace('>', '')
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{safe_title} — GC Tracker Admin</title>
<style>
  body{{background:#111;color:#ddd;font-family:monospace;padding:40px;max-width:800px}}
  h2{{color:#eee;margin-bottom:4px}} p{{color:#888;margin-top:0}}
  button{{padding:8px 20px;background:#c00;color:#fff;border:none;border-radius:5px;
          font-size:1rem;cursor:pointer;margin-top:16px}}
  button:disabled{{background:#555;cursor:default}}
  #log{{margin-top:20px;background:#1a1a1a;border:1px solid #333;border-radius:6px;
        padding:16px;min-height:120px;white-space:pre-wrap;font-size:.82rem;line-height:1.5}}
  .done{{color:#4ade80}} .err{{color:#f88}}
</style></head><body>
{_admin_nav(nav_current)}
<h2>🛠 {safe_title}</h2>
<p>{safe_desc}</p>
{options_html}
<button id="run-btn" data-api-path="{safe_api}">▶ Run Now</button>
<div id="log">Waiting…</div>
<script src="/static/admin-task.js"></script>
</body></html>"""


@app.route("/admin/build-coords")
def admin_build_coords():
    """Admin page to geocode all stores and build gc_store_coords.json."""
    denied = _require_admin()
    if denied:
        return denied
    html = _admin_task_page(
        title="Build Store Coordinates",
        api_path="/api/build-store-coords",
        description="Pulls 'storeName' (e.g. 'South Austin, TX') from Algolia for each "
                    "active store, then geocodes that string via Nominatim (~1 req/sec). "
                    "Takes ~5 min. Skips stores already in gc_store_coords.json unless "
                    "'Force re-geocode all' is checked.",
        options_html='<label style="display:block;margin-top:14px;color:#bbb;cursor:pointer">'
                     '<input type="checkbox" id="force-cb" style="vertical-align:middle"> '
                     'Force re-geocode all stores (even cached ones)</label>',
        nav_current="/admin/build-coords",
    )
    return Response(html, content_type="text/html")


@app.route("/admin/validate-stores")
def admin_validate_stores():
    """Admin page to validate and clean up the store list."""
    denied = _require_admin()
    if denied:
        return denied
    html = _admin_task_page(
        title="Validate Stores",
        api_path="/api/validate-stores",
        description="Checks every store for 404s, auto-removes dead stores, "
                    "renames any whose slugs changed, then rebuilds the store list from GC live data. "
                    "Takes ~0.5s per store.",
        nav_current="/admin/validate-stores",
    )
    return Response(html, content_type="text/html")


@app.route("/admin/pg-backfill")
def admin_pg_backfill():
    """Admin page to run the one-time full Postgres catalog backfill from the
    live _cat_cache. Corrective fix for Phase A's stale-file backfill — see
    POSTGRES_MIGRATION_PLAN.md. Upsert-only and idempotent, safe to re-run."""
    denied = _require_admin()
    if denied:
        return denied
    html = _admin_task_page(
        title="Postgres Full Backfill",
        api_path="/api/pg-full-backfill",
        description="Upserts EVERY item currently in the live _cat_cache (available items "
                    "plus sold/delisted history — both intentionally kept) into the Postgres "
                    "items table, correcting the incomplete Phase A backfill. Upsert-only, "
                    "never deletes. Takes a few minutes at full catalog scale (~440K items). "
                    "Run /api/pg-parity-check afterward to confirm pg_total == json_total.",
        nav_current="/admin/pg-backfill",
    )
    return Response(html, content_type="text/html")


@app.route("/admin/users")
def admin_users():
    """Session-protected user account summary page."""
    denied = _require_admin()
    if denied:
        return denied

    # Auto-purge any users whose scheduled deletion date has passed
    from datetime import timezone as _tz
    now_iso = datetime.now(_tz.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    with _user_db() as conn:
        due = [r["id"] for r in conn.execute(
            "SELECT id FROM users WHERE deleted_at IS NOT NULL AND deleted_at <= ?", (now_iso,)
        ).fetchall()]
        for uid_del in due:
            conn.execute("DELETE FROM user_data WHERE user_id=?", (uid_del,))
            conn.execute("DELETE FROM users WHERE id=?", (uid_del,))
        if due:
            conn.commit()

    # Load all users + their data
    with _user_db() as conn:
        users = [dict(r) for r in conn.execute(
            "SELECT u.id, u.username, u.email, u.created_at, u.deleted_at, "
            "       u.last_login, d.last_run, d.updated_at "
            "FROM users u "
            "LEFT JOIN user_data d ON d.user_id = u.id "
            "ORDER BY u.created_at DESC"
        ).fetchall()]

        # Count watchlist/keyword items per user
        # Cross-reference watchlist against live catalog so sold/gone items don't inflate the count
        _load_cat_cache()
        for u in users:
            row = conn.execute(
                "SELECT watchlist, keywords, favorites FROM user_data WHERE user_id=?",
                (u["id"],)
            ).fetchone()
            if row:
                try:
                    _wl = json.loads(row["watchlist"] or "{}")
                    u["wl_count"] = sum(1 for sku in _wl if sku in _cat_cache)
                except: u["wl_count"] = 0
                try: u["kw_count"]  = len(json.loads(row["keywords"]  or "[]"))
                except: u["kw_count"] = 0
                try: u["fav_count"] = len(json.loads(row["favorites"] or "[]"))
                except: u["fav_count"] = 0
            else:
                u["wl_count"] = u["kw_count"] = u["fav_count"] = 0

    def _fmt(ts):
        if not ts: return "—"
        try:
            dt = datetime.strptime(ts, "%Y-%m-%dT%H:%M:%SZ")
            return dt.strftime("%b %d, %Y  %H:%M UTC")
        except: return ts

    csrf_token = _admin_page_csrf()
    rows_html = ""
    for u in users:
        last_scan      = _fmt(u.get("last_run"))
        joined         = _fmt(u.get("created_at"))
        last_login_fmt = _fmt(u.get("last_login"))
        uid            = int(u["id"])
        uname_safe  = _html.escape(str(u["username"]))
        deleted_at  = u.get("deleted_at") or ""
        if deleted_at:
            # Show delete-on date and Cancel / Delete Now options
            try:
                del_dt = datetime.strptime(deleted_at, "%Y-%m-%dT%H:%M:%SZ")
                del_label = del_dt.strftime("Deletes %b %d")
            except Exception:
                del_label = "Scheduled"
            action_html = (
                f'<span style="color:#a05050;font-size:.75rem">{del_label}</span> '
                f'<form method="POST" action="/admin/delete-user" style="display:inline">'
                f'<input type="hidden" name="id" value="{uid}">'
                f'<input type="hidden" name="_csrf" value="{csrf_token}">'
                f'<input type="hidden" name="action" value="cancel">'
                f'<button type="submit" style="background:#1a3a1a;color:#8fc88f;border:none;border-radius:4px;padding:2px 8px;cursor:pointer;font-size:.75rem">Undo</button>'
                f'</form> '
                f'<form method="POST" action="/admin/delete-user" style="display:inline"'
                f' onsubmit="return confirm(\'Permanently delete {uname_safe} right now?\')">'
                f'<input type="hidden" name="id" value="{uid}">'
                f'<input type="hidden" name="_csrf" value="{csrf_token}">'
                f'<input type="hidden" name="action" value="now">'
                f'<button type="submit" style="background:#600;color:#fcc;border:none;border-radius:4px;padding:2px 8px;cursor:pointer;font-size:.75rem">Delete Now</button>'
                f'</form>'
            )
            row_style = ' style="opacity:.6"'
        else:
            action_html = (
                f'<form method="POST" action="/admin/delete-user" style="display:inline"'
                f' onsubmit="return confirm(\'Schedule {uname_safe} for deletion in 10 days?\')">'
                f'<input type="hidden" name="id" value="{uid}">'
                f'<input type="hidden" name="_csrf" value="{csrf_token}">'
                f'<input type="hidden" name="action" value="schedule">'
                f'<button type="submit" style="background:#600;color:#fcc;border:none;border-radius:4px;padding:3px 10px;cursor:pointer;font-size:.78rem">✕ Delete</button>'
                f'</form>'
            )
            row_style = ''
        rows_html += (
            f'<tr{row_style}>'
            f'<td>{uname_safe}</td>'
            f'<td data-value="{_html.escape(u.get("created_at",""))}">{_html.escape(str(joined))}</td>'
            f'<td data-value="{_html.escape(u.get("last_run","") or "")}">{_html.escape(str(last_scan))}</td>'
            f'<td data-value="{_html.escape(u.get("last_login","") or "")}">{_html.escape(str(last_login_fmt))}</td>'
            f'<td style="text-align:center">{int(u["wl_count"])}</td>'
            f'<td style="text-align:center">{int(u["kw_count"])}</td>'
            f'<td style="text-align:center">{int(u["fav_count"])}</td>'
            f'<td style="text-align:center">{action_html}</td>'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Users</title>
<style>
body{{background:#111;color:#ddd;font-family:monospace;padding:24px;font-size:.88rem}}
h1{{color:#fff;margin-bottom:4px}}
.stat{{color:#eee;font-size:1.1rem;font-weight:bold;margin-bottom:6px}}
.sub{{color:#666;margin-bottom:20px;font-size:.82rem}}
table{{border-collapse:collapse;width:100%;max-width:1100px}}
th{{background:#1e1e1e;padding:8px 14px;text-align:left;border-bottom:2px solid #333;color:#aaa}}
th[data-col]:not([data-col="-1"]){{cursor:pointer;user-select:none}}
th[data-col]:not([data-col="-1"]):hover{{color:#fff}}
th.sort-asc::after{{content:" ↑";color:#7af}}
th.sort-desc::after{{content:" ↓";color:#7af}}
td{{padding:7px 14px;border-bottom:1px solid #222}}
tr:hover td{{background:#1a1a1a}}
a{{color:#888;text-decoration:none;font-size:.78rem}}
</style>
<script src="/static/admin.js"></script>
</head><body>
{_admin_nav('/admin/users')}
<h1>👤 User Accounts</h1>
<div class="stat">{len(users)} user{"s" if len(users) != 1 else ""}</div>
<div class="sub"><a href="/admin/devices">→ Device log</a></div>
<table>
<tr>
  <th data-col="0">Username</th>
  <th data-col="1">Joined</th>
  <th data-col="2">Last scan</th>
  <th data-col="3">Last login</th>
  <th data-col="4" style="text-align:center">Watch</th>
  <th data-col="5" style="text-align:center">Want</th>
  <th data-col="6" style="text-align:center">Favs</th>
  <th data-col="-1"></th>
</tr>
{rows_html if rows_html else '<tr><td colspan="8" style="color:#555;padding:20px">No accounts yet.</td></tr>'}
</table>
</body></html>"""

    return Response(html, mimetype="text/html")


@app.route("/admin/delete-user", methods=["POST"])
def admin_delete_user():
    """Schedule a user for deletion (soft-delete) or cancel/confirm from the admin panel."""
    denied = _require_admin()
    if denied:
        return denied
    # CSRF: validate form token
    submitted = request.form.get("_csrf", "")
    expected  = session.get("_admin_csrf") or ""
    if not expected or not submitted or not hmac.compare_digest(submitted, expected):
        return Response("Invalid CSRF token — go back and reload the page.", status=403, content_type="text/plain")
    user_id = request.form.get("id", "")
    action  = request.form.get("action", "schedule")  # "schedule" | "cancel" | "now"
    try:
        user_id = int(user_id)
    except (ValueError, TypeError):
        return Response("Invalid user id.", status=400, content_type="text/plain")
    user = _user_by_id(user_id)
    if not user:
        return Response("User not found.", status=404, content_type="text/plain")
    with _user_db() as conn:
        if action == "now":
            conn.execute("DELETE FROM user_data WHERE user_id=?", (user_id,))
            conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        elif action == "cancel":
            conn.execute("UPDATE users SET deleted_at=NULL WHERE id=?", (user_id,))
        else:  # schedule
            from datetime import timezone
            delete_on = (datetime.now(timezone.utc) + timedelta(days=10)).strftime("%Y-%m-%dT%H:%M:%SZ")
            conn.execute("UPDATE users SET deleted_at=? WHERE id=?", (delete_on, user_id))
    return redirect("/admin/users")


@app.route("/admin/clear-lock")
def admin_clear_lock():
    """Force-release the global scan lock if it's stuck after a crash.
    Protected by admin session."""
    denied = _require_admin()
    if denied:
        return denied
    if _lock.locked():
        try:
            _lock.release()
            return Response("✓ Lock cleared — scans can now run.", content_type="text/plain")
        except RuntimeError:
            return Response("Lock was already free (release failed).", content_type="text/plain")
    return Response("Lock was not held — nothing to clear.", content_type="text/plain")


@app.route("/admin/listing-patterns")
def admin_listing_patterns():
    """Analyze date_listed distribution across the cached inventory to reveal
    how GC batches new listings — by day, hour-of-day, and minute within hour.
    Protected by admin session."""
    denied = _require_admin()
    if denied:
        return denied

    _load_cat_cache()
    from collections import Counter

    dates, hours, minutes, exact_times, items_no_date = [], [], [], [], 0
    for sku, item in _cat_cache.items():
        dl = item.get("date_listed", "")
        if not dl:
            items_no_date += 1
            continue
        exact_times.append(dl)
        # "2026-04-15T14:23:00Z"
        try:
            date_part  = dl[:10]          # "2026-04-15"
            hour_part  = int(dl[11:13])   # 14
            minute_part= int(dl[14:16])   # 23
            dates.append(date_part)
            hours.append(hour_part)
            minutes.append(minute_part)
        except Exception:
            pass

    total = len(exact_times) + items_no_date
    by_date  = Counter(dates).most_common(60)
    by_hour  = sorted(Counter(hours).items())
    by_minute= sorted(Counter(minutes).items())

    # Look for clustering: what fraction of items land on the exact :00 second?
    on_zero_second = sum(1 for t in exact_times if t.endswith("T00:00:00Z") or t[17:19] == "00")
    on_midnight    = sum(1 for t in exact_times if t[11:19] == "00:00:00")

    # Sample of 40 most recent timestamps (sorted desc)
    recent = sorted(exact_times, reverse=True)[:40]

    S = '<style>body{font-family:monospace;background:#111;color:#ddd;padding:24px;max-width:900px}' \
        'h2{color:#f5c518}table{border-collapse:collapse;width:100%}' \
        'td,th{border:1px solid #333;padding:6px 10px;text-align:right}' \
        'th{background:#222;text-align:center}td:first-child{text-align:left}' \
        '.bar{display:inline-block;background:#c00;height:12px;vertical-align:middle}' \
        '.note{color:#888;font-size:.85em;margin:8px 0}</style>'

    def bar(n, mx):
        w = int(n / mx * 200) if mx else 0
        return f'<span class="bar" style="width:{w}px"></span> {n:,}'

    html = [f'<html><head><title>GC Listing Patterns</title>{S}</head><body>']
    html.append(_admin_nav('/admin/listing-patterns'))
    html.append(f'<h2>GC Listing Pattern Analysis</h2>')
    html.append(f'<p class="note">Total items in cache: <b>{total:,}</b> &nbsp;|&nbsp; '
                f'With date_listed: <b>{len(exact_times):,}</b> &nbsp;|&nbsp; '
                f'Missing date: <b>{items_no_date:,}</b></p>')
    html.append(f'<p class="note">Items landing at exactly midnight UTC: <b>{on_midnight:,}</b> '
                f'({on_midnight/len(exact_times)*100:.1f}%)</p>')
    html.append(f'<p class="note">Items with :00 seconds: <b>{on_zero_second:,}</b> '
                f'({on_zero_second/len(exact_times)*100:.1f}%) — '
                f'(high % = timestamps truncated to the minute, not exact)</p>')

    # By hour of day
    mx_h = max(c for _,c in by_hour) if by_hour else 1
    html.append('<h2>Items by Hour of Day (UTC)</h2><table><tr><th>Hour (UTC)</th><th>Count</th><th>Distribution</th></tr>')
    for h, c in by_hour:
        html.append(f'<tr><td>{h:02d}:00</td><td>{c:,}</td><td>{bar(c, mx_h)}</td></tr>')
    html.append('</table>')

    # By minute within the hour
    mx_m = max(c for _,c in by_minute) if by_minute else 1
    html.append('<h2>Items by Minute Within Hour</h2>'
                '<p class="note">Spikes at :00 or other specific minutes = batch publishing</p>'
                '<table><tr><th>Minute</th><th>Count</th><th>Distribution</th></tr>')
    for m, c in by_minute:
        html.append(f'<tr><td>:{m:02d}</td><td>{c:,}</td><td>{bar(c, mx_m)}</td></tr>')
    html.append('</table>')

    # By date (most recent first)
    mx_d = max(c for _,c in by_date) if by_date else 1
    html.append('<h2>Items by Date Listed (top 60)</h2><table><tr><th>Date</th><th>Count</th><th>Distribution</th></tr>')
    for d, c in sorted(by_date, reverse=True):
        html.append(f'<tr><td>{d}</td><td>{c:,}</td><td>{bar(c, mx_d)}</td></tr>')
    html.append('</table>')

    # 40 most recent timestamps raw
    html.append('<h2>40 Most Recent date_listed Values</h2>'
                '<p class="note">Look for identical timestamps (batch) vs spread-out (item-by-item)</p>'
                '<table><tr><th>Timestamp (UTC)</th></tr>')
    for t in recent:
        html.append(f'<tr><td>{t}</td></tr>')
    html.append('</table>')

    html.append('</body></html>')
    return Response("".join(html), content_type="text/html")


@app.route("/api/reset", methods=["POST"])
@optional_user_context
def api_reset():
    """Delete inventory state and cache to start fresh.
    Preserves favorites, watchlist, and want list."""
    denied = _require_admin_api()
    if denied:
        return denied
    deleted = []
    for f in [STATE_FILE, CAT_CACHE_FILE, OUTPUT_FILE,
              DATA_DIR / "gc_last_scan.txt",
              DATA_DIR / "gc_invalid_stores.json",
              DATA_DIR / "gc_condition_diag.json",
              DATA_DIR / "gc_debug_listing.html"]:
        if f.exists():
            f.unlink()
            deleted.append(f.name)
    global _cat_cache
    _cat_cache = {}
    return jsonify({"deleted": deleted, "status": "Reset complete. Ready for a fresh baseline."})

@app.route("/api/clear-blocklist", methods=["POST"])
@optional_user_context
def api_clear_blocklist():
    """Remove the invalid stores blocklist so all stores are re-evaluated."""
    denied = _require_admin_api()
    if denied:
        return denied
    f = DATA_DIR / "gc_invalid_stores.json"
    if f.exists():
        f.unlink()
    return jsonify({"status": "Blocklist cleared. Run Validate Stores to re-check all stores."})

def _build_stores_noscript() -> str:
    """Build a <noscript> block listing all known GC store locations for SEO.
    Generated at request time so it always reflects the live store cache."""
    try:
        stores = json.loads(STORES_CACHE.read_text()).get("stores", []) if STORES_CACHE.exists() else []
    except Exception:
        stores = []
    if not stores:
        return ''
    # City names link to the per-store landing pages — the crawl path from "/" to
    # all ~240 store pages (sitemap alone is weaker than real links). (2026-07 audit S4)
    links = ", ".join(
        f'<a href="/store/{_store_slug(s)}" style="color:#777">{_html.escape(s)}</a>'
        for s in stores
    )
    return (
        '<noscript><p style="padding:12px 20px;text-align:center;color:#666;'
        'font-size:.72rem;line-height:1.7">'
        'GC Used Inventory Tracker requires JavaScript. '
        'This tool searches used guitar gear at Guitar Center locations nationwide including: '
        + links +
        '.</p></noscript>'
    )

@app.route("/")
@optional_user_context
def index():
    return HTML_TEMPLATE.replace('<!-- __STORES_NOSCRIPT__ -->', _build_stores_noscript())


# ── Per-store SEO landing pages (v2.14.5, 2026-07 audit S1) ──────────────────
# Server-rendered, zero-JS pages so Google has a city-specific URL/title/snippet
# to rank for "guitar center <city> inventory" queries (the homepage was ranking
# page-1 for these with ~zero CTR because it shows a generic title). Purely
# additive — the main app is untouched. Unauthenticated route over the 92K cache,
# so pages are memoized per store keyed by cache mtime (same discipline as
# /api/browse: no per-request full-cache pass beyond the first hit after a scan).

_STORE_PAGE_CACHE: dict = {}   # slug -> (cat_cache_mtime, html)

def _store_slug(name: str) -> str:
    s = re.sub(r'[^a-z0-9]+', '-', (name or '').lower())
    return re.sub(r'-+', '-', s).strip('-')

def _store_slug_map() -> dict:
    """slug -> store name, from the live store cache (4KB read; cheap per request)."""
    try:
        stores = json.loads(STORES_CACHE.read_text()).get("stores", []) if STORES_CACHE.exists() else []
    except Exception:
        stores = []
    return {_store_slug(s): s for s in stores if s}

def _render_store_page(store_name: str, slug: str) -> str:
    _load_cat_cache()
    items = [v for _, v in list(_cat_cache.items())
             if v.get("store") == store_name and v.get("available", True)]
    items.sort(key=lambda i: i.get("date_listed") or "", reverse=True)
    count = len(items)
    cat_counts: dict = {}
    for i in items:
        c = i.get("category") or "Other"
        cat_counts[c] = cat_counts.get(c, 0) + 1
    esc = _html.escape
    city = esc(store_name)
    title = f"Guitar Center {city} Used Gear — Live Inventory ({count} items)"
    # Build desc as PLAIN text — it gets esc()'d exactly once at each interpolation
    # point below. v2.15.0 escaped category names here too, so "&" reached the
    # rendered meta description as a literal "&amp;" (double-escape; Google showed
    # it verbatim in snippets). v2.15.1 fix.
    desc = (f"Browse {count} used items currently at the Guitar Center {store_name} store: "
            + ", ".join(f"{c} ({n})" for c, n in sorted(cat_counts.items(), key=lambda x: -x[1])[:4])
            + ". Updated after every scan — free watch list and want list at GC Used Inventory Tracker.")
    rows = []
    for i in items[:50]:
        price = i.get("price") or 0
        rows.append(
            "<tr><td><a href=\"" + esc(i.get("url") or "https://www.guitarcenter.com/") + "\" rel=\"nofollow noopener\">"
            + esc(i.get("name") or "") + "</a></td>"
            + "<td>" + esc(i.get("brand") or "") + "</td>"
            + "<td>" + (f"${price:,.2f}" if price else "") + "</td>"
            + "<td>" + esc(i.get("condition") or "") + "</td>"
            + "<td>" + esc(_fmt_date(i.get("date_listed") or "")) + "</td></tr>"
        )
    cats_html = " · ".join(
        esc(c) + " <b>" + str(n) + "</b>"
        for c, n in sorted(cat_counts.items(), key=lambda x: -x[1])
    )
    item_list_ld = json.dumps({
        "@context": "https://schema.org", "@type": "ItemList",
        "name": f"Used gear at Guitar Center {store_name}",
        "numberOfItems": count,
        "itemListElement": [
            {"@type": "ListItem", "position": n + 1,
             "name": i.get("name") or "", "url": i.get("url") or ""}
            for n, i in enumerate(items[:20])
        ],
    }, separators=(",", ":"))
    breadcrumb_ld = json.dumps({
        "@context": "https://schema.org", "@type": "BreadcrumbList",
        "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "GC Used Inventory Tracker", "item": "https://gcgeartracker.com/"},
            {"@type": "ListItem", "position": 2, "name": f"Guitar Center {store_name}", "item": f"https://gcgeartracker.com/store/{slug}"},
        ],
    }, separators=(",", ":"))
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<meta name="description" content="{esc(desc)}">
<link rel="canonical" href="https://gcgeartracker.com/store/{slug}">
<meta property="og:type" content="website">
<meta property="og:url" content="https://gcgeartracker.com/store/{slug}">
<meta property="og:title" content="{title}">
<meta property="og:description" content="{esc(desc)}">
<meta property="og:image" content="https://gcgeartracker.com/static/og-image.png">
<link rel="icon" href="/static/favicon.svg" type="image/svg+xml">
<script type="application/ld+json">{item_list_ld}</script>
<script type="application/ld+json">{breadcrumb_ld}</script>
<style>
body{{background:#111;color:#ddd;font-family:-apple-system,system-ui,sans-serif;margin:0;padding:20px;line-height:1.5}}
main{{max-width:960px;margin:0 auto}}
h1{{color:#fff;font-size:1.35rem;margin:8px 0 4px}}
.sub{{color:#888;font-size:.9rem;margin-bottom:14px}}
.cats{{color:#aaa;font-size:.82rem;margin-bottom:18px}} .cats b{{color:#eee}}
.cta{{display:inline-block;background:#c00;color:#fff;padding:9px 16px;border-radius:6px;text-decoration:none;font-weight:600;margin-bottom:20px}}
table{{width:100%;border-collapse:collapse;font-size:.85rem}}
th{{text-align:left;color:#888;font-weight:600;padding:7px 10px;border-bottom:1px solid #333}}
td{{padding:7px 10px;border-bottom:1px solid #222}}
td a{{color:#e88;text-decoration:none}} td a:hover{{text-decoration:underline}}
.foot{{color:#555;font-size:.75rem;margin-top:22px}} .foot a{{color:#888}}
</style>
</head>
<body><main>
<p class="sub"><a href="/" style="color:#888">← GC Used Inventory Tracker</a></p>
<h1>Guitar Center {city} — Used Gear Inventory</h1>
<p class="sub">{count} used items currently tracked at this store. Newest 50 shown below; the full list is in the free tracker.</p>
<p class="cats">{cats_html}</p>
<a class="cta" href="/?store={slug}">Browse all {count} {city} items in the tracker →</a>
<table>
<thead><tr><th>Item</th><th>Brand</th><th>Price</th><th>Condition</th><th>Listed</th></tr></thead>
<tbody>{"".join(rows) if rows else '<tr><td colspan="5" style="color:#777">No used items at this store right now — check back after the next scan.</td></tr>'}</tbody>
</table>
<p class="foot">Prices and availability change constantly — click any item for the live Guitar Center listing.
Independent tool — not affiliated with Guitar Center, Inc. · <a href="/privacy">Privacy Policy</a></p>
</main></body>
</html>"""

@app.route("/store/<slug>")
def store_page(slug):
    name = _store_slug_map().get(slug)
    if not name:
        return "Not found", 404
    try:
        mtime = CAT_CACHE_FILE.stat().st_mtime
    except OSError:
        mtime = 0.0
    hit = _STORE_PAGE_CACHE.get(slug)
    if hit and hit[0] == mtime:
        return hit[1]
    html_out = _render_store_page(name, slug)
    _STORE_PAGE_CACHE[slug] = (mtime, html_out)
    return html_out

@app.route("/cl")
def cl_page():
    return CL_TEMPLATE

@app.route("/newdeals")
def newdeals_page():
    denied = _require_admin()
    if denied: return denied
    return NEWDEALS_TEMPLATE

@app.route("/api/new-scan", methods=["POST"])
def api_new_scan():
    """Fetch all new GC inventory from Algolia, dedupe by SKU, cache to disk."""
    denied = _require_admin_api()
    if denied: return denied
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import datetime as _dt2
    try:
        hits0, nb_pages = _fetch_new_page(0)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    all_hits = list(hits0)
    if nb_pages > 1:
        with ThreadPoolExecutor(max_workers=12) as pool:
            futs = {pool.submit(_fetch_new_page, p): p for p in range(1, nb_pages)}
            for fut in as_completed(futs):
                try:
                    hits, _ = fut.result()
                    all_hits.extend(hits)
                except Exception:
                    pass

    items = {}
    for hit in all_hits:
        sku = str(hit.get("sku") or hit.get("objectID") or "").strip()
        if not sku or sku in items:
            continue
        name = _clean_name(hit.get("displayName") or hit.get("name") or "")
        if not name:
            continue
        try:    price      = float(hit.get("price") or 0)
        except: continue
        try:    list_price = float(hit.get("listPrice") or 0)
        except: list_price = 0.0
        if price <= 0:
            continue
        pct_off = int((1.0 - price / list_price) * 100) if list_price > price > 0 else 0
        # Category: prefer the structured categories array (same as used gear parsing)
        cats_arr = hit.get("categories") or []
        if cats_arr and isinstance(cats_arr, list) and isinstance(cats_arr[0], dict):
            category = cats_arr[0].get("lvl0") or ""
        else:
            # Fallback: categoryPageIds — skip bare "New"/"Used" sentinel values
            cat_ids  = hit.get("categoryPageIds") or []
            category = next((c for c in cat_ids if c and c.lower() not in ("new", "used", "")), "")
        brand    = (hit.get("brand") or "").strip()
        seo_url  = hit.get("seoUrl") or ""
        items[sku] = {
            "name":        name,
            "brand":       brand,
            "category":    category,
            "price":       price,
            "list_price":  list_price,
            "pct_off":     pct_off,
            "url":         ("https://www.guitarcenter.com" + seo_url) if seo_url else "",
            "is_software": _is_software_item(name, category),
        }

    last_updated = _dt2.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    _save_new_deals_cache(items, last_updated)
    return jsonify({"ok": True, "count": len(items), "last_updated": last_updated})


@app.route("/api/new-browse", methods=["POST"])
def api_new_browse():
    """Browse cached new deals with filters + pagination."""
    denied = _require_admin_api()
    if denied: return denied
    import re as _re2
    data    = request.json or {}
    cache   = _load_new_deals_cache()
    if not cache:
        return jsonify({"no_cache": True, "items": [], "total": 0})

    items = list(cache.get("items", {}).values())

    # Software/plugin filter (excluded by default)
    if not bool(data.get("include_software")):
        items = [i for i in items if not i.get("is_software", False)]

    # Keyword search
    fq = (data.get("filter_q") or "").strip().lower()
    if fq:
        tokens = fq.split()
        items = [i for i in items if all(
            t in (i["name"] + " " + i["brand"]).lower() for t in tokens)]

    # Brand / category filters
    brands = [b for b in (data.get("filter_brands") or []) if b]
    if brands:
        items = [i for i in items if i["brand"] in brands]
    cats = [c for c in (data.get("filter_categories") or []) if c]
    if cats:
        items = [i for i in items if i["category"] in cats]

    # % off minimum
    min_pct = int(data.get("filter_min_pct_off") or 0)
    if min_pct > 0:
        items = [i for i in items if i["pct_off"] >= min_pct]

    # Price range
    try:
        pmin = float(data["filter_price_min"]) if data.get("filter_price_min") not in (None, "") else None
    except (TypeError, ValueError):
        pmin = None
    try:
        pmax = float(data["filter_price_max"]) if data.get("filter_price_max") not in (None, "") else None
    except (TypeError, ValueError):
        pmax = None
    if pmin is not None: items = [i for i in items if i["price"] >= pmin]
    if pmax is not None: items = [i for i in items if i["price"] <= pmax]

    # Want list keyword filter — whole-word match per keyword, OR logic across keywords
    if bool(data.get("filter_want_list")):
        keywords = [k.lstrip("=").strip() for k in (data.get("keywords") or []) if k.strip()]
        if keywords:
            pats = [_re2.compile(r'\b' + _re2.escape(k) + r'\b', _re2.IGNORECASE) for k in keywords]
            items = [i for i in items if any(
                p.search(i["name"] + " " + i["brand"]) for p in pats)]

    # Collect available brand + category facets for dropdowns
    all_brands = sorted(set(i["brand"] for i in items if i["brand"]))
    all_cats   = sorted(set(i["category"] for i in items if i["category"]))
    total      = len(items)

    # Sort
    sort_field = data.get("sort") or "pct_off"
    reverse    = (data.get("dir") or "desc") == "desc"
    if sort_field == "price":
        items.sort(key=lambda i: i["price"], reverse=reverse)
    elif sort_field == "list_price":
        items.sort(key=lambda i: i["list_price"], reverse=reverse)
    elif sort_field == "name":
        items.sort(key=lambda i: i["name"].lower(), reverse=reverse)
    elif sort_field == "brand":
        items.sort(key=lambda i: i["brand"].lower(), reverse=reverse)
    elif sort_field == "category":
        items.sort(key=lambda i: i["category"].lower(), reverse=reverse)
    else:  # pct_off default
        items.sort(key=lambda i: i["pct_off"], reverse=reverse)

    # Paginate
    per_page    = 50
    total_pages = max(1, (total + per_page - 1) // per_page)
    page        = max(1, min(int(data.get("page") or 1), total_pages))
    offset      = (page - 1) * per_page

    return jsonify({
        "items":        items[offset:offset + per_page],
        "total":        total,
        "page":         page,
        "total_pages":  total_pages,
        "brands":       all_brands,
        "categories":   all_cats,
        "last_updated": cache.get("last_updated", ""),
    })

@app.route("/privacy")
def privacy_page():
    return PRIVACY_TEMPLATE

@app.route("/google73eeaa5f083d2e84.html")
def google_site_verification():
    return "google-site-verification: google73eeaa5f083d2e84.html", 200, {"Content-Type": "text/html"}

@app.route("/robots.txt")
def robots_txt():
    content = (
        "User-agent: *\n"
        "Allow: /\n"
        "Disallow: /admin/\n"
        "Disallow: /api/\n"
        "\n"
        "Sitemap: https://gcgeartracker.com/sitemap.xml\n"
    )
    return content, 200, {"Content-Type": "text/plain"}

@app.route("/sitemap.xml")
def sitemap_xml():
    # Per-store landing pages included with lastmod = last scan date, so Google
    # recrawls city pages after every scan day. (2026-07 audit S3)
    lastmod = ""
    try:
        last_scan_file = DATA_DIR / "gc_last_scan.txt"
        if last_scan_file.exists():
            raw = last_scan_file.read_text().strip()[:10]  # YYYY-MM-DD
            if len(raw) == 10:
                lastmod = f"<lastmod>{raw}</lastmod>"
    except Exception:
        pass
    urls = [
        f'  <url><loc>https://gcgeartracker.com/</loc>{lastmod}'
        '<changefreq>daily</changefreq><priority>1.0</priority></url>',
        '  <url><loc>https://gcgeartracker.com/privacy</loc>'
        '<changefreq>monthly</changefreq><priority>0.3</priority></url>',
    ]
    for slug in sorted(_store_slug_map()):
        urls.append(
            f'  <url><loc>https://gcgeartracker.com/store/{slug}</loc>{lastmod}'
            '<changefreq>daily</changefreq><priority>0.6</priority></url>'
        )
    content = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        + "\n".join(urls) +
        '\n</urlset>\n'
    )
    return content, 200, {"Content-Type": "application/xml"}

@app.route("/.well-known/security.txt")
def security_txt():
    # RFC 9116 — gives researchers a private channel to report issues instead of
    # posting "this isn't secure" publicly. Update Expires before it lapses.
    content = (
        "Contact: mailto:cboehmig@gmail.com\n"
        "Expires: 2027-06-05T00:00:00Z\n"
        "Preferred-Languages: en\n"
        "Canonical: https://gcgeartracker.com/.well-known/security.txt\n"
    )
    return content, 200, {"Content-Type": "text/plain; charset=utf-8"}

NEWDEALS_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>GC New Deals — Admin</title>
<link rel="icon" type="image/svg+xml" href="/static/og-image.svg">
<link rel="stylesheet" href="/static/gc.css">
<link rel="stylesheet" href="/static/newdeals.css">
<!-- __GA__ -->
</head>
<body>
<div class="nd-page">

  <div class="nd-header">
    <a href="/" class="nd-back-link">&#8592; Tracker</a>
    <h1 class="nd-title">GC New Deals <span class="nd-admin-badge">Admin</span></h1>
  </div>

  <div class="nd-status">
    <span id="nd-refresh-time">No data cached yet &#8212; click Refresh to load.</span>
    <span id="nd-item-count"></span>
    <button id="nd-refresh-btn" data-action="refresh">&#8635; Refresh Data</button>
  </div>

  <div class="nd-top-bar" id="nd-top-bar">
    <div class="nd-chips">
      <button class="chip-btn" id="nd-wl-btn" data-action="toggle-wantlist">&#127919; Want List</button>
      <label class="nd-sw-label">
        <input type="checkbox" id="nd-include-sw"> Include Software / Plugins
      </label>
    </div>
    <div class="nd-filter-bar">
      <div class="nd-search-wrap">
        <input type="text" id="nd-search" placeholder="Search items&#8230;" autocomplete="off">
      </div>
      <select id="nd-brand-sel"><option value="">All Brands</option></select>
      <select id="nd-cat-sel"><option value="">All Categories</option></select>
      <select id="nd-pct-sel">
        <option value="0">Any discount</option>
        <option value="20">20%+ off</option>
        <option value="30">30%+ off</option>
        <option value="40" selected>40%+ off</option>
        <option value="50">50%+ off</option>
        <option value="60">60%+ off</option>
      </select>
      <input type="number" id="nd-price-min" placeholder="$Min" min="0">
      <span class="nd-price-sep">&#8211;</span>
      <input type="number" id="nd-price-max" placeholder="$Max" min="0">
      <button id="nd-clear-btn" data-action="clear-filters">&#10005; Clear</button>
    </div>
  </div>

  <div class="nd-results-hdr">
    <span id="nd-result-count"></span>
  </div>

  <div id="nd-empty-msg" class="nd-empty">No data cached yet &#8212; click &#8635; Refresh Data to load inventory.</div>

  <div id="nd-results-wrap" style="display:none">
    <table class="nd-table">
      <thead>
        <tr>
          <th class="nd-th" data-sort="pct_off">% Off</th>
          <th class="nd-th" data-sort="price">Sale Price</th>
          <th class="nd-th" data-sort="list_price">MSRP</th>
          <th class="nd-th nd-th-name" data-sort="name">Name</th>
          <th class="nd-th" data-sort="brand">Brand</th>
          <th class="nd-th" data-sort="category">Category</th>
        </tr>
      </thead>
      <tbody id="nd-tbody"></tbody>
    </table>
  </div>

  <div id="nd-paginator" class="nd-paginator"></div>

  <div id="dev-footer">
    <span><a href="/">&#8592; Main Tracker</a> &nbsp;&#183;&nbsp; GC New Deals (Admin)</span>
  </div>

</div>
<script src="/static/newdeals.js" defer></script>
</body>
</html>
"""

CL_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CL Used Gear Search</title>
<link rel="stylesheet" href="/static/cl.css">
<!-- __GA__ -->
</head>
<body>

<header>
  <h1>CL Used Gear Search <span>craigslist aggregator</span></h1>
  <span id="hdr-user"></span>
  <button id="hdr-signout" style="display:none">Sign Out</button>
</header>

<div class="cl-wrap">
  <!-- City sidebar -->
  <div class="cl-left">
    <div class="search-wrap">
      <input id="cl-city-search" type="text" placeholder="Search cities…" autocomplete="off">
      <div class="cl-sel-btns">
        <button class="cl-sel-btn" id="cl-favs-btn">★ Favorites</button>
        <button class="cl-sel-btn" id="cl-select-all-btn">Select All</button>
        <button class="cl-sel-btn" id="cl-clear-all-btn">Clear All</button>
      </div>
    </div>
    <div id="cl-city-list"></div>
  </div>

  <!-- Search + results -->
  <div class="cl-right">
    <div class="cl-search-bar">
      <input id="cl-query" type="text" placeholder="e.g. telecaster, les paul, fender twin…"
        autocomplete="off">
      <span id="cl-status"></span>
      <button id="cl-search-btn">Search</button>
    </div>

    <div class="cl-results-hdr" id="cl-toolbar">
      <button id="cl-watchlist-toggle" class="cl-chip">★ Watch List</button>
      <button id="cl-wantlist-btn" class="cl-chip">🎯 Want List</button>
      <a id="cl-wl-link" style="display:none">Clear Want List Search</a>
    </div>

    <div class="cl-results-hdr" id="cl-results-hdr" style="display:none">
      <span id="cl-count"></span>
      <input id="cl-res-search" type="text" placeholder="Filter results…" autocomplete="off">
    </div>

    <div id="cl-body">
      <div class="cl-empty">Select cities on the left, enter a search term, and press Search.<br><br>
        Searches all selected Craigslist markets simultaneously for used musical gear.</div>
    </div>
  </div>
</div>

<!-- Auth modal -->
<div id="auth-modal" class="open">
  <div class="auth-box">
    <h2>Sign In</h2>
    <p>Sign in with your GC Tracker account to search Craigslist.</p>
    <div id="cl-google-wrap" style="display:none">
      <button class="auth-google-btn" id="cl-auth-google-btn">
        <svg width="18" height="18" viewBox="0 0 18 18"><path fill="#4285F4" d="M17.64 9.2c0-.637-.057-1.251-.164-1.84H9v3.481h4.844c-.209 1.125-.843 2.078-1.796 2.717v2.258h2.908c1.702-1.566 2.684-3.875 2.684-6.615z"/><path fill="#34A853" d="M9 18c2.43 0 4.467-.806 5.956-2.18l-2.908-2.259c-.806.54-1.837.86-3.048.86-2.344 0-4.328-1.584-5.036-3.711H.957v2.332A8.997 8.997 0 0 0 9 18z"/><path fill="#FBBC05" d="M3.964 10.71A5.41 5.41 0 0 1 3.682 9c0-.593.102-1.17.282-1.71V4.958H.957A8.996 8.996 0 0 0 0 9c0 1.452.348 2.827.957 4.042l3.007-2.332z"/><path fill="#EA4335" d="M9 3.58c1.321 0 2.508.454 3.44 1.345l2.582-2.58C13.463.891 11.426 0 9 0A8.997 8.997 0 0 0 .957 4.958L3.964 7.29C4.672 5.163 6.656 3.58 9 3.58z"/></svg>
        Sign in with Google
      </button>
      <div class="auth-divider"><span>or sign in with username</span></div>
    </div>
    <div class="auth-field"><label>Username</label><input id="auth-user" type="text" autocomplete="username"></div>
    <div class="auth-field"><label>Password</label><input id="auth-pw" type="password" autocomplete="current-password"></div>
    <button class="auth-submit" id="cl-login-submit">Sign In</button>
    <div class="auth-err" id="auth-err"></div>
  </div>
</div>

<script src="/static/cl.js"></script>
</body>
</html>"""

@app.route("/download/excel")
@optional_user_context
def download_excel():
    if not OUTPUT_FILE.exists():
        return "No Excel file yet — run the tracker first.", 404
    return send_file(OUTPUT_FILE, as_attachment=True,
                     download_name="gc_new_inventory.xlsx")

@app.route("/api/stores")
@optional_user_context
def api_stores():
    return jsonify({
        "stores":    get_store_list(),
        "info":      get_store_info(),
    })

@app.route("/api/stores/refresh", methods=["POST"])
@optional_user_context
def api_stores_refresh():
    # Admin only: this fans out dozens of synchronous outbound requests to
    # guitarcenter.com AND overwrites the shared store-list cache. Left open it was
    # both an unauthenticated outbound-amplification DoS and a way for anyone to wipe
    # the global store list. Not called by any frontend — it's a maintenance action.
    denied = _require_admin_api()
    if denied:
        return denied
    stores = refresh_store_list()
    info   = get_store_info()
    return jsonify({"stores": stores,
                    "count": len(stores), "info": info})

# (Dead legacy POST /api/favorites removed in v2.14.5 — favorites live in SQLite via
#  /api/sync; no frontend called it. Login-gated since v2.12.28, deleted per 2026-07
#  audit E3. FAVORITES_FILE stays: admin export/import/reset still reference it.)


# Synthetic brand-facet label so brandless items (brand == "") are filterable
# in both /api/browse and /api/saved-search-counts.
NO_BRAND_LABEL = "(none)"


# ── Shared query-matching helpers (v2.16.0) ──────────────────────────────────
# Hoisted from api_browse's function locals so /api/saved-search-counts uses
# IDENTICAL semantics (its old inline copy had drifted: it treated filter_strict
# as whole-word when browse treats it as fuzzy/contains, and it searched 6 fields
# where browse searches name+brand — so counts didn't match applied results).
# Also adds the v2.16.0 operators: leading '-' = NOT on a term, ';' = OR between
# clauses (lowest precedence). Comma/space AND, quotes, wildcards unchanged.
# v2.16.3 adds 'prefix : b1; b2; ...' (_expand_colon_prefix) so a common AND/NOT
# prefix applies to every OR branch instead of just the one it's typed next to.
# These are pure functions of their inputs — no request state.

_SIMPLE_KW_RE = re.compile(r'^\w+$')   # a single word token
_KW_SPLIT_RE  = re.compile(r'\W+')     # tokenizer matching \b boundaries

# ── Colon-prefix OR expansion (v2.16.3) ───────────────────────────────────────
# User-reported gap: 'Mesa, -combo; Angel; Blues; ...' only applies "Mesa" and
# "-combo" to the FIRST clause — ';' has no cross-clause memory by design (no
# parentheses, deliberately, to keep this a flat O(items) matcher on an
# unauthenticated endpoint). 'prefix : branch1; branch2; ...' lets a common
# AND/NOT condition apply to every OR branch by expanding BEFORE the existing,
# already-fuzzed clause compilers ever see it — zero new matching primitives.
_COLON_PREFIX_MAX_BRANCHES = 30  # independent ceiling; expansion multiplies
                                  # token count, so it needs its own cap even
                                  # though the per-entry char/keyword caps exist.

def _find_prefix_colon(s):
    """Index of the first ':' not inside a double-quoted phrase, else -1."""
    in_quotes = False
    for i, ch in enumerate(s):
        if ch == '"':
            in_quotes = not in_quotes
        elif ch == ':' and not in_quotes:
            return i
    return -1

def _expand_colon_prefix(s, join=', '):
    """'prefix : b1; b2; b3' -> 'prefix<join>b1; prefix<join>b2; prefix<join>b3'
    so the prefix (its own comma/dash AND/NOT terms) applies to every OR
    branch instead of just the one it's typed next to. Only the first
    non-quoted ':' is the marker; strings with no such colon are returned
    completely unchanged, so every existing want list / saved search / query
    with no colon in it parses byte-identically to before. `join` matches
    each caller's own AND separator: want-list clauses are comma-AND (', '),
    filter_q clauses are space-AND (' ')."""
    idx = _find_prefix_colon(s)
    if idx == -1:
        return s
    prefix = s[:idx].strip()
    rest = s[idx + 1:]
    if not prefix:
        return s
    branches = [b.strip() for b in rest.split(';') if b.strip()][:_COLON_PREFIX_MAX_BRANCHES]
    if not branches:
        return s
    return '; '.join(prefix + join + b for b in branches)

def _compile_query(query_str, fuzzy=False):
    """Parse a query string into AND-joined terms.
    Syntax:
      Allen          → whole-word match  (won't match Allentown, McAllen)
      "Jam Pedals"   → exact phrase match
      Thorpy, Dane   → comma = AND; each part uses same rules
      OD*            → wildcard: * is a glob wildcard (OD808, OD-1, etc.)
      fuzzy=True     → plain terms use contains matching instead of whole-word
    """
    terms = []
    for part in query_str.split(','):
        part = part.strip()
        if not part:
            continue
        if part.startswith('"') and part.endswith('"') and len(part) > 2:
            terms.append(('exact', part[1:-1].lower()))
        elif '*' in part:
            pieces = [re.escape(p) for p in part.split('*')]
            terms.append(('regex', re.compile('.*'.join(pieces), re.IGNORECASE)))
        elif fuzzy:
            terms.append(('contains', part.lower()))
        else:
            terms.append(('word', re.compile(r'\b' + re.escape(part) + r'\b', re.IGNORECASE)))
    return terms

def _matches_all(text_lower, terms):
    for mode, val in terms:
        if mode in ('exact', 'contains'):
            if val not in text_lower:
                return False
        else:
            if not val.search(text_lower):
                return False
    return bool(terms)

def _matches_any(text_lower, terms):
    """True if ANY term matches — used for the negation lists."""
    for mode, val in terms:
        if mode in ('exact', 'contains'):
            if val in text_lower:
                return True
        elif val.search(text_lower):
            return True
    return False

def _kw_or_pattern(base):
    """Regex pattern string for one OR-able want-list keyword, mirroring the
    single-term branches of _compile_query so the alternation matches identically:
      "phrase"  -> escaped substring (exact, no boundaries)
      od*       -> escaped pieces joined by .*  (wildcard)
      big muff  -> \\bbig\\ muff\\b             (whole-word / phrase)
    Compiled case-insensitively by the caller."""
    if base.startswith('"') and base.endswith('"') and len(base) > 2:
        return re.escape(base[1:-1])
    if '*' in base:
        return '.*'.join(re.escape(p) for p in base.split('*'))
    return r'\b' + re.escape(base) + r'\b'

# search-box (filter_q) clause compiler: ';' OR-clauses of space-separated tokens,
# '-tok' negated. Shared 12-token budget across all clauses, ≤4 clauses — same
# unauthenticated-DoS budget as the old single-clause path (v2.13.0 cap).
_FQ_TOKEN_RE = re.compile(r'-?"[^"]+"|\S+')

def _compile_fq_clauses(fq, fuzzy=False, max_tokens=12, max_clauses=4):
    fq = _expand_colon_prefix(fq, join=' ')  # v2.16.3 — see _expand_colon_prefix
    clauses = []
    budget = max_tokens
    for cl in fq.split(';'):
        cl = cl.strip()
        if not cl or budget <= 0:
            continue
        toks = _FQ_TOKEN_RE.findall(cl)[:budget]
        budget -= len(toks)
        pos, neg = [], []
        for tok in toks:
            if len(tok) > 1 and tok[0] == '-':
                neg.extend(_compile_query(tok[1:], fuzzy=fuzzy))
            else:
                pos.extend(_compile_query(tok, fuzzy=fuzzy))
        if pos or neg:
            clauses.append((pos, neg))
        if len(clauses) >= max_clauses:
            break
    return clauses

def _fq_text_match(text_lower, clauses):
    """True if ANY clause matches: all positives present, no negative present."""
    for pos, neg in clauses:
        if pos and not _matches_all(text_lower, pos):
            continue
        if neg and _matches_any(text_lower, neg):
            continue
        return True
    return False

def _wl_bool_compile(base):
    """Compile a want-list entry that uses the v2.16.0 ';' (OR) / '-' (NOT)
    syntax. Returns a list of (pos_terms, neg_terms, required_tokens) clauses,
    or None if the entry uses no new syntax (caller falls through to the
    legacy paths — old entries stay byte-identical in behavior). A clause needs
    at least one positive part; an all-negative entry returns None (a want-list
    entry matching "everything except X" would highlight the whole catalog)."""
    parts_all = [p.strip() for cl in base.split(';') for p in cl.split(',')]
    has_neg = any(len(p) > 1 and p[0] == '-' for p in parts_all)
    if ';' not in base and not has_neg:
        return None
    clauses = []
    for cl in base.split(';'):
        cl = cl.strip()
        if not cl:
            continue
        pos, neg, req = [], [], set()
        for p in cl.split(','):
            p = p.strip()
            if not p:
                continue
            if len(p) > 1 and p[0] == '-':
                neg.extend(_compile_query(p[1:]))
            else:
                pos.extend(_compile_query(p))
                pl = p.lower()
                # Required-tokens pre-filter (sound, same trick as the phrase and
                # comma-AND paths): only from plain (non-quoted, non-wildcard) parts.
                if '*' not in pl and not (pl.startswith('"') and pl.endswith('"') and len(pl) > 2):
                    req |= set(_KW_SPLIT_RE.split(pl)) - {''}
        if pos:
            clauses.append((pos, neg, req))
    return clauses or None


@app.route("/api/saved-search-counts", methods=["POST"])
def api_saved_search_counts():
    """Return match counts for each saved search in a single batch call."""
    # Require a logged-in session — this endpoint loads the full 92K-item cache
    # and filters it for each search entry, so an unbounded unauthenticated
    # request would be a trivial CPU DoS.
    if not session.get("user_id"):
        return jsonify({"error": "Not logged in."}), 401
    data     = request.json or {}
    searches = data.get("searches", [])
    if not searches:
        return jsonify({"counts": []})
    # Hard cap so even authenticated users can't send thousands of searches.
    searches = searches[:50]
    # Use the mtime-memoized in-memory cache (v2.13.0) instead of re-reading and
    # re-parsing the 51MB file from disk on every call (~400ms, GIL-held — it stalled
    # every other request thread). Same read-only snapshot idiom as /api/browse.
    # (2026-07 audit E1)
    _load_cat_cache()
    all_items = [v for _, v in list(_cat_cache.items())]
    if not all_items:
        return jsonify({"counts": [0] * len(searches)})

    counts = []
    for search in searches:
        stores   = set(search.get("stores") or [])
        f        = search.get("filters") or {}
        fq       = (f.get("filter_q") or "").lower().strip()[:200]   # clamp len, parity with /api/browse (v2.13.0)
        f_brands = set(f.get("filter_brands") or [])
        f_conds  = set(f.get("filter_conditions") or [])
        f_cats   = set(f.get("filter_categories") or [])
        f_subs   = set(f.get("filter_subcategories") or [])
        f_strict = bool(f.get("filter_strict"))
        f_pdrop  = bool(f.get("filter_price_drop_only"))

        items = [i for i in all_items if i.get("store") in stores] if stores else list(all_items)

        if fq:
            # v2.16.0: use the SAME shared clause compiler + the SAME name+brand text
            # as /api/browse's _apply_base, so a saved search's count equals what
            # applying it actually returns. The old inline copy here had drifted two
            # ways: it treated filter_strict as whole-word (browse treats it as
            # fuzzy/contains, per the v2.10.5 semantics flip) and it searched 6
            # fields where browse searches name+brand — both made counts wrong.
            fq_clauses = _compile_fq_clauses(fq, fuzzy=f_strict)
            items = [i for i in items if fq_clauses and _fq_text_match(
                ((i.get("name") or "") + " " + (i.get("brand") or "")).lower(), fq_clauses)]

        if f_brands: items = [i for i in items if (i.get("brand") in f_brands) or (not i.get("brand") and NO_BRAND_LABEL in f_brands)]
        if f_conds:  items = [i for i in items if i.get("condition") in f_conds]
        if f_cats:   items = [i for i in items if i.get("category") in f_cats]
        if f_subs:   items = [i for i in items if i.get("subcategory") in f_subs]
        if f_pdrop:  items = [i for i in items if (i.get("price_drop") or 0) > 0]
        def _sc_float(v):
            try: return float(v) if v is not None and v != '' else None
            except (TypeError, ValueError): return None
        sc_pmin = _sc_float(f.get("filter_price_min"))
        sc_pmax = _sc_float(f.get("filter_price_max"))
        if sc_pmin is not None: items = [i for i in items if (i.get("price_raw") or 0) >= sc_pmin]
        if sc_pmax is not None: items = [i for i in items if (i.get("price_raw") or 0) <= sc_pmax]

        counts.append(len(items))

    return jsonify({"counts": counts})


# ── Algolia key health check ────────────────────────────────────────────────
# The Algolia search key is the single point of failure for scans: if GC ever
# rotates it, scans start returning 401/403 and silently stop finding inventory.
# This endpoint runs the same used-inventory query the scanner uses (hitsPerPage:0)
# so an external monitor can catch a dead key / schema drift within a day. Result
# is cached ~15 min so it can't be hammered to burn GC's Algolia quota
# (≤ ~96 real probes/day no matter how often it's hit). Public, returns no secret.
_ALGOLIA_HEALTH = {"ts": 0.0, "result": None}
_ALGOLIA_HEALTH_TTL = 900  # seconds
# Refresh lock: without it, N concurrent requests arriving after TTL expiry each fire
# a live probe at GC before any writes the result back — the "≤96 probes/day" cap only
# holds if misses are serialized. (2026-07 audit L1)
_ALGOLIA_HEALTH_LOCK = threading.Lock()

@app.route("/api/health/algolia")
def api_health_algolia():
    import time as _t
    now = _t.time()
    cached = _ALGOLIA_HEALTH["result"]
    if cached is not None and (now - _ALGOLIA_HEALTH["ts"]) < _ALGOLIA_HEALTH_TTL:
        return jsonify(dict(cached, cached=True))
    # Serialize the refresh — losers of the race return the (stale) cached value
    # instead of firing their own probe at GC's Algolia.
    if not _ALGOLIA_HEALTH_LOCK.acquire(blocking=False):
        if cached is not None:
            return jsonify(dict(cached, cached=True))
        _ALGOLIA_HEALTH_LOCK.acquire()  # cold start: wait for the first probe
        _ALGOLIA_HEALTH_LOCK.release()
        fresh = _ALGOLIA_HEALTH["result"]
        return jsonify(dict(fresh, cached=True)) if fresh is not None else (jsonify({"ok": False, "error": "unavailable"}), 503)
    try:
        # Re-check under the lock — another thread may have refreshed while we waited.
        cached = _ALGOLIA_HEALTH["result"]
        if cached is not None and (_t.time() - _ALGOLIA_HEALTH["ts"]) < _ALGOLIA_HEALTH_TTL:
            return jsonify(dict(cached, cached=True))
        return _algolia_health_probe(now)
    finally:
        _ALGOLIA_HEALTH_LOCK.release()


def _algolia_health_probe(now: float):
    result = {
        "ok": False, "nbHits": 0, "http_status": None,
        "checked_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "cached": False,
    }
    try:
        payload = {"requests": [{
            "indexName":    ALGOLIA_INDEX,
            "facetFilters": ["condition.lvl0:Used"],
            "hitsPerPage":  0,
            "page":         0,
            "query":        "",
        }]}
        r = _http.post(ALGOLIA_URL, headers=ALGOLIA_HEADERS, json=payload, timeout=15)
        result["http_status"] = r.status_code
        if r.status_code == 200:
            result["nbHits"] = (r.json().get("results") or [{}])[0].get("nbHits", 0)
            result["ok"] = result["nbHits"] > 0
    except Exception as e:
        result["error"] = type(e).__name__
    _ALGOLIA_HEALTH["ts"] = now
    _ALGOLIA_HEALTH["result"] = result
    return jsonify(result)


# ── Postgres Tier 1 read path (Phase C, v2.16.20) ─────────────────────────────
# Pure-SQL equivalent of api_browse()'s _apply_base()/facet-count/sort/paginate
# logic, for the dominant "no want-list keywords, no filter_q free text" case
# (POSTGRES_MIGRATION_PLAN.md §3, Tier 1). Shadow-mode only this session — see
# the ?pg_shadow=1/_is_admin() gate at the top of api_browse(); nothing here
# is reachable by a normal user yet. Tier 2 (keyword/free-text search active)
# is explicitly out of scope — _pg_tier1_eligible() returns False for it and
# api_browse() falls through to the existing unmodified Python path.
#
# sort_field is user-controlled (the request body), so every column reference
# below comes from a fixed whitelist (_PG_SORT_MAP / the "condition" special
# case) rather than ever being interpolated from the request directly — every
# other value in every WHERE/ORDER BY fragment is passed as a psycopg2 %(name)s
# parameter, never string-formatted into the SQL text.

# sort_field -> (sql column, lowercase-compare?). Mirrors api_browse()'s sort
# `elif` chain: "price"/"date"/"price_drop_since" get their own raw (unlowered)
# numeric/text comparison exactly like the Python elifs; everything else falls
# in the Python `else` branch, which lowercases — same here. "condition" isn't
# in this map; it gets the quality-rank CASE below, matching the Python
# elif for "condition" (v2.10.18 quality ranking, not alphabetical).
_PG_SORT_MAP = {
    "price":             ("price", False),
    "date":              ("date_listed", False),
    "price_drop_since":  ("price_drop_since", False),
    "name":              ("name", True),
    "brand":             ("brand", True),
    "category":          ("category", True),
    "subcategory":       ("subcategory", True),
    "location":          ("location", True),
    "store":             ("store", True),
    "condition_note":    ("condition_note", True),
    "url":               ("url", True),
    "image_id":          ("image_id", True),
}
_PG_TIER1_SORTABLE = set(_PG_SORT_MAP) | {"condition"}
_PG_COND_RANK_SQL = (
    "CASE condition "
    "WHEN 'Excellent' THEN 0 WHEN 'Great' THEN 1 WHEN 'Good' THEN 2 "
    "WHEN 'Fair' THEN 3 WHEN 'Poor' THEN 4 ELSE 99 END"
)
_PG_COND_KNOWN_SQL = "condition IN ('Excellent','Great','Good','Fair','Poor')"


def _pg_tier1_eligible(fq: str, has_kw: bool, sort_field: str) -> bool:
    """True iff this /api/browse request is the Tier 1 case: no filter_q free
    text, no want-list keywords active, and a sort_field the SQL path actually
    knows how to translate (the UI only ever sends the columns in
    _PG_TIER1_SORTABLE — see static/gc.js's _SORT_COLS — so an unrecognized
    value here means a client we don't recognize, not a real gap; falling
    through to the legacy path is the safe choice either way).
    f_want_only does NOT disqualify Tier 1 — see _pg_tier1_browse, it degrades
    to a correctly-empty result exactly like the JSON path would with no
    keywords active."""
    return (not fq) and (not has_kw) and (sort_field in _PG_TIER1_SORTABLE)


def _pg_tier1_browse(*, store_set, search_all, user_last_scan,
                      f_brands, f_conds, f_cats, f_subs, f_watched, wl_ids,
                      f_want_only, f_price_drop_only, f_vintage_only,
                      f_price_min, f_price_max, sort_field, sort_dir, user_sorted,
                      new_ids, fav_stores, page, per_page) -> dict:
    """Returns the exact same JSON shape api_browse() returns for a Tier 1
    request, computed entirely in Postgres (four small aggregate/paginated
    queries against the `items` table) instead of materializing and
    filtering/sorting the full catalog in Python. Raises on any DB error —
    the caller (api_browse()) catches and falls back to the legacy path."""
    import psycopg2.extras as _pg_extras

    # ── Shared WHERE fragments (parameterized; see module note above) ──────
    where = ["available"]
    params = {}

    if search_all:
        pass
    else:
        where.append("store = ANY(%(stores)s)")
        params["stores"] = list(store_set) if store_set else []

    if user_last_scan:
        where.append("(first_seen = '' OR first_seen <= %(user_last_scan)s)")
        params["user_last_scan"] = user_last_scan

    # ── Non-facet filters (api_browse()'s _apply_base(), minus fq/want_only-
    # via-keywords which can't fire in Tier 1) ──────────────────────────────
    if f_want_only:
        # No keywords active (a Tier 1 precondition) => kwMatch is False for
        # every item in the JSON path => "want-list matches only" can never
        # match anything. Reproduce that exactly rather than special-casing
        # eligibility around it.
        where.append("FALSE")
    if f_price_drop_only:
        where.append("price_drop > 0")
    if f_vintage_only:
        where.append("is_vintage")
    if f_watched:
        where.append("sku = ANY(%(wl_ids)s)")
        params["wl_ids"] = list(wl_ids)
    if f_price_min is not None:
        where.append("price >= %(price_min)s")
        params["price_min"] = f_price_min
    if f_price_max is not None:
        where.append("price <= %(price_max)s")
        params["price_max"] = f_price_max

    base_where_sql = " AND ".join(where)

    # ── Facet WHERE fragments — one per facet, each independently omittable
    # so the 4 facet-count queries below can each apply "all OTHER facets"
    # (api_browse()'s contextual-count semantics) ───────────────────────────
    def _facet_clause(col: str, values, param_key: str):
        if not values:
            return "TRUE", {}
        if col == "brand":
            real = [v for v in values if v != NO_BRAND_LABEL]
            want_no_brand = NO_BRAND_LABEL in values
            if real and want_no_brand:
                return f"(brand = ANY(%({param_key})s) OR brand = '')", {param_key: real}
            if want_no_brand and not real:
                return "brand = ''", {}
            return f"brand = ANY(%({param_key})s)", {param_key: real}
        return f"{col} = ANY(%({param_key})s)", {param_key: list(values)}

    brand_clause, brand_p = _facet_clause("brand", f_brands, "_f_brand")
    cond_clause,  cond_p  = _facet_clause("condition", f_conds, "_f_cond")
    cat_clause,   cat_p   = _facet_clause("category", f_cats, "_f_cat")
    sub_clause,   sub_p   = _facet_clause("subcategory", f_subs, "_f_sub")

    filtered_where_sql = f"{base_where_sql} AND {brand_clause} AND {cond_clause} AND {cat_clause} AND {sub_clause}"
    filtered_params = dict(params); filtered_params.update(brand_p); filtered_params.update(cond_p)
    filtered_params.update(cat_p); filtered_params.update(sub_p)

    # ── Sort clause (whitelisted columns only — see module note above) ─────
    reverse = (sort_dir == "desc")
    order_parts = []
    if sort_field == "condition":
        order_parts.append(f"{_PG_COND_KNOWN_SQL} DESC")  # unknowns always last
        order_parts.append(f"{_PG_COND_RANK_SQL} {'DESC' if reverse else 'ASC'}")
    else:
        col, lower = _PG_SORT_MAP[sort_field]
        expr = f"LOWER({col})" if lower else col
        order_parts.append(f"{expr} {'DESC' if reverse else 'ASC'}")
    if not user_sorted:
        # NEW-on-top tiering (api_browse()'s post-sort partition). No keywords
        # active in Tier 1 => kwMatch is always False => the "new+want-match"
        # tier is always empty => this reduces to exactly two tiers: NEW
        # first, then everything else, each internally in primary-sort order —
        # a single leading ORDER BY term does that.
        order_parts.insert(0, "(sku = ANY(%(new_ids)s)) DESC")
    # Final tiebreak, always: Python's list.sort() is stable (ties keep their
    # PRE-sort relative order, even with reverse=True — see the Python docs),
    # and the pre-sort order of `filtered` in the JSON path traces back to
    # _build_base_item_list()'s iteration of _cat_cache, i.e. plain ascending
    # SKU order. On a low-cardinality sort column (condition/category/store/
    # brand/etc, or even price/date with enough rows) that tie group can be
    # huge, and Postgres gives no ordering guarantee at all among ties without
    # an explicit tiebreak — found via the Phase C A/B diff harness (multiple
    # "sort X" cases returned entirely different, though equally-valid-per-key,
    # item orders). sku ASC reproduces the JSON path's tiebreak exactly.
    order_parts.append("sku ASC")
    order_sql = ", ".join(order_parts)

    offset_params = dict(filtered_params)
    offset_params["new_ids"] = list(new_ids)

    with _pg_conn() as conn:
        with conn.cursor(cursor_factory=_pg_extras.RealDictCursor) as cur:

            # Q1 — total_unfiltered + new_count: BEFORE any _apply_base filter,
            # matching api_browse()'s `all_items`/`total_unfiltered` scope
            # (store selection + per-user scan-date gating only).
            q1_where = ["available"]
            q1_params = {"new_ids": list(new_ids)}
            if not search_all:
                q1_where.append("store = ANY(%(stores)s)")
                q1_params["stores"] = list(store_set) if store_set else []
            if user_last_scan:
                q1_where.append("(first_seen = '' OR first_seen <= %(user_last_scan)s)")
                q1_params["user_last_scan"] = user_last_scan
            cur.execute(
                f"SELECT COUNT(*) AS total_unfiltered, "
                f"COUNT(*) FILTER (WHERE sku = ANY(%(new_ids)s)) AS new_count "
                f"FROM items WHERE {' AND '.join(q1_where)}", q1_params)
            row = cur.fetchone()
            total_unfiltered = row["total_unfiltered"]
            new_count = row["new_count"]

            # Q2 — facet counts: one query, one pass per facet (mirrors the
            # v2.16.12 "single pass" optimization, in SQL: 4 branches unioned
            # instead of 4 separate round trips). Each branch's WHERE includes
            # base_where + the OTHER three facets, exactly matching
            # api_browse()'s contextual-count semantics.
            facet_params = dict(params)
            facet_params.update(brand_p); facet_params.update(cond_p)
            facet_params.update(cat_p); facet_params.update(sub_p)
            facet_params["no_brand_label"] = NO_BRAND_LABEL
            facet_sql = " UNION ALL ".join([
                f"SELECT 'brand' AS facet, COALESCE(NULLIF(brand,''), %(no_brand_label)s) AS value, "
                f"COUNT(*) AS n FROM items WHERE {base_where_sql} AND {cond_clause} AND {cat_clause} "
                f"AND {sub_clause} GROUP BY value",
                f"SELECT 'condition' AS facet, condition AS value, COUNT(*) AS n FROM items "
                f"WHERE {base_where_sql} AND {brand_clause} AND {cat_clause} AND {sub_clause} "
                f"AND condition <> '' GROUP BY value",
                f"SELECT 'category' AS facet, category AS value, COUNT(*) AS n FROM items "
                f"WHERE {base_where_sql} AND {brand_clause} AND {cond_clause} AND {sub_clause} "
                f"AND category <> '' GROUP BY value",
                f"SELECT 'subcategory' AS facet, subcategory AS value, COUNT(*) AS n FROM items "
                f"WHERE {base_where_sql} AND {brand_clause} AND {cond_clause} AND {cat_clause} "
                f"AND subcategory <> '' GROUP BY value",
            ])
            cur.execute(facet_sql, facet_params)
            brand_ctx, cond_ctx, cat_ctx, sub_ctx = {}, {}, {}, {}
            _ctx_map = {"brand": brand_ctx, "condition": cond_ctx, "category": cat_ctx, "subcategory": sub_ctx}
            for r in cur.fetchall():
                _ctx_map[r["facet"]][r["value"]] = r["n"]
            # Always include currently-selected values so users can deselect them
            for b in f_brands:
                brand_ctx.setdefault(b, 0)
            for c in f_conds:
                cond_ctx.setdefault(c, 0)
            for c in f_cats:
                cat_ctx.setdefault(c, 0)
            for s in f_subs:
                sub_ctx.setdefault(s, 0)

            # Q3 — total_filtered + store_count over the fully-filtered set
            # (v2.16.21) NULLIF(store, '') before COUNT(DISTINCT ...): a plain
            # COUNT(DISTINCT store) counts an empty-string store as one more
            # distinct value, but api_browse()'s own store_count explicitly
            # excludes falsy store values (`if i.get("store")`, see the
            # legacy JSON-path store_count line below) — a real production
            # item with store == "" made this a 296-vs-297 mismatch caught by
            # spot-checking ?pg_shadow=1 against live data after deploy (the
            # 436K-row synthetic test data never had an empty-store item, so
            # the offline diff harness missed this one). COUNT(DISTINCT ...)
            # already ignores NULLs, so NULLIF(store, '') reproduces the
            # Python truthy-check exactly.
            cur.execute(
                f"SELECT COUNT(*) AS total_filtered, "
                f"COUNT(DISTINCT NULLIF(store, '')) AS store_count "
                f"FROM items WHERE {filtered_where_sql}", filtered_params)
            row = cur.fetchone()
            total_filtered = row["total_filtered"]
            store_count = row["store_count"]

            total_pages = max(1, -(-total_filtered // per_page))
            page = min(page, total_pages)
            start = (page - 1) * per_page

            # Q4 — the actual page of items
            cur.execute(
                f"SELECT sku, name, brand, category, subcategory, condition, condition_note, "
                f"price, list_price, price_drop, price_drop_since, store, location, url, "
                f"image_id, is_vintage, date_listed "
                f"FROM items WHERE {filtered_where_sql} ORDER BY {order_sql} "
                f"LIMIT %(_limit)s OFFSET %(_offset)s",
                {**offset_params, "_limit": per_page, "_offset": start})
            page_rows = cur.fetchall()

    page_items = []
    for r in page_rows:
        sku = r["sku"]
        store = r["store"] or ""
        price_raw = float(r["price"] or 0)
        page_items.append({
            "id":               sku,
            "name":             r["name"] or "",
            "brand":            r["brand"] or "",
            "price":            f"${price_raw:,.2f}" if price_raw else "",
            "price_raw":        price_raw,
            "list_price_raw":   float(r["list_price"] or 0),
            "price_drop":       float(r["price_drop"] or 0),
            "price_drop_since": r["price_drop_since"] or "",
            "store":            store,
            "location":         r["location"] or "",
            "url":              r["url"] or "",
            "category":         r["category"] or "",
            "subcategory":      r["subcategory"] or "",
            "condition":        r["condition"] or "",
            "date":             _fmt_date(r["date_listed"] or ""),
            "date_raw":         r["date_listed"] or "",
            "image_id":         r["image_id"] or "",
            "is_vintage":       bool(r["is_vintage"]),
            "condition_note":   r["condition_note"] or "",
            "watched":          sku in wl_ids,
            "isNew":            sku in new_ids,
            "kwMatch":          False,  # Tier 1 precondition: no keywords active
            "isFav":            store in fav_stores if fav_stores else False,
        })

    _cond_order = {"Excellent": 0, "Great": 1, "Good": 2, "Fair": 3, "Poor": 4}
    return {
        "items":            page_items,
        "page":             page,
        "per_page":         per_page,
        "total_count":      total_filtered,
        "total_unfiltered": total_unfiltered,
        "total_pages":      total_pages,
        "store_count":      store_count,
        "new_count":        new_count,
        "new_want_count":   0,  # no keywords active in Tier 1 => always 0
        "no_store_data":    False,
        # (v2.16.20) Secondary alphabetical tiebreak on brand name: two brands
        # with the same count used to fall back to dict-insertion order, which
        # is really "SKU order" — an accidental, unspecified tiebreak nobody
        # relied on, and one the Postgres Tier 1 path can't cheaply reproduce
        # (SQL aggregation doesn't preserve row order). Made deterministic and
        # identical in both paths instead of trying to replicate the old
        # incidental order in SQL — found via the Phase C A/B diff harness.
        "brands":        [{"name": b, "count": c} for b, c in sorted(brand_ctx.items(), key=lambda x: (-x[1], x[0]))],
        "conditions":    [{"name": c, "count": n} for c, n in sorted(cond_ctx.items(), key=lambda x: _cond_order.get(x[0], 5))],
        "categories":    [{"name": c, "count": n} for c, n in sorted(cat_ctx.items())],
        "subcategories": [{"name": s, "count": n} for s, n in sorted(sub_ctx.items())],
    }


@app.route("/api/browse", methods=["POST"])
@optional_user_context
def api_browse():
    """Return cached inventory for selected stores with server-side
    pagination, sorting, and filtering.  Sends only one page at a time
    so the browser never has to hold 80K items in memory."""
    data = request.json or {}
    stores = data.get("stores", [])
    search_all = bool(data.get("all_stores"))
    if not stores and not search_all:
        return jsonify({"items": [], "no_store_data": True})

    # Pagination params
    page     = max(int(data.get("page", 1)), 1)
    per_page = min(max(int(data.get("per_page", 50)), 10), 200)

    # Sort params  (defaults: date descending = newest first)
    sort_field = data.get("sort_field", "date")
    sort_dir   = data.get("sort_dir", "desc")
    fav_stores = set(data.get("fav_stores", []))

    # Per-user filtering: only show items first_seen ≤ this user's last scan time
    user_last_scan = (data.get("user_last_scan") or "").strip()

    # Filter params — all dropdowns are multi-select arrays
    # Cap query length: filter_q compiles to regex (incl. '*' → '.*') that runs over
    # every cached item. An unbounded query with many wildcards is an unauthenticated
    # CPU DoS over the ~92K-item cache, so clamp it to a sane length.
    fq       = (data.get("filter_q") or "").lower().strip()[:200]
    f_brands = data.get("filter_brands") or []
    f_conds  = data.get("filter_conditions") or []
    f_cats   = data.get("filter_categories") or []
    f_subs   = data.get("filter_subcategories") or []
    f_watched = bool(data.get("filter_watched"))
    f_want_only = bool(data.get("filter_want_list_only"))
    f_price_drop_only = bool(data.get("filter_price_drop_only"))
    f_vintage_only = bool(data.get("vintage_only"))
    f_strict = bool(data.get("filter_strict"))
    def _to_float(v):
        try: return float(v) if v is not None and v != '' else None
        except (TypeError, ValueError): return None
    f_price_min = _to_float(data.get("filter_price_min"))
    f_price_max = _to_float(data.get("filter_price_max"))
    force_fav_sort = bool(data.get("force_fav_sort"))
    # (moved up from the sort block below, v2.16.20 — Tier 1 shadow read path
    # needs this before the sort block runs; reading it here changes nothing
    # for the legacy JSON path, it's just an earlier data.get() call)
    user_sorted = bool(data.get("user_sorted"))

    # ── Postgres Tier 1 shadow read path (Phase C, v2.16.20) ───────────────
    # Admin/dev-only, gated behind ?pg_shadow=1 — never reachable by normal
    # users this session. See POSTGRES_MIGRATION_PLAN.md §3/§7 Phase C and
    # NEXT_SESSION_PROMPT.md. Eligibility (no want-list keywords, no filter_q
    # free text, a whitelisted sort_field) is re-checked below once fq/_has_kw
    # are known — this early flag only decides whether it's even worth trying.
    _pg_shadow_requested = request.args.get("pg_shadow") == "1"

    _load_cat_cache()
    # Watchlist and keywords now come from the client (localStorage)
    wl_ids     = set(data.get("watchlist_ids", []))
    keywords   = data.get("keywords", [])
    # Want-list keywords are matched against every cached item (see _kw_match below).
    # An unbounded list is a CPU-DoS over the ~92K-item cache, but the old flat
    # keywords[:50] silently dropped matches for real power users — guitar want
    # lists of 70-220+ terms are common. Guard in two steps:
    #   1. Dedupe case-insensitively (cheap; kills redundant regexes from re-adds) and
    #      clamp each term to 100 chars.
    #   2. Cap by accountability: logged-in users own large, cross-device-synced want
    #      lists and are account-rate-limitable, so they get generous headroom;
    #      anonymous callers are the real DoS vector (their list is localStorage-only),
    #      so they stay tighter.
    # The matcher below keeps plain single-word terms ~free regardless of count, so the
    # cap mainly bounds the rarer phrase/wildcard terms.
    if isinstance(keywords, list):
        _seen = set()
        _dedup = []
        for k in keywords:
            ks = str(k)[:100]
            kl = ks.strip().lower()
            if kl and kl not in _seen:
                _seen.add(kl)
                _dedup.append(ks)
        _kw_cap = 750 if session.get("user_id") else 250
        keywords = _dedup[:_kw_cap]
    else:
        keywords = []
    new_ids    = set(data.get("new_ids", []))
    store_set  = set(stores) if not search_all else None

    # ── Unified query matching (shared by keyword list and filter_q) ─────────
    import re as _re

    # Query helpers (_compile_query, _matches_all, _matches_any, _kw_or_pattern,
    # _SIMPLE_KW_RE, _KW_SPLIT_RE, _wl_bool_compile, _compile_fq_clauses,
    # _expand_colon_prefix) are module-level since v2.16.0 — shared with
    # /api/saved-search-counts so the two endpoints can't drift. See "Shared
    # query-matching helpers".

    # ── Want-list keyword matcher (fast path) ────────────────────────────────
    # Matching each keyword as its own \bword\b regex is O(items × keywords); at ~92K
    # items a 200-term want list cost seconds per browse (the cause of the >50-term
    # "missing matches" bug — the old fix just truncated the list). Bucket instead:
    #   • plain single word (^\w+$)  -> lowercased SET, tested via token membership once
    #     per item — O(items), independent of how many such keywords there are.
    #   • phrase / quoted / wildcard -> folded into ONE alternation regex, searched once
    #     per item instead of once per keyword.
    #   • comma-AND ("a, b")         -> rare; kept on the per-keyword term-list path.
    # '=' is the legacy strict marker (whole-word is the default now), stripped first.
    #
    # Cost is NOT uniform across term types, so the DoS guard is per-type, not a flat
    # cap (a flat cap on "complex" terms would drop real multi-word want-list entries
    # like "Big Muff" / "OD-1" once a user had >N of them — re-introducing the >50-term
    # bug this whole rewrite exists to fix). Benchmarked over the full ~92K cache:
    #   • plain single word (^\w+$)  -> lowercased SET, token-membership once per item.
    #     O(items) regardless of count. Stays at the generous _kw_cap (750/250).
    #   • ordinary phrase / hyphenated ("Big Muff", "OD-1") -> exact \b...\b regex, but
    #     gated by a cheap SOUND pre-filter: all of the phrase's word-tokens must be
    #     present in the item before the regex runs. Since few items contain every word
    #     of a given phrase, the regex almost never fires → ~free for real lists
    #     (a realistic 220-term list of 160 words + 60 phrases benches ~0.35s). Given a
    #     generous cap (300) purely as a backstop; real want lists are far below it.
    #   • wildcard (*) / quoted-exact -> the genuinely expensive paths ('.*' can't be
    #     pre-filtered; quoted-exact is a raw substring). Folded into one alternation and
    #     capped HARD at 30 — advanced syntax a real want list uses a handful of, but 250
    #     wildcards alone cost ~8s/browse, GIL-held, on the public unbounded /api/browse
    #     (unauthenticated CPU-DoS). (v2.13.1)
    #   • comma-AND ("a, b") -> a whole-word subterm matches iff its token is present, so
    #     it takes the SAME cheap required-tokens pre-filter as phrases and gets its own
    #     generous cap — no longer starved by the wildcard/quoted cap, which silently
    #     dropped comma terms once a want list had >30 combined exotic terms. (v2.14.4)
    _PHRASE_KW_CAP = 300
    _EXOTIC_KW_CAP = 30    # wildcard (*) + quoted-exact only — the expensive alternation
    _AND_KW_CAP    = 200   # comma-AND — cheap once sound-prefiltered, so its own cap
    _BOOL_KW_CAP   = 50    # v2.16.0 ';'/'-' entries — benched ~0.8s worst-adversarial
                           # at 50 over the full cache, in line with the v2.13.1 budget
    _phrase_n = 0
    _exotic_n = 0
    _and_n    = 0
    _bool_n   = 0
    _kw_word_set = set()
    _kw_phrases  = []   # (compiled \b...\b regex, set-of-required-word-tokens) — pre-filtered
    _kw_or_pats  = []   # wildcard + quoted-exact patterns -> one alternation (capped)
    _kw_and      = []   # comma-AND: (term-list, set-of-required-plain-word-tokens) — pre-filtered
    _kw_bool     = []   # v2.16.0 OR/NOT entries: list of (pos, neg, req) clause lists
    for _kw in keywords:
        _base = _kw.lstrip('=').strip()
        if not _base:
            continue
        # v2.16.3: expand 'prefix : b1; b2; ...' BEFORE any routing decision, so a
        # colon-prefix entry with a single branch and no dash (e.g. 'Mesa: Angel')
        # correctly falls through to the plain comma-AND path below instead of the
        # dead literal-string branch it'd hit if expansion ran after routing.
        _base = _expand_colon_prefix(_base, join=', ')
        # v2.16.0 OR/NOT syntax — must be checked FIRST: an entry like 'OD*, -combo'
        # or 'foo*; bar' contains '*' and would otherwise be misrouted to the exotic
        # branch and lose its operators. _wl_bool_compile returns None for every
        # legacy entry (no ';' and no leading-dash part), so old lists are untouched.
        _wl_clauses = _wl_bool_compile(_base) if (';' in _base or '-' in _base) else None
        if _wl_clauses is not None:
            if '*' in _base:
                # Wildcards can't be pre-filtered — charge them to the tight exotic
                # budget so the bool path can't smuggle 100 wildcard scans past the
                # 30-cap (the v2.13.1 DoS boundary).
                if _exotic_n < _EXOTIC_KW_CAP:
                    _kw_bool.append(_wl_clauses)
                    _exotic_n += 1
            elif _bool_n < _BOOL_KW_CAP:
                _kw_bool.append(_wl_clauses)
                _bool_n += 1
        elif _SIMPLE_KW_RE.match(_base):
            _kw_word_set.add(_base.lower())
        elif '*' in _base or (_base.startswith('"') and _base.endswith('"') and len(_base) > 2):
            if _exotic_n < _EXOTIC_KW_CAP:
                _p = _kw_or_pattern(_base)
                if _p:
                    _kw_or_pats.append(_p)
                    _exotic_n += 1
        elif ',' in _base:
            if _and_n < _AND_KW_CAP:
                _t = _compile_query(_base)
                if _t:
                    # Required tokens = words of the plain (non-wildcard/quoted) subterms;
                    # all must be present before the per-term regexes run (cheap gate).
                    _req = set()
                    for _part in _base.split(','):
                        _pp = _part.strip().lower()
                        if _pp and '*' not in _pp and not (_pp.startswith('"') and _pp.endswith('"') and len(_pp) > 2):
                            _req |= set(_KW_SPLIT_RE.split(_pp)) - {''}
                    _kw_and.append((_t, _req))
                    _and_n += 1
        else:
            # Ordinary phrase or punctuated single term -> exact whole-word match, but
            # cheap to skip via the required-tokens subset test below.
            if _phrase_n < _PHRASE_KW_CAP:
                _words = set(_KW_SPLIT_RE.split(_base.lower())) - {''}
                _kw_phrases.append((_re.compile(r'\b' + _re.escape(_base) + r'\b', _re.IGNORECASE), _words))
                _phrase_n += 1
    _kw_combo = _re.compile('|'.join(_kw_or_pats), _re.IGNORECASE) if _kw_or_pats else None
    _has_kw = bool(_kw_word_set or _kw_phrases or _kw_combo or _kw_and or _kw_bool)

    def _kw_match(name_l, brand_l):
        text = name_l + " " + brand_l
        _toks = set(_KW_SPLIT_RE.split(text)) if (_kw_word_set or _kw_phrases or _kw_and or _kw_bool) else None
        if _kw_word_set and not _kw_word_set.isdisjoint(_toks):
            return True
        if _kw_phrases:
            for _rx, _words in _kw_phrases:
                # Sound pre-filter: a whole-word phrase can only match if every one of
                # its word-tokens is present; the regex confirms order/adjacency/exact
                # separators, so this stays behavior-identical to the per-keyword path.
                if _words <= _toks and _rx.search(text):
                    return True
        if _kw_combo is not None and _kw_combo.search(text):
            return True
        if _kw_and:
            # Same sound pre-filter as phrases: all plain-word subterm tokens must be
            # present before running the per-term regexes (behavior-identical, just cheap).
            for _terms, _req in _kw_and:
                if _req <= _toks and _matches_all(text, _terms):
                    return True
        if _kw_bool:
            # v2.16.0 OR/NOT entries: entry matches if ANY of its clauses does —
            # all positives present (required-tokens pre-filtered) and no negative.
            for _clauses in _kw_bool:
                for _pos, _neg, _req in _clauses:
                    if _req <= _toks and _matches_all(text, _pos) and not (_neg and _matches_any(text, _neg)):
                        return True
        return False

    # Check if any cache entries have store field
    has_store_data = any(v.get("store") for v in _cat_cache.values())
    if not has_store_data:
        return jsonify({"items": [], "no_store_data": True,
                        "message": "Run 'Check for New Items' once to populate store data."})

    # ── Postgres Tier 1 shadow read path, continued (Phase C, v2.16.20) ────
    # Only reachable with ?pg_shadow=1 AND an admin session AND the request
    # actually qualifies as Tier 1 (no want-list keywords, no filter_q free
    # text, a whitelisted sort_field — see _pg_tier1_eligible). Every real
    # user, and every admin request without the flag, falls straight through
    # to the unchanged legacy JSON path below — this branch never affects
    # what gets served to normal traffic this session (Phase D is the real
    # cutover, not this one).
    if (_pg_shadow_requested and _is_admin() and _PG_POOL is not None
            and _pg_tier1_eligible(fq, _has_kw, sort_field)):
        try:
            _pg_t0 = time.time()
            _pg_result = _pg_tier1_browse(
                store_set=store_set, search_all=search_all,
                user_last_scan=user_last_scan,
                f_brands=f_brands, f_conds=f_conds, f_cats=f_cats, f_subs=f_subs,
                f_watched=f_watched, wl_ids=wl_ids, f_want_only=f_want_only,
                f_price_drop_only=f_price_drop_only, f_vintage_only=f_vintage_only,
                f_price_min=f_price_min, f_price_max=f_price_max,
                sort_field=sort_field, sort_dir=sort_dir, user_sorted=user_sorted,
                new_ids=new_ids, fav_stores=fav_stores, page=page, per_page=per_page,
            )
            _pg_result["_pg_shadow"] = True
            _pg_result["_pg_shadow_ms"] = round((time.time() - _pg_t0) * 1000, 1)
            return jsonify(_pg_result)
        except Exception as e:
            # Never let a Postgres/SQL bug affect a real response — this whole
            # branch is diagnostic. Log and fall through to the legacy path
            # below exactly as if pg_shadow had never been passed.
            print(f"[pg] tier1 shadow browse failed, falling back to JSON path: "
                  f"{type(e).__name__}: {e}")

    # ── Build full item list for selected stores (lightweight dicts) ──────
    # E2 (v2.16.4): the expensive per-item projection (price/date formatting,
    # name/brand lowercasing) now happens ONCE per cache generation in
    # _build_base_item_list(), memoized by cache mtime — see that function's
    # docstring. This pass is just the genuinely per-request part: store
    # selection, per-user scan-date gating, and the per-user annotations
    # (watched/isNew/kwMatch/isFav) that can't be cached across users.
    all_items = []
    for b in _build_base_item_list():
        if store_set is not None and b["store"] not in store_set:
            continue
        # Per-user browse gating: hide items first seen after this user's last scan.
        # Ensures each device only sees inventory that existed when IT scanned —
        # items added to cache by another device's newer scan stay hidden here.
        if user_last_scan:
            first_seen = b["first_seen"]
            if first_seen and first_seen > user_last_scan:
                continue
        sku   = b["id"]
        store = b["store"]
        kw_hit = _kw_match(b["name_lower"], b["brand_lower"]) if _has_kw else False
        all_items.append({
            "id":               sku,
            "name":             b["name"],
            "brand":            b["brand"],
            "price":            b["price"],
            "price_raw":        b["price_raw"],
            "list_price_raw":   b["list_price_raw"],
            "price_drop":       b["price_drop"],
            "price_drop_since": b["price_drop_since"],
            "store":            store,
            "location":         b["location"],
            "url":              b["url"],
            "category":         b["category"],
            "subcategory":      b["subcategory"],
            "condition":        b["condition"],
            "date":             b["date"],
            "date_raw":         b["date_raw"],
            "image_id":         b["image_id"],
            "is_vintage":       b["is_vintage"],
            "condition_note":   b["condition_note"],
            "watched":          sku in wl_ids,
            "isNew":            sku in new_ids,
            "kwMatch":          kw_hit,
            "isFav":            store in fav_stores if fav_stores else False,
        })

    total_unfiltered = len(all_items)
    new_count_unfiltered = sum(1 for i in all_items if i.get("isNew"))

    # ── Contextual facet counts ───────────────────────────────────────────
    # Step 1: apply non-facet filters (text search, want list, price drops, watched)
    def _apply_base(items):
        r = items
        if fq:
            # v2.16.0: ';' splits OR clauses; within a clause space-separated tokens
            # are AND'd (old behavior, byte-identical when no ';' or '-' present) and
            # a leading '-' negates a token. Token budget stays 12 TOTAL across all
            # clauses + ≤4 clauses — same unauthenticated-DoS ceiling as v2.13.0.
            fq_clauses = _compile_fq_clauses(fq, fuzzy=f_strict)
            r = [i for i in r if fq_clauses and _fq_text_match(
                ((i["name"] or "") + " " + (i["brand"] or "")).lower(), fq_clauses)]
        if f_want_only:       r = [i for i in r if i["kwMatch"]]
        if f_price_drop_only: r = [i for i in r if i.get("price_drop", 0) > 0]
        if f_vintage_only:    r = [i for i in r if i.get("is_vintage")]
        if f_watched:         r = [i for i in r if i["watched"]]
        if f_price_min is not None:
            r = [i for i in r if (i.get("price_raw") or 0) >= f_price_min]
        if f_price_max is not None:
            r = [i for i in r if (i.get("price_raw") or 0) <= f_price_max]
        return r

    base_items = _apply_base(all_items)

    # Brandless items are stored with brand == "". Surface them under the synthetic
    # NO_BRAND_LABEL facet entry so they become filterable — otherwise they're
    # unreachable, since the facet builder skips empties and the filter is membership.
    def _brand_ok(b, bs):
        return (b in bs) if b else (NO_BRAND_LABEL in bs)

    # Step 2: for each facet, count items passing all OTHER facet filters.
    # (v2.16.12) Single linear pass over base_items computing all four facets'
    # counts together, instead of 4 separate calls each doing its own O(N)
    # list-comprehension filter + loop. Measured via cProfile against the real
    # ~92K-item cache: the old 4-pass version was ~187ms cumulative per
    # /api/browse call — the single biggest per-request cost on this endpoint,
    # well above the memoized base-list build (E2, v2.16.4). Semantics are
    # unchanged: each facet's count still reflects all OTHER active facet
    # filters, just computed together instead of via 4 separate filtered
    # sub-lists.
    _bset, _cset, _catset, _subset = set(f_brands), set(f_conds), set(f_cats), set(f_subs)
    brand_ctx, cond_ctx, cat_ctx, sub_ctx = {}, {}, {}, {}
    for i in base_items:
        _brand, _cond, _cat, _sub = i["brand"], i["condition"], i["category"], i["subcategory"]
        _brand_pass = (not _bset) or _brand_ok(_brand, _bset)
        _cond_pass  = (not _cset) or (_cond in _cset)
        _cat_pass   = (not _catset) or (_cat in _catset)
        _sub_pass   = (not _subset) or (_sub in _subset)
        if _cond_pass and _cat_pass and _sub_pass:
            v = _brand or NO_BRAND_LABEL
            brand_ctx[v] = brand_ctx.get(v, 0) + 1
        if _brand_pass and _cat_pass and _sub_pass and _cond:
            cond_ctx[_cond] = cond_ctx.get(_cond, 0) + 1
        if _brand_pass and _cond_pass and _sub_pass and _cat:
            cat_ctx[_cat] = cat_ctx.get(_cat, 0) + 1
        if _brand_pass and _cond_pass and _cat_pass and _sub:
            sub_ctx[_sub] = sub_ctx.get(_sub, 0) + 1
    _cond_order = {"Excellent": 0, "Great": 1, "Good": 2, "Fair": 3, "Poor": 4}

    # Always include currently-selected values so users can deselect them
    for b in f_brands:
        if b not in brand_ctx: brand_ctx[b] = 0
    for c in f_conds:
        if c not in cond_ctx: cond_ctx[c] = 0
    for c in f_cats:
        if c not in cat_ctx: cat_ctx[c] = 0
    for s in f_subs:
        if s not in sub_ctx: sub_ctx[s] = 0

    # ── Apply filters ─────────────────────────────────────────────────────
    filtered = base_items
    if f_brands:
        bs = set(f_brands); filtered = [i for i in filtered if _brand_ok(i["brand"], bs)]
    if f_conds:
        cs = set(f_conds); filtered = [i for i in filtered if i["condition"] in cs]
    if f_cats:
        cs = set(f_cats); filtered = [i for i in filtered if i["category"] in cs]
    if f_subs:
        ss = set(f_subs); filtered = [i for i in filtered if i["subcategory"] in ss]

    # ── Sort ──────────────────────────────────────────────────────────────
    # NEW items float to top ONLY on default sort (no explicit column click)
    # When user explicitly clicks a sort column, sort purely by that column
    reverse = (sort_dir == "desc")
    # user_sorted is read earlier now (v2.16.20, see the Tier 1 shadow block above)

    # (v2.16.20) Deterministic tiebreak: pre-sort by sku ascending before the
    # primary sort below. list.sort() is stable, so this becomes the tiebreak
    # for equal-key items in every branch below, regardless of the requested
    # direction — Python's reverse=True preserves stable tie order rather than
    # reversing it, so a plain (primary, sku) tuple key would have reversed
    # the tiebreak too on a "desc" request; this two-pass approach keeps the
    # tiebreak always-ascending-by-sku no matter which way the primary sort
    # runs. Previously ties fell back to whatever order `filtered` already
    # had — i.e. _cat_cache's arbitrary historical insertion order, which
    # Postgres has no equivalent of and can't reproduce. Made explicit here
    # and matched exactly by the new Tier 1 SQL path's own trailing
    # "ORDER BY ..., sku ASC" (see _pg_tier1_browse) — found via the Phase C
    # A/B diff harness (several "sort X" cases returned entirely different,
    # though equally-valid-per-key, item orders on tied values).
    filtered.sort(key=lambda x: x["id"])

    if sort_field == "price":
        filtered.sort(key=lambda x: x.get("price_raw") or 0, reverse=reverse)
    elif sort_field == "date":
        filtered.sort(key=lambda x: x.get("date_raw") or "", reverse=reverse)
    elif sort_field == "price_drop_since":
        filtered.sort(key=lambda x: x.get("price_drop_since") or "", reverse=reverse)
    elif sort_field == "condition":
        # Quality ranking (v2.10.18), not alphabetical: best → worst when ascending
        # (Excellent→Great→Good→Fair→Poor), reverse on descending. Alphabetical
        # order had no meaning for users (and put Poor near the top because 'P' > 'G').
        # First click on the column = 'asc' (see JS line ~8389: non-date fields default
        # to +1 on first click), so users see Excellent-first by default.
        # Unknown/blank conditions sort to the end either way.
        _unknown_rank = 99 if not reverse else -1
        filtered.sort(
            key=lambda x: _cond_order.get(x.get("condition") or "", _unknown_rank),
            reverse=reverse,
        )
    else:
        filtered.sort(key=lambda x: (x.get(sort_field) or "").lower(), reverse=reverse)

    # Only apply NEW-on-top tier for the default (non-user-clicked) sort
    # Three tiers: new+want-list match → new only → everything else.
    # (v2.16.12) A single linear partition instead of a second full O(N log N)
    # sort — `filtered` is already correctly ordered by the primary sort above,
    # and both list.sort() and this partition are stable, so concatenating the
    # three tiers in original relative order is equivalent to re-sorting by
    # priority, just O(N) instead of O(N log N) on top of the first sort.
    # Measured (cProfile, real ~92K-item cache): the two full-list sorts
    # combined cost ~100ms/call on the common no-filter browse call.
    if not user_sorted:
        _new_want, _new_only, _rest = [], [], []
        for x in filtered:
            if x.get("isNew"):
                (_new_want if x.get("kwMatch") else _new_only).append(x)
            else:
                _rest.append(x)
        filtered = _new_want + _new_only + _rest

    total_filtered = len(filtered)
    total_pages    = max(1, -(-total_filtered // per_page))  # ceil division
    page           = min(page, total_pages)
    start          = (page - 1) * per_page
    page_items     = filtered[start:start + per_page]

    # Count items that are both want-list matches AND new (for the status notification)
    new_want_count = sum(1 for i in filtered if i.get("isNew") and i.get("kwMatch"))
    # Unique stores in filtered result set
    store_count = len(set(i.get("store", "") for i in filtered if i.get("store")))

    return jsonify({
        "items":            page_items,
        "page":             page,
        "per_page":         per_page,
        "total_count":      total_filtered,
        "total_unfiltered": total_unfiltered,
        "total_pages":      total_pages,
        "store_count":      store_count,
        "new_count":        new_count_unfiltered,
        "new_want_count":   new_want_count,
        "no_store_data":    False,
        # Contextual facet counts — each facet reflects all OTHER active filters
        # (v2.16.20) Secondary alphabetical tiebreak on brand name: two brands
        # with the same count used to fall back to dict-insertion order, which
        # is really "SKU order" — an accidental, unspecified tiebreak nobody
        # relied on, and one the Postgres Tier 1 path can't cheaply reproduce
        # (SQL aggregation doesn't preserve row order). Made deterministic and
        # identical in both paths instead of trying to replicate the old
        # incidental order in SQL — found via the Phase C A/B diff harness.
        "brands":        [{"name": b, "count": c} for b, c in sorted(brand_ctx.items(), key=lambda x: (-x[1], x[0]))],
        "conditions":    [{"name": c, "count": n} for c, n in sorted(cond_ctx.items(), key=lambda x: _cond_order.get(x[0], 5))],
        "categories":    [{"name": c, "count": n} for c, n in sorted(cat_ctx.items())],
        "subcategories": [{"name": s, "count": n} for s, n in sorted(sub_ctx.items())],
    })


# (Dead legacy global-file endpoints removed in v2.14.5 — /api/watchlist GET+POST,
# /api/watchlist/items, /api/keywords GET+POST. Per-user watch/want lists live in
# SQLite via /api/sync; no frontend called any of these (grepped all static/*.js).
# Login-gated since v2.12.32, deleted per 2026-07 audit E3. load_watchlist/
# save_watchlist helpers stay — the scan loop still updates the legacy global
# watchlist file's sold flags.)


@app.route("/api/state")
@optional_user_context
def api_state():
    _load_cat_cache()
    total_items = sum(1 for v in _cat_cache.values() if v.get("available", True))
    last_scan_file = DATA_DIR / "gc_last_scan.txt"
    last_scan = last_scan_file.read_text().strip() if last_scan_file.exists() else None
    return jsonify({
        "total_items":  total_items,
        "excel_exists": OUTPUT_FILE.exists(),
        "is_first_run": total_items == 0,
        "last_scan":    last_scan,
    })

@app.route("/api/run", methods=["POST"])
@optional_user_context
def api_run():
    global _current_run_time
    if not _lock.acquire(blocking=False):
        # A scan is already running — subscribe this client to it instead of rejecting
        joined_id = _current_run_id
        joined_time = _current_run_time
        if joined_id and _subscribe_to_run(joined_id) is not None:
            return jsonify({"status": "joined", "run_id": joined_id, "run_time": joined_time})
        # Race: scan just finished between the lock check and subscribe — tell client to retry
        return jsonify({"error": "Scan just finished, please try again."}), 409
    # Rate-limit unauthenticated scan triggers to prevent Algolia quota exhaustion.
    # Logged-in users (user_id in session) are exempt — they have an account to hold accountable.
    user_id = session.get("user_id")
    if not user_id:
        ip  = _client_ip()
        now = time.time()
        if now - _scan_last.get(ip, 0) < _SCAN_COOLDOWN:
            _lock.release()
            return jsonify({"error": "Please wait a moment before starting another scan."}), 429
        _scan_last[ip] = now
    _stop_event.clear()
    data     = request.json
    # Cap the stores array — each entry is an Algolia fan-out; there are only ~240
    # real stores, so anything beyond that is garbage or abuse. (2026-07 audit L2)
    selected = (data.get("stores") or [])[:300]
    baseline = data.get("baseline", False)
    # If user is logged in, use their server-stored last_run (synced across devices).
    # Otherwise fall back to the device's own localStorage value.
    if user_id:
        _udata = _get_user_data(user_id)
        device_last_run    = _udata.get("last_run", "")
        device_last_anchor = _udata.get("last_anchor", "")
    else:
        device_last_run    = (data.get("device_last_run")    or "").strip()
        device_last_anchor = (data.get("device_last_anchor") or "").strip()
    # Compute run_time here so we can return it to the client immediately.
    run_time_now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    _current_run_time = run_time_now
    run_id, run_q = _create_run_queue()
    # Also mirror to legacy _q for any other endpoints that read it
    while not _q.empty():
        try: _q.get_nowait()
        except queue.Empty: break
    t = threading.Thread(
        target=_run,
        args=(selected, baseline, run_id, device_last_run, run_time_now, device_last_anchor, user_id),
        daemon=True,
    )
    t.start()
    return jsonify({"status": "started", "run_id": run_id, "run_time": run_time_now})

@app.route("/api/set-cookies", methods=["POST"])
@optional_user_context
def api_set_cookies():
    """Import browser cookies into the HTTP session."""
    denied = _require_admin_api()
    if denied:
        return denied
    cookie_string = request.json.get("cookies", "")
    count = 0
    for part in cookie_string.split(";"):
        part = part.strip()
        if "=" in part:
            k, v = part.split("=", 1)
            _http.cookies.set(k.strip(), v.strip(), domain=".guitarcenter.com")
            count += 1
    _save_cookies()
    return jsonify({"imported": count, "status": "Cookies set — try running a store now."})



@app.route("/api/export-data")
@optional_user_context
def api_export_data():
    """Export all data files as a JSON bundle for migration."""
    denied = _require_admin_api()
    if denied:
        return denied
    bundle = {}
    for name, path in [
        ("state",     STATE_FILE),
        ("cat_cache", CAT_CACHE_FILE),
        ("stores",    STORES_CACHE),
        ("favorites", FAVORITES_FILE),
        ("watchlist", WATCHLIST_FILE),
        ("keywords",  KEYWORDS_FILE),
    ]:
        if path.exists():
            try:
                bundle[name] = json.loads(path.read_text())
            except Exception:
                pass
    from flask import Response
    return Response(
        json.dumps(bundle),
        mimetype="application/json",
        headers={"Content-Disposition": "attachment; filename=gc_data_export.json"}
    )


@app.route("/api/import-data", methods=["POST"])
@optional_user_context
def api_import_data():
    """Import a data bundle exported from another instance. Requires admin session."""
    denied = _require_admin_api()
    if denied:
        return denied
    bundle = request.json or {}
    written = []
    mapping = {
        "state":     STATE_FILE,
        "cat_cache": CAT_CACHE_FILE,
        "stores":    STORES_CACHE,
        "favorites": FAVORITES_FILE,
        "watchlist": WATCHLIST_FILE,
        "keywords":  KEYWORDS_FILE,
    }
    for name, path in mapping.items():
        if name in bundle:
            path.write_text(json.dumps(bundle[name]))
            written.append(name)
    global _cat_cache
    _cat_cache = {}
    _load_cat_cache()
    return jsonify({"imported": written, "status": "Import complete — reload the page."})


@app.route("/api/cl-search")
@optional_user_context
def api_cl_search():
    # Require login. Each call fans out to up to ~75 Craigslist markets (10 concurrent,
    # 12s timeouts each), so leaving it open was an unauthenticated outbound-request
    # amplification / resource-abuse vector. The CL feature is sign-in-only by design.
    if not session.get("user_id"):
        return jsonify({"error": "Not logged in."}), 401
    q = request.args.get("q", "").strip()[:200]
    cities_param = request.args.get("cities", "").strip()
    if not q:
        return jsonify({"error": "No search term provided."})
    _valid_cities = set(_CL_CITIES)
    cities = [c.strip() for c in cities_param.split(",") if c.strip() in _valid_cities] if cities_param else []
    title_only = request.args.get("title_only", "").lower() in ("1", "true", "yes")
    try:
        results = _cl_search(q, cities or None, title_only=title_only)
        return jsonify({"results": results, "count": len(results)})
    except Exception:
        # Don't leak internal exception text to the caller.
        return jsonify({"error": "Search failed. Please try again."})


_CL_CITIES = [
    "atlanta","austin","boston","chicago","dallas","denver","detroit",
    "houston","lasvegas","losangeles","miami","minneapolis","nashville",
    "newyork","philadelphia","phoenix","portland","raleigh","sacramento",
    "saltlakecity","sanantonio","sandiego","sfbay","seattle","stlouis",
    "washingtondc","baltimore","charlotte","cleveland","columbus","fortworth",
    "indianapolis","jacksonville","kansascity","memphis","milwaukee",
    "oklahomacity","orlando","pittsburgh","richmond","riverside","tampabay",
    "tucson","tulsa","virginiabeach","albuquerque","boise","buffalo",
    "cincinnati","desmoines","elpaso","fresno","grandrapids","greensboro",
    "hartford","honolulu","knoxville","louisville","madison","neworleans",
    "norfolk","omaha","providence","rochester","spokane","syracuse",
    "toledo","wichita",
]

_CL_LABELS = {
    "sfbay":"SF Bay Area","newyork":"New York","losangeles":"Los Angeles",
    "washingtondc":"Washington DC","saltlakecity":"Salt Lake City",
    "sandiego":"San Diego","sanantonio":"San Antonio","lasvegas":"Las Vegas",
    "tampabay":"Tampa Bay","kansascity":"Kansas City","grandrapids":"Grand Rapids",
    "desmoines":"Des Moines","fortworth":"Fort Worth","oklahomacity":"Oklahoma City",
    "virginiabeach":"Virginia Beach","neworleans":"New Orleans","stlouis":"St. Louis",
}

def _cl_city_label(city_id: str) -> str:
    return _CL_LABELS.get(city_id, city_id.title())

def _cl_fmt_date(iso: str) -> str:
    try:
        from datetime import datetime as dt
        d = dt.fromisoformat(iso.replace("Z",""))
        return f"{d.month}/{d.day}/{str(d.year)[2:]}"
    except Exception:
        return iso[:10] if iso else ""

def _cl_slugify(text: str) -> str:
    """Convert a title to a CL-style URL slug for matching.
    E.g. 'Fender Telecaster 2019 MIM' → 'fender-telecaster-2019-mim'"""
    s = text.lower().strip()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return s.strip('-')

def _cl_parse_html(html: str, city_id: str) -> list[dict]:
    """Parse CL search results — ItemList JSON-LD + URLs from HTML anchor tags."""
    items = []
    label = _cl_city_label(city_id)

    # Extract post URLs from the HTML — CL puts them in <a class="cl-app-anchor"> or similar
    # Pattern: href="https://cityname.craigslist.org/msa/d/title/1234567890.html"
    post_urls = re.findall(
        r'href="(https?://[a-z]+\.craigslist\.org/[^"]+/d/[^"]+\.html)"',
        html)
    # Dedupe while preserving order
    seen = set()
    post_urls_ordered = []
    for u in post_urls:
        if u not in seen:
            seen.add(u)
            post_urls_ordered.append(u)

    # Build a slug→URL lookup for title-based matching (replaces fragile position-based matching).
    # CL post URLs contain a slugified title: /msa/d/fender-telecaster-2019/1234567890.html
    # We extract the slug and match it against JSON-LD item names.
    _slug_to_urls: dict[str, list[str]] = {}  # slug → [url, ...] (multiple posts can have similar slugs)
    for u in post_urls_ordered:
        try:
            # Extract slug from URL path: /section/d/SLUG/ID.html
            path_parts = u.split('/')
            d_idx = path_parts.index('d') if 'd' in path_parts else -1
            if d_idx >= 0 and d_idx + 1 < len(path_parts):
                slug = path_parts[d_idx + 1]
                if slug not in _slug_to_urls:
                    _slug_to_urls[slug] = []
                _slug_to_urls[slug].append(u)
        except Exception:
            pass

    def _match_url_by_title(name: str) -> str:
        """Find the best matching URL for a JSON-LD item name by comparing title slugs."""
        name_slug = _cl_slugify(name)
        if not name_slug:
            return ""
        name_words = set(name_slug.split('-'))
        best_url = ""
        best_score = 0
        for slug, urls in _slug_to_urls.items():
            if not urls:
                continue
            slug_words = set(slug.split('-'))
            # Score = number of overlapping words (Jaccard-like)
            overlap = len(name_words & slug_words)
            # Require at least 2 word matches to avoid false positives
            if overlap > best_score and overlap >= 2:
                best_score = overlap
                best_url = urls[0]
        # If we found a match, remove the URL from the pool so it can't be reused
        if best_url:
            for slug, urls in _slug_to_urls.items():
                if best_url in urls:
                    urls.remove(best_url)
                    break
        return best_url

    # Find the ItemList JSON-LD block
    for block in re.findall(
            r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html, re.DOTALL):
        try:
            data = json.loads(block)
        except Exception:
            continue
        if not isinstance(data, dict) or data.get("@type") != "ItemList":
            continue

        entries = data.get("itemListElement", [])
        for i, entry in enumerate(entries):
            if not isinstance(entry, dict):
                continue
            item = entry.get("item", {})
            if not isinstance(item, dict):
                continue
            name   = item.get("name", "")
            if not name:
                continue
            # Prefer URL directly from JSON-LD (ListItem.url or item.url/sameAs)
            # — these are authoritative and immune to index-mismatch bugs.
            # Fall back to title-slug matching when JSON-LD omits the URL.
            url = (entry.get("url") or item.get("url") or item.get("sameAs") or "").strip()
            if not url:
                url = _match_url_by_title(name)
            if not url:
                continue
            offers = item.get("offers", {})
            price  = offers.get("price", "")
            try:    price = f"${float(price):,.0f}" if price else ""
            except: price = str(price)
            avail  = offers.get("availableAtOrFrom", {})
            addr   = avail.get("address", {}) if isinstance(avail, dict) else {}
            hood   = addr.get("addressLocality","") or addr.get("addressRegion","")
            loc    = label
            date   = _cl_fmt_date(
                offers.get("validFrom","") or offers.get("availabilityStarts","") or
                item.get("datePosted","") or item.get("dateCreated","") or
                item.get("uploadDate","")
            )
            # Extract thumbnail image
            img = item.get("image", "")
            if isinstance(img, list):
                img = img[0] if img else ""
            if isinstance(img, dict):
                img = img.get("url", "") or img.get("contentUrl", "")

            items.append({"title": name, "url": url, "price": price,
                          "location": loc, "date": date, "cityId": city_id,
                          "image": img or ""})
        if items:
            break  # Found and parsed the ItemList, done

    return items


def _cl_search(query: str, cities: list = None, title_only: bool = False) -> list[dict]:
    """Search Craigslist musical instruments across US cities.
    If title_only=True, adds srchType=T to restrict matches to listing titles."""
    import time as _time
    results   = []
    seen_urls = set()
    search_cities = cities if cities else _CL_CITIES

    def _search_city(city_id):
        try:
            # Each thread gets its own session to avoid thread-safety issues
            s = http.Session()
            s.headers.update({
                "User-Agent": random.choice(_USER_AGENTS),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
            })
            # Small random delay to avoid hammering CL simultaneously
            _time.sleep(random.uniform(0.05, 0.3))
            srch_param = "&srchType=T" if title_only else ""
            url = (f"https://{city_id}.craigslist.org/search/msa"
                   f"?query={http.utils.quote(query)}&sort=date{srch_param}")
            r = s.get(url, timeout=12)
            if r.status_code == 200:
                return _cl_parse_html(r.text, city_id)
        except Exception:
            pass
        return []

    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_search_city, c): c for c in search_cities}
        for future in as_completed(futures):
            for item in future.result():
                title_key = f"{item['title'].lower().strip()}|{item['price']}|{item['cityId']}"
                if title_key not in seen_urls:
                    seen_urls.add(title_key)
                    results.append(item)

    results.sort(key=lambda x: x.get("date",""), reverse=True)
    return results


@app.route("/api/cl-parse-test")
@optional_user_context
def api_cl_parse_test():
    """Test the CL parser on a live page and show what it finds."""
    denied = _require_admin_api()
    if denied:
        return denied
    city = request.args.get("city", "sfbay")
    # Allowlist the city before interpolating it into an outbound URL (matches
    # /api/cl-debug). Admin-only, but removes the SSRF primitive entirely.
    if city not in _CL_CITIES:
        return jsonify({"error": "Unknown city."}), 400
    q    = request.args.get("q", "telecaster")
    try:
        url  = f"https://{city}.craigslist.org/search/msa?query={http.utils.quote(q)}&sort=date"
        r    = _http.get(url, timeout=12)
        html = r.text
        blocks = re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', html, re.DOTALL)
        parsed_blocks = []
        for i, b in enumerate(blocks):
            try:
                d = json.loads(b)
                parsed_blocks.append({
                    "index": i,
                    "type": d.get("@type","") if isinstance(d, dict) else type(d).__name__,
                    "keys": list(d.keys())[:15] if isinstance(d, dict) else [],
                    "sample": b[:1200],
                })
            except Exception as e:
                parsed_blocks.append({"index": i, "parse_error": str(e), "raw": b[:400]})
        # Also show raw keys from first listing for date debugging
        raw_keys = {}
        for block in re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', html, re.DOTALL):
            try:
                d = json.loads(block)
                entries = []
                if isinstance(d, list): entries = d
                elif d.get("@type") == "CollectionPage": entries = d.get("mainEntity",{}).get("itemListElement",[])
                elif d.get("@type") in ("ItemList","ListItem"): entries = [d]
                if entries:
                    item = entries[0].get("item", entries[0]) if isinstance(entries[0], dict) else {}
                    offers = item.get("offers", {})
                    raw_keys = {"item_keys": list(item.keys()), "offer_keys": list(offers.keys())}
                    break
            except Exception:
                pass
        results = _cl_parse_html(html, city)
        return jsonify({
            "html_size": len(html),
            "json_ld_block_count": len(blocks),
            "blocks": parsed_blocks,
            "first_listing_keys": raw_keys,
            "parser_results_count": len(results),
            "parser_sample": results[:3],
        })
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/api/cl-debug")
@optional_user_context
def api_cl_debug():
    """Probe a CL city to find the right section code and response format."""
    denied = _require_admin_api()
    if denied:
        return denied
    city = request.args.get("city", "sfbay")
    if city not in _CL_CITIES:
        return jsonify({"error": f"Unknown city '{city}'. Must be one of the supported CL cities."}), 400
    q    = request.args.get("q", "telecaster")
    out  = {}
    for section in ["msa", "msg", "mso", "mlt"]:
        # Try plain HTML
        try:
            url = f"https://{city}.craigslist.org/search/{section}?query={http.utils.quote(q)}&sort=date"
            r   = _http.get(url, timeout=10)
            out[f"{section}_html"] = {
                "status": r.status_code,
                "size":   len(r.text),
                "has_results": any(x in r.text for x in ["result-row","cl-search-result","listing-id"]),
                "snippet": r.text[2000:2500],
            }
        except Exception as e:
            out[f"{section}_html"] = {"error": str(e)}
        # Try format=json
        try:
            url = f"https://{city}.craigslist.org/search/{section}?query={http.utils.quote(q)}&sort=date&format=json"
            r   = _http.get(url, timeout=10)
            out[f"{section}_json"] = {
                "status":       r.status_code,
                "content_type": r.headers.get("Content-Type",""),
                "size":         len(r.text),
                "snippet":      r.text[:1000],
            }
        except Exception as e:
            out[f"{section}_json"] = {"error": str(e)}
    return jsonify(out)


@app.route("/api/debug-fetch")
@optional_user_context
def api_debug_fetch():
    """Test Algolia API fetch for a store."""
    denied = _require_admin_api()
    if denied:
        return denied
    store = request.args.get("store", "Austin")
    try:
        data     = fetch_page(store, 1)
        products = parse_products(data, store)
        results  = data.get("results", [{}])
        first    = results[0] if results else {}
        return jsonify({
            "store":          store,
            "nb_hits":        first.get("nbHits", 0),
            "nb_pages":       first.get("nbPages", 0),
            "products_found": len(products),
            "sample":         products[:3],
            "raw_hit_sample": first.get("hits", [{}])[:2],
        })
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/api/debug-condition")
@optional_user_context
def api_debug_condition():
    """Inspect the saved listing HTML to find exactly where condition data lives."""
    denied = _require_admin_api()
    if denied:
        return denied
    debug_file = DATA_DIR / "gc_debug_listing.html"
    if not debug_file.exists():
        return jsonify({"error": "No debug file yet — run the tracker once to save a listing page, then visit this URL."})
    html = debug_file.read_text(errors="replace")

    report = {"html_size": len(html)}

    # 1. All "Condition" occurrences in raw HTML (catches server-rendered text)
    condition_hits = []
    for m in re.finditer(r'.{0,60}[Cc]ondition.{0,60}', html):
        txt = m.group(0).strip()
        if txt not in condition_hits:
            condition_hits.append(txt)
        if len(condition_hits) >= 15:
            break
    report["condition_in_raw_html"] = condition_hits

    # 2. Dig into __NEXT_DATA__ and find ALL keys that contain condition-like words
    nd_condition_fields = {}
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if m:
        try:
            nd = json.loads(m.group(1))
            report["has_next_data"] = True
            report["next_data_size"] = len(m.group(1))

            # Walk every key/value pair and collect anything condition-related
            def walk(obj, path=""):
                if isinstance(obj, dict):
                    for k, v in obj.items():
                        full = f"{path}.{k}" if path else k
                        if any(c in k.lower() for c in ("condition", "grade", "quality", "rating")):
                            nd_condition_fields[full] = str(v)[:120]
                        if isinstance(v, (dict, list)):
                            walk(v, full)
                elif isinstance(obj, list):
                    for i, item in enumerate(obj[:5]):  # only first 5 items
                        walk(item, f"{path}[{i}]")
            walk(nd)
        except Exception as e:
            report["next_data_parse_error"] = str(e)
    else:
        report["has_next_data"] = False

    report["next_data_condition_fields"] = nd_condition_fields

    # 3. All JSON-LD blocks — show the full offers object for first item
    ld_offers = []
    for block in re.findall(r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>', html, re.DOTALL):
        try:
            d = json.loads(block)
            if d.get("@type") == "CollectionPage":
                items = d.get("mainEntity", {}).get("itemListElement", [])
                for entry in items[:3]:
                    item = entry.get("item", {})
                    ld_offers.append({
                        "name": item.get("name", "")[:60],
                        "offers": item.get("offers", {}),
                    })
        except Exception:
            pass
    report["jsonld_first_3_offers"] = ld_offers

    # 4. Show raw HTML snippet around first .gc product URL
    m2 = re.search(r'https?://www\.guitarcenter\.com/Used/[^"\'<>\s]+\.gc', html)
    if m2:
        start = max(0, m2.start() - 300)
        end = min(len(html), m2.end() + 600)
        report["html_around_first_product_url"] = html[start:end]

    return jsonify(report)

@app.route("/api/debug-condition/reset", methods=["GET", "POST"])
@optional_user_context
def api_debug_condition_reset():
    denied = _require_admin_api()
    if denied:
        return denied
    debug_file = DATA_DIR / "gc_debug_listing.html"
    if debug_file.exists():
        debug_file.unlink()
    return jsonify({"status": "cleared"})

@app.route("/api/debug-condition/diag")
@optional_user_context
def api_debug_condition_diag():
    """Read the condition extraction diagnostic log."""
    denied = _require_admin_api()
    if denied:
        return denied
    diag_file = DATA_DIR / "gc_condition_diag.json"
    if not diag_file.exists():
        return jsonify({"error": "No diagnostic file yet — run the tracker first."})
    return diag_file.read_text()

@app.route("/api/stop", methods=["POST"])
@optional_user_context
def api_stop():
    # Require the run_id of the active scan so external actors cannot stop
    # someone else's scan without first knowing the ID.  The client receives
    # run_id from /api/run and must echo it here.  Admin sessions are exempt
    # so the admin clear-lock page still works without a run_id.
    if not _is_admin():
        data   = request.json or {}
        req_id = (data.get("run_id") or "").strip()
        if not req_id or req_id != _current_run_id:
            return jsonify({"error": "Invalid or missing run_id."}), 403
    _stop_event.set()
    # Force-release lock after a short delay to prevent stuck state
    def _force_unlock():
        import time; time.sleep(5)
        if _lock.locked():
            try: _lock.release()
            except RuntimeError: pass
    threading.Thread(target=_force_unlock, daemon=True).start()
    return jsonify({"status": "stopping"})

@app.route("/api/populate-store-data", methods=["POST"])
@optional_user_context
def api_populate_store_data():
    """One-time migration: scan stores to tag cache entries with their store name."""
    denied = _require_admin_api()
    if denied:
        return denied
    if not _lock.acquire(blocking=False):
        return jsonify({"error": "A run is already in progress."}), 409
    _stop_event.clear()
    while not _q.empty():
        try: _q.get_nowait()
        except queue.Empty: break
    data = request.json or {}
    stores = data.get("stores", [])  # empty = all stores
    t = threading.Thread(target=_populate_store_data, args=(stores,), daemon=True)
    t.start()
    return jsonify({"status": "started"})


def _populate_store_data(selected_stores: list = None):
    """Fetch pages of each store and tag cache entries with their store name."""
    def send(msg): _q.put(msg)
    try:
        _load_cat_cache()
        stores = selected_stores if selected_stores else get_store_list()
        total  = len(stores)
        updated = 0
        label = f"{total} selected store(s)" if selected_stores else f"all {total} stores"
        send({"type": "progress", "msg": f"Tagging cache entries for {label}…"})
        send({"type": "progress", "msg": "You can stop at any time — progress is saved as it goes."})

        for i, store in enumerate(stores, 1):
            if _stop_event.is_set():
                send({"type": "progress", "msg": "⏹ Stopped."})
                break
            if i % 20 == 1:
                send({"type": "progress", "msg": f"  [{i}/{total}] {store}…"})
            try:
                page = 1
                while page <= 50:
                    data = fetch_page(store, page)
                    products = parse_products(data, store)
                    if not products:
                        break
                    for p in products:
                        sku = p["id"]
                        if sku in _cat_cache and not _cat_cache[sku].get("store"):
                            _cat_cache[sku]["store"] = store
                            _cat_cache[sku]["name"]  = _cat_cache[sku].get("name") or p.get("name","")
                            _cat_cache[sku]["url"]   = _cat_cache[sku].get("url")  or p.get("url","")
                            _cat_cache[sku]["price"] = _cat_cache[sku].get("price") or p.get("price",0)
                            _cat_cache[sku]["brand"] = _cat_cache[sku].get("brand") or p.get("brand","")
                            _cat_cache[sku]["location"] = _cat_cache[sku].get("location") or p.get("location","")
                            _cat_cache[sku]["category"] = _cat_cache[sku].get("category") or p.get("category","")
                            _cat_cache[sku]["subcategory"] = _cat_cache[sku].get("subcategory") or p.get("subcategory","")
                            _cat_cache[sku]["date_listed"] = _cat_cache[sku].get("date_listed") or p.get("date_listed","")
                            updated += 1
                    if len(products) < PAGE_SIZE:
                        break
                    page += 1
                    _sleep(1.0, 0.5)
            except Exception:
                pass
            _sleep(1.5, 0.8)

        _save_cat_cache()
        send({"type": "progress", "msg": f"\n✓ Done — {updated} cache entries tagged with store names."})
        send({"type": "done", "baseline": False, "stopped": _stop_event.is_set(),
              "scanned": total, "new_count": 0, "new_items": [], "all_items": [],
              "gap_fill": True, "fixed": updated})
    except Exception as e:
        # Don't leak exception text over the (public) SSE stream — log it server-side.
        # (2026-07 audit / round-2 deferred L3)
        print(f"[scan] operation failed: {type(e).__name__}: {e}")
        send({"type": "done", "error": "Operation failed — see server logs.", "scanned": 0, "new_count": 0, "new_items": []})
    finally:
        _lock.release()



@app.route("/api/pg-parity-check")
@optional_user_context
def api_pg_parity_check():
    denied = _require_admin_api()
    if denied:
        return denied
    if not (_PSYCOPG2_AVAILABLE and PG_DATABASE_URL):
        return jsonify({"error": "Postgres not configured"}), 503
    try:
        return jsonify(_pg_parity_check())
    except Exception as e:
        return jsonify({"error": f"{type(e).__name__}: {e}"}), 502

@app.route("/api/pg-full-backfill", methods=["POST"])
@optional_user_context
def api_pg_full_backfill():
    denied = _require_admin_api()
    if denied:
        return denied
    if not (_PSYCOPG2_AVAILABLE and PG_DATABASE_URL):
        return jsonify({"error": "Postgres not configured"}), 503
    if not _lock.acquire(blocking=False):
        return jsonify({"error": "A run is already in progress."}), 409
    _stop_event.clear()
    while not _q.empty():
        try: _q.get_nowait()
        except queue.Empty: break
    t = threading.Thread(target=_pg_full_backfill, daemon=True)
    t.start()
    return jsonify({"status": "started"})

@app.route("/api/validate-stores", methods=["POST"])
@optional_user_context
def api_validate_stores():
    denied = _require_admin_api()
    if denied:
        return denied
    if not _lock.acquire(blocking=False):
        return jsonify({"error": "A run is already in progress."}), 409
    _stop_event.clear()
    while not _q.empty():
        try: _q.get_nowait()
        except queue.Empty: break
    t = threading.Thread(target=_validate_stores, daemon=True)
    t.start()
    return jsonify({"status": "started"})

@app.route("/api/store-coords")
@optional_user_context
def api_store_coords():
    """Return cached store coordinates JSON (built by /api/build-store-coords)."""
    if STORE_COORDS_FILE.exists():
        try:
            return jsonify(json.loads(STORE_COORDS_FILE.read_text()))
        except Exception:
            pass
    return jsonify({})

@app.route("/api/build-store-coords", methods=["POST"])
@optional_user_context
def api_build_store_coords():
    """Trigger a one-time geocoding run to build gc_store_coords.json.
    Uses the existing SSE stream — progress shows up in the log panel."""
    denied = _require_admin_api()
    if denied:
        return denied
    if not _lock.acquire(blocking=False):
        return jsonify({"error": "A run is already in progress."}), 409
    _stop_event.clear()
    while not _q.empty():
        try: _q.get_nowait()
        except queue.Empty: break

    force = bool((request.json or {}).get("force", False))

    def _run():
        try:
            def _send(msg):
                _q.put({"type": "progress", "msg": msg})
            _build_store_coords(_send, force=force)
            _q.put({"type": "done", "baseline": False, "stopped": False,
                    "new_ids": [], "items": []})
        except Exception as e:
            _q.put({"type": "progress", "msg": f"Error: {e}"})
            _q.put({"type": "done", "baseline": False, "stopped": True,
                    "new_ids": [], "items": []})
        finally:
            try: _lock.release()
            except Exception: pass

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started"})

@app.route("/api/fill-gaps", methods=["POST"])
@optional_user_context
def api_fill_gaps():
    """Re-scrape listing pages for selected stores to fill missing condition/category data."""
    denied = _require_admin_api()
    if denied:
        return denied
    if not _lock.acquire(blocking=False):
        return jsonify({"error": "A run is already in progress."}), 409
    _stop_event.clear()
    while not _q.empty():
        try: _q.get_nowait()
        except queue.Empty: break
    data = request.json or {}
    stores = data.get("stores", [])
    if not stores:
        _lock.release()
        return jsonify({"error": "No stores selected."}), 400
    t = threading.Thread(target=_fill_gaps, args=(stores,), daemon=True)
    t.start()
    return jsonify({"status": "started"})

@app.route("/api/progress")
@optional_user_context
def api_progress():
    run_id = request.args.get("run_id", "")
    # Each SSE connection gets its own subscriber queue so fan-out works correctly.
    # For non-run endpoints (populate, validate, etc.) fall back to the legacy global queue.
    if run_id:
        my_q = _subscribe_to_run(run_id)
        if my_q is None:
            # Run already finished or never existed — send an empty done so client recovers
            def _empty():
                yield f"data: {json.dumps({'type':'done','new_count':0,'new_items':[],'scanned':0})}\n\n"
            return Response(stream_with_context(_empty()), mimetype="text/event-stream",
                            headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})
    else:
        my_q = _q
    def generate():
        try:
            while True:
                try:
                    msg = my_q.get(timeout=30)
                    yield f"data: {json.dumps(msg)}\n\n"
                    if msg.get("type") == "done":
                        break
                except queue.Empty:
                    yield f"data: {json.dumps({'type':'ping'})}\n\n"
        finally:
            if run_id and my_q is not _q:
                _cleanup_subscriber(run_id, my_q)
    return Response(stream_with_context(generate()), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


def _check_store_url(store_name: str) -> tuple[bool, str]:
    """Check if a store name works in the GC filter URL.
    Returns (is_valid, working_name). Tries variations if the original fails."""
    def _try(name: str) -> bool:
        try:
            query = f"filters=stores:{name.replace(' ', '%20')}"
            url   = f"https://www.guitarcenter.com/Used/?{query}&page=1"
            r = _http.get(url, timeout=10, allow_redirects=True)
            return r.status_code != 404
        except Exception:
            return True  # network error — assume valid

    if _try(store_name):
        return True, store_name

    # Try stripping state suffix (e.g. "Albany NY" → "Albany")
    parts = store_name.rsplit(' ', 1)
    if len(parts) == 2 and len(parts[1]) == 2 and parts[1].isupper():
        bare = parts[0]
        if _try(bare):
            return True, bare

    return False, store_name


def _validate_stores():
    """Check every store with a page-1 fetch, auto-fix names, remove 404s, then rebuild."""
    def send(msg): _q.put(msg)
    try:
        stores = get_store_list()
        total  = len(stores)
        removed = []
        renamed = []
        send({"type": "progress", "msg": f"Step 1: Validating {total} stores…"})
        send({"type": "progress", "msg": "About 0.5s per store. You can stop at any time."})

        updated_stores = list(stores)
        for i, store in enumerate(stores, 1):
            if _stop_event.is_set():
                send({"type": "progress", "msg": "⏹ Stopped by user."})
                break
            if i % 25 == 1:
                send({"type": "progress", "msg": f"  [{i}/{total}] checking…"})
            is_valid, working_name = _check_store_url(store)
            if not is_valid:
                _remove_invalid_store(store)
                removed.append(store)
                if store in updated_stores:
                    updated_stores.remove(store)
                send({"type": "progress", "msg": f"  ✗ Removed: {store}"})
            elif working_name != store:
                idx = updated_stores.index(store) if store in updated_stores else -1
                if idx >= 0:
                    updated_stores[idx] = working_name
                renamed.append(f"{store} → {working_name}")
                send({"type": "progress", "msg": f"  ✎ Renamed: {store} → {working_name}"})
            _sleep(0.5, 0.3)  # 0.2–0.8s between store checks

        # Save corrected names back to cache
        if renamed:
            try:
                d = json.loads(STORES_CACHE.read_text()) if STORES_CACHE.exists() else {}
                d["stores"] = sorted(set(updated_stores))
                STORES_CACHE.write_text(json.dumps(d))
            except Exception:
                pass

        if removed:
            send({"type": "progress", "msg": f"\n  Removed {len(removed)}: {', '.join(removed)}"})
        if renamed:
            send({"type": "progress", "msg": f"  Renamed {len(renamed)}: {', '.join(renamed)}"})
        if not removed and not renamed:
            send({"type": "progress", "msg": "\n  All stores validated — none removed or renamed."})

        if not _stop_event.is_set():
            send({"type": "progress", "msg": "\nStep 2: Rebuilding store list from GC's live data…"})
            try:
                new_stores = refresh_store_list()
                send({"type": "progress", "msg": f"  ✓ Store list rebuilt — {len(new_stores)} stores."})
            except Exception as e:
                send({"type": "progress", "msg": f"  Rebuild failed: {e}"})

        final_stores = get_store_list()
        send({"type": "progress", "msg": f"\n✓ Done — {len(final_stores)} valid stores in list."})
        send({"type": "done", "baseline": False, "stopped": _stop_event.is_set(),
              "scanned": total, "new_count": 0, "new_items": [], "all_items": [],
              "gap_fill": True, "fixed": len(removed)})
    except Exception as e:
        # Don't leak exception text over the (public) SSE stream — log it server-side.
        # (2026-07 audit / round-2 deferred L3)
        print(f"[scan] operation failed: {type(e).__name__}: {e}")
        send({"type": "done", "error": "Operation failed — see server logs.", "scanned": 0, "new_count": 0, "new_items": []})
    finally:
        _lock.release()


def _fill_gaps(selected_stores: list[str]):
    """Fetch individual product pages for items missing category or condition data."""
    def send(msg): _q.put(msg)
    try:
        _load_cat_cache()

        # Find cache entries that need fixing — missing category OR empty condition
        gaps = {
            sku: data for sku, data in _cat_cache.items()
            if data.get("url")
            and (not data.get("category") or not data.get("condition"))
        }

        total = len(gaps)
        if total == 0:
            send({"type": "progress", "msg": "No gaps found — all items already have category and condition data."})
            send({"type": "done", "baseline": False, "stopped": False,
                  "scanned": 0, "new_count": 0, "new_items": [], "all_items": [],
                  "gap_fill": True, "fixed": 0})
            return

        send({"type": "progress", "msg": f"Found {total} items with missing data. Fetching product pages in parallel…"})
        send({"type": "progress", "msg": f"(You can stop at any time.)"})

        fixed = 0
        gap_list = list(gaps.items())

        def _fetch_gap(item):
            sku, data = item
            url  = data.get("url", "")
            name = data.get("name", "")
            try:
                _sleep(0.3, 0.2)  # 0.1–0.5s jitter
                cat, subcat, condition = fetch_page_data(url, name)
                return sku, cat, subcat, condition
            except Exception:
                return sku, "", "", ""

        with ThreadPoolExecutor(max_workers=5) as pool:
            futures = {pool.submit(_fetch_gap, item): item for item in gap_list}
            for future in as_completed(futures):
                if _stop_event.is_set():
                    send({"type": "progress", "msg": "⏹ Stopped by user."})
                    break
                sku, cat, subcat, condition = future.result()
                data = gaps[sku]
                _cat_cache[sku].update({
                    "category":          cat or data.get("category", ""),
                    "subcategory":       subcat or data.get("subcategory", ""),
                    "condition":         condition or data.get("condition", ""),
                    "condition_fetched": True,
                })
                fixed += 1
                if fixed % 10 == 0:
                    send({"type": "progress", "msg": f"  …{fixed}/{total} items updated"})

        _save_cat_cache()
        send({"type": "progress", "msg": f"\n✓ Done — {fixed} item(s) updated. Re-run your stores to see the refreshed data."})
        send({"type": "done", "baseline": False, "stopped": _stop_event.is_set(),
              "scanned": fixed, "new_count": 0, "new_items": [], "all_items": [],
              "gap_fill": True, "fixed": fixed})
    except Exception as e:
        # Don't leak exception text over the (public) SSE stream — log it server-side.
        # (2026-07 audit / round-2 deferred L3)
        print(f"[scan] operation failed: {type(e).__name__}: {e}")
        send({"type": "done", "error": "Operation failed — see server logs.", "scanned": 0, "new_count": 0, "new_items": []})
    finally:
        _lock.release()



def _run(selected_stores: list[str], baseline: bool, run_id: str = "", device_last_run: str = "", run_time: str = "", device_last_anchor: str = "", user_id: int | None = None):
    def send(msg):
        if run_id:
            _broadcast(run_id, msg)   # fan-out to all subscribers
        _q.put(msg)                   # also send to legacy queue for backwards compat
    try:
        # Use the run_time passed in from api_run (computed before thread start)
        # so the client and server share the exact same timestamp.
        if not run_time:
            run_time = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

        stores_to_scan = selected_stores if not baseline else []
        nationwide = baseline or len(stores_to_scan) == 0
        label = "nationwide scan" if nationwide else f"{len(stores_to_scan)} store(s)"
        send({"type":"progress","msg":f"Starting {label}…"})

        all_products, ids_this_run = [], set()
        # Track scan coverage gaps so we never let a run that MISSED some stores'
        # or pages' data (timeout, transient error, user stop) silently mark those
        # stores' items sold, or advance the NEW-detection anchor past items we
        # never actually got a chance to see (v2.16.11 — see scrape_store()).
        incomplete_stores: set[str] = set()
        scan_incomplete = False

        if nationwide:
            # ── Nationwide: query ALL used inventory via parallel page fetches ────
            PARALLEL_WORKERS = 15  # concurrent API requests
            send({"type":"progress","msg":"Fetching all used inventory nationwide via API…"})
            # First fetch page 1 to learn total pages
            try:
                data1 = fetch_page(None, 1)
            except Exception as e:
                send({"type":"progress","msg":f"  API error on page 1: {e}"})
                data1 = None
            if not data1:
                scan_incomplete = True
            if data1:
                products1 = parse_products(data1, None)
                for p in products1:
                    if p["id"] not in ids_this_run:
                        all_products.append(p)
                        ids_this_run.add(p["id"])
                try:
                    nb_pages = data1.get("results", [{}])[0].get("nbPages", 1)
                    nb_hits  = data1.get("results", [{}])[0].get("nbHits", 0)
                    send({"type":"progress","msg":f"  {nb_hits:,} items across {nb_pages} pages — fetching {PARALLEL_WORKERS} pages at a time…"})
                except Exception:
                    nb_pages = 1
                # Fetch remaining pages in parallel batches
                remaining = list(range(2, min(nb_pages + 1, 1001)))
                def _fetch_one_page(pg):
                    if _stop_event.is_set():
                        return pg, None, None
                    try:
                        d = fetch_page(None, pg)
                        return pg, d, None
                    except Exception as exc:
                        return pg, None, exc
                # Hard wall-clock ceiling per batch — a normal Algolia page fetch
                # completes in well under a second, so this is purely a safety net
                # against a connection that stalls past requests' own timeout=
                # (v2.16.10, see STORE_SCAN_TIMEOUT below for the matching store-scan fix).
                PAGE_BATCH_TIMEOUT = 90
                batch_idx = 0
                while batch_idx < len(remaining) and not _stop_event.is_set():
                    batch = remaining[batch_idx:batch_idx + PARALLEL_WORKERS]
                    batch_idx += len(batch)
                    pool = ThreadPoolExecutor(max_workers=PARALLEL_WORKERS)
                    try:
                        futures = {pool.submit(_fetch_one_page, pg): pg for pg in batch}
                        try:
                            for fut in as_completed(futures, timeout=PAGE_BATCH_TIMEOUT):
                                pg, data, err = fut.result()
                                if err:
                                    send({"type":"progress","msg":f"  API error on page {pg}: {err}"})
                                    scan_incomplete = True
                                    continue
                                if data is None:
                                    continue
                                products = parse_products(data, None)
                                for p in products:
                                    if p["id"] not in ids_this_run:
                                        all_products.append(p)
                                        ids_this_run.add(p["id"])
                        except _FutureTimeoutError:
                            stuck = [futures[f] for f in futures if not f.done()]
                            scan_incomplete = True
                            send({"type":"progress","msg":f"  ⚠ page(s) {stuck} stalled past {PAGE_BATCH_TIMEOUT}s — skipping, continuing scan."})
                    finally:
                        # wait=False: never block here on a genuinely stuck worker thread —
                        # that's the exact hang this fix exists to prevent. The thread is
                        # abandoned and will finish or die on its own without holding up the scan.
                        pool.shutdown(wait=False)
                    # Progress update after each batch
                    pages_done = min(batch_idx + 1, nb_pages)
                    send({"type":"progress","msg":f"  page {pages_done}/{nb_pages}… ({len(all_products):,} items so far)"})
                if _stop_event.is_set():
                    send({"type":"progress","msg":"⏹ Stopped by user."})
                    scan_incomplete = True
            send({"type":"progress","msg":f"  Fetched {len(all_products):,} items total."})
        else:
            # ── Normal scan: query selected stores in parallel ────────────────
            STORE_WORKERS = 10
            # Hard wall-clock ceiling on the whole batch of stores, scaled by count
            # (12s/store budget — generous headroom over normal completion time,
            # since up to STORE_WORKERS run concurrently). This is what fixes the
            # bug where a single-store scan would sit pinging with no progress and
            # no completion for 10+ minutes: fetch_page()'s requests timeout= only
            # bounds socket-level stalls (no bytes for N seconds), not a slow trickle
            # that keeps the connection technically alive — this bounds the whole
            # wait regardless of what's stalling underneath (v2.16.10).
            STORE_SCAN_TIMEOUT = max(120, len(stores_to_scan) * 12)
            send({"type":"progress","msg":f"Scanning {len(stores_to_scan)} stores ({STORE_WORKERS} at a time)…"})
            completed = [0]
            lock = threading.Lock()
            def _scan_one_store(store):
                if _stop_event.is_set():
                    return store, [], set(), False
                _rotate_ua()
                products, ids, complete = scrape_store(store, send, _stop_event)
                with lock:
                    completed[0] += 1
                    send({"type":"progress","msg":f"  [{completed[0]}/{len(stores_to_scan)}] {store} — {len(products)} items"})
                return store, products, ids, complete
            pool = ThreadPoolExecutor(max_workers=STORE_WORKERS)
            try:
                futures = {pool.submit(_scan_one_store, s): s for s in stores_to_scan}
                try:
                    for fut in as_completed(futures, timeout=STORE_SCAN_TIMEOUT):
                        if _stop_event.is_set():
                            send({"type":"progress","msg":"⏹ Stopped by user."})
                            scan_incomplete = True
                            break
                        store, products, ids, complete = fut.result()
                        if not complete:
                            incomplete_stores.add(store)
                        for p in products:
                            if p["id"] not in ids_this_run:
                                all_products.append(p)
                        ids_this_run |= ids
                except _FutureTimeoutError:
                    stuck = [futures[f] for f in futures if not f.done()]
                    incomplete_stores.update(stuck)
                    send({"type":"progress","msg":f"  ⚠ {stuck} stalled past {STORE_SCAN_TIMEOUT}s — skipping, continuing with what we have."})
            finally:
                # wait=False: never block here on a genuinely stuck worker thread — see note above.
                pool.shutdown(wait=False)

        # (cache-ID snapshot removed — NEW detection now uses startDate timestamps)

        # ── Anchor date for NEW detection (per-user, v2.10.18) ───────────────────
        # The anchor represents "the max date_listed of items this user was exposed
        # to at their last scan." Anything with date_listed > anchor is genuinely new
        # to THIS user. This handles Algolia's 6-12h indexing pipeline delay: items
        # can appear in search results with date_listed values older than the last
        # scan time, which would make them invisible to timestamp-based detection
        # but they'd silently push existing items down the date-sorted table (the
        # "0 new / reordered" bug).
        #
        # IMPORTANT: We use the *per-user* stored anchor (passed in as
        # device_last_anchor), NOT max(date_listed in _cat_cache). _cat_cache is the
        # global shared inventory written by EVERY user's scan, so reading it here
        # contaminates the anchor with other users' activity — if Alice scanned five
        # minutes ago, Bob's threshold would jump to Alice's freshest item and Bob
        # would see 0 new items even when items are genuinely new to him. (Bug
        # introduced in v2.10.11, fixed in v2.10.18.)
        anchor_date = device_last_anchor or ""

        # ── Apply data from Algolia API to products, tracking price drops ────────
        # Categories, condition, brand all come from the API now — no page scraping needed.
        for p in all_products:
            sku    = p["id"]
            cached = _cat_cache.get(sku, {})
            cat    = p.get("category") or cached.get("category", "")
            subcat = p.get("subcategory") or cached.get("subcategory", "")
            condition = p.get("condition") or cached.get("condition", "")
            brand     = p.get("brand") or cached.get("brand", "")
            location  = p.get("location") or cached.get("location", p.get("store", ""))
            # Price drop detection — use Algolia's native priceDrop flag + listPrice field
            new_price       = p.get("price") or 0
            new_list_price  = p.get("list_price") or 0
            has_price_drop  = bool(p.get("has_price_drop", False))
            price_drop_amt  = round(new_list_price - new_price, 2) if (has_price_drop and new_list_price > new_price) else 0
            # Track when we FIRST detected this drop (preserves timestamp across scans)
            prev_had_drop   = bool(cached.get("has_price_drop", False))
            if has_price_drop and not prev_had_drop:
                price_drop_since = run_time          # newly dropped this scan
            elif has_price_drop:
                price_drop_since = cached.get("price_drop_since", run_time)  # preserve
            else:
                price_drop_since = ""                # no longer dropped
            _cat_cache[sku] = {
                "category":          cat,
                "subcategory":       subcat,
                "condition":         condition,
                "brand":             brand,
                "name":              p.get("name", ""),
                "url":               p.get("url", ""),
                "store":             p.get("store", ""),
                "location":          location,
                "price":             new_price,
                "list_price":        new_list_price,
                "has_price_drop":    has_price_drop,
                "price_drop":        price_drop_amt,
                "price_drop_since":  price_drop_since,
                "available":         True,
                "date_listed":       p.get("date_listed") or cached.get("date_listed", ""),
                "image_id":          p.get("image_id") or cached.get("image_id", ""),
                # GC vintage flag (premiumGear=="Vintage"), captured at scan time
                "is_vintage":        bool(p.get("is_vintage", cached.get("is_vintage", False))),
                # Staff-written "Condition & Details" note extracted from longDescription
                # (v2.16.6) — often empty, that's expected (not every item has one).
                "condition_note":    p.get("condition_note") or cached.get("condition_note", ""),
                # first_seen: when our system first encountered this item
                "first_seen":        cached.get("first_seen", run_time),
            }
            p["category"]    = cat
            p["subcategory"] = subcat
            p["condition"]   = condition
            p["brand"]       = brand
            p["location"]    = location
            p["price_drop"]  = price_drop_amt
            p["list_price"]  = new_list_price

        # ── Mark sold items (not found in this scan) ────────────────────────────
        # Only mark items sold for stores/coverage we're confident we saw a
        # COMPLETE picture of this run — a store whose fetch errored, timed out,
        # or got abandoned mid-scan (incomplete_stores) is not evidence its items
        # sold, just evidence we didn't finish looking. For a nationwide scan,
        # completeness isn't per-store — if page 1 failed, a page stalled, or the
        # scan was stopped early, we don't know which SKUs we missed, so skip
        # sold-marking entirely rather than risk wiping out items still in stock.
        # (v2.16.11 — see scrape_store()'s `complete` flag.)
        if not _stop_event.is_set() and (not nationwide or not scan_incomplete):
            scanned_store_set = set(stores_to_scan) - incomplete_stores
            wl = load_watchlist()
            wl_changed = False
            for sku, cached in _cat_cache.items():
                if sku in ids_this_run:
                    continue
                # Nationwide scan: any item not found is gone
                # Store scan: only mark items from stores we completely scanned
                if nationwide or cached.get("store") in scanned_store_set:
                    if cached.get("available", True):
                        cached["available"] = False
                        if sku in wl:
                            wl[sku]["sold"] = True
                            wl_changed = True
            if wl_changed:
                save_watchlist(wl)

        # ── Update watchlist with latest data ─────────────────────────────────
        wl = load_watchlist()
        changed = False
        for sku, item in wl.items():
            if sku in _cat_cache and not wl[sku].get("sold"):
                cached = _cat_cache[sku]
                wl[sku].update({
                    "price":      cached.get("price", item.get("price")),
                    "condition":  cached.get("condition", item.get("condition", "")),
                    "brand":      cached.get("brand", item.get("brand", "")),
                    "location":   cached.get("location", item.get("location", "")),
                    "category":   cached.get("category", item.get("category", "")),
                    "subcategory":cached.get("subcategory", item.get("subcategory", "")),
                    "date_listed":cached.get("date_listed", item.get("date_listed", "")),
                })
                changed = True
        if changed:
            save_watchlist(wl)

        send({"type":"progress","msg":f"  {len(all_products):,} products scanned."})
        _save_cat_cache()
        # Phase B dual-write (v2.16.15): mirror this scan into Postgres in the
        # background -- never blocks the scan's own completion. See
        # _pg_sync_scan()'s docstring and POSTGRES_MIGRATION_PLAN.md.
        threading.Thread(
            target=_pg_sync_scan,
            args=(set(ids_this_run), nationwide, scan_incomplete, set(incomplete_stores), list(stores_to_scan)),
            daemon=True,
        ).start()
        # Read global last-scan time (fallback when device has no history)
        last_scan_file = DATA_DIR / "gc_last_scan.txt"
        global_prev_scan = last_scan_file.read_text().strip() if last_scan_file.exists() else ""
        # Per-device prev_scan: prefer the device's own last-run timestamp sent
        # from localStorage. Falls back to the global scan time so first-time
        # devices on an existing server don't see the entire catalog as NEW.
        prev_scan_time = device_last_run or global_prev_scan
        # Record this scan's completion time globally (for devices with no history)
        last_scan_file.write_text(run_time)

        def fmt(p):
            date_src = p.get("date_listed") or _cat_cache.get(p["id"], {}).get("date_listed", "")
            lp = p.get("list_price") or 0
            return {
                "id":               p["id"],
                "name":             p["name"],
                "brand":            p.get("brand", ""),
                "price":            f"${p['price']:,.2f}" if p["price"] else "",
                "price_raw":        p.get("price") or 0,
                "list_price_raw":   lp,
                "price_drop":       p.get("price_drop", 0),
                "price_drop_since": _cat_cache.get(p["id"], {}).get("price_drop_since", ""),
                "store":            p["store"],
                "location":         p.get("location") or p.get("store", ""),
                "url":              p["url"],
                "category":         p.get("category", ""),
                "subcategory":      p.get("subcategory", ""),
                "condition":        p.get("condition", ""),
                "date":             _fmt_date(date_src),
                "date_raw":         date_src,
                "image_id":         p.get("image_id") or _cat_cache.get(p["id"], {}).get("image_id", ""),
                "condition_note":   p.get("condition_note") or _cat_cache.get(p["id"], {}).get("condition_note", ""),
            }

        # ── Per-device new-item detection ─────────────────────────────────────
        # An item is NEW if date_listed > threshold, where threshold is whichever
        # is more recent: the anchor_date (most recent item in pre-scan cache) or
        # prev_scan_time (last wall-clock scan time). The anchor approach is primary
        # because it's immune to Algolia's indexing pipeline delay — items that appear
        # in search results after our last scan but carry older date_listed values
        # (the "0 new / table reordered" bug) won't pollute the sort without being flagged.
        # GC sometimes stores date-only values ("2026-05-05") with no time component.
        # A plain string compare like "2026-05-05" > "2026-05-05T08:00:00Z" is False
        # (shorter string sorts before longer at that position), so items listed today
        # would never be flagged new once any scan ran today. Fix: treat date-only
        # values as end-of-day ("2026-05-05T23:59:59Z") so they stay new all day.
        def _norm_item_date(d):
            return d + "T23:59:59Z" if d and len(d) == 10 else d

        # Threshold = the max date_listed the user was actually exposed to at their
        # last scan (the "top of their table"). Anything with a newer date_listed is
        # genuinely new to this user.
        # We intentionally do NOT mix in prev_scan_time (wall-clock scan time) here.
        # Wall-clock timestamps are lexicographically larger than date-only strings
        # (e.g. "2026-05-18T08:00:00Z" > "2026-05-17"), so including prev_scan_time
        # in a max() would inflate the threshold above items that are genuinely new —
        # exactly the "5 items between the known item and the 10 new ones, none flagged"
        # bug. Use anchor-only; fall back to prev_scan_time only on first scan.
        _norm_anchor = _norm_item_date(anchor_date) if anchor_date else ""
        threshold = _norm_anchor if _norm_anchor else prev_scan_time

        new_ids_list = []
        if not baseline and threshold:
            for p in all_products:
                item_date = p.get("date_listed") or _cat_cache.get(p["id"], {}).get("date_listed", "")
                if item_date and _norm_item_date(item_date) > threshold:
                    new_ids_list.append(p["id"])

        send({"type":"progress","msg":f"  {len(new_ids_list):,} new items since last scan."})

        # ── Compute new per-user anchor (post-scan) ─────────────────────────────
        # The anchor we persist for this user is the max date_listed across the
        # cache AFTER this scan. It represents "everything I've now been exposed
        # to" — next time this user scans, anything older than this anchor will
        # be treated as already-seen even if Algolia surfaces it freshly (the
        # indexing-delay protection that anchor_date was designed for).
        new_anchor = ""
        # Only let this scan's data advance the anchor if we're confident we saw
        # a COMPLETE picture — one or more stores that errored/timed out/got
        # abandoned (incomplete_stores), or a nationwide scan that hit a page
        # failure or got stopped early (scan_incomplete), means this run's max
        # date_listed may be missing whatever the UNSEEN store(s)/page(s)' freshest
        # listings actually are. Advancing the anchor anyway would silently block
        # those genuinely-new items from ever being flagged NEW once we do see
        # them — the "item shows up but never gets tagged NEW" bug this guards
        # against. Trade-off: on a run with any coverage gap, the anchor simply
        # doesn't move forward this time (safe — worst case, an already-seen item
        # gets re-flagged NEW on a later clean scan) rather than risk moving past
        # something we never actually saw. (v2.16.11)
        coverage_ok = not scan_incomplete and not incomplete_stores
        if all_products and coverage_ok:
            # Use THIS scan's products only — not _cat_cache, which is global and
            # shared across all users. Using _cat_cache re-introduces the contamination
            # bug: another user's scan populates it with fresher items, inflating this
            # user's anchor and causing 0-new on their next scan.
            # (v2.10.18 fixed the threshold for the current scan but not persistence.)
            _scan_dates = [p.get("date_listed", "") for p in all_products if p.get("date_listed")]
            if _scan_dates:
                new_anchor = max(_scan_dates)
        elif not coverage_ok:
            send({"type":"progress","msg":"  ⚠ scan had gaps (some stores/pages didn't complete) — NEW-detection anchor not advanced this run."})
        # Don't let the anchor regress: preserve the old anchor if this scan
        # produced no dates (e.g. stopped early). Never include prev_scan_time —
        # wall-clock timestamps inflate the anchor and block same-date new items.
        new_anchor = max(new_anchor, anchor_date or "")

        # Persist server-side for logged-in users (atomic with this scan completing).
        # Guests receive scan_anchor in the SSE done payload and roundtrip via localStorage.
        # NOTE: We persist last_anchor on baseline scans too — a baseline establishes
        # the starting point that future scans compare against.
        if user_id:
            try:
                _set_user_data(user_id, last_anchor=new_anchor, last_run=run_time)
            except Exception:
                pass  # Non-fatal — client will also sync via /api/sync after done

        # For large scans, don't send full item lists via SSE — client will use server-side browse
        large_scan = len(all_products) > 1000
        items_for_sse = [] if large_scan else [fmt(p) for p in all_products[:500]]
        send({
            "type":        "done",
            "baseline":    baseline,
            "stopped":     _stop_event.is_set(),
            "scanned":     len(all_products),
            "new_ids":     new_ids_list,
            "scan_time":   run_time,
            "scan_anchor": new_anchor,
            "items":       items_for_sse,
            "use_browse":  large_scan,
        })
    except Exception as e:
        send({"type":"done","error":str(e),"scanned":0,"new_count":0,"new_items":[]})
    finally:
        _lock.release()


# ── HTML ──────────────────────────────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Guitar Center Used Gear Tracker — Browse Inventory by Store Location</title>
<meta name="description" content="Browse used gear at any Guitar Center location. Search guitars, amps, pedals, drums, and more by store, city, condition, and price — updated in real time. Free watch list and want list.">
<link rel="canonical" href="https://gcgeartracker.com/">
<meta property="og:type" content="website">
<meta property="og:url" content="https://gcgeartracker.com/">
<meta property="og:site_name" content="GC Used Inventory Tracker">
<meta property="og:title" content="Guitar Center Used Gear Tracker — Browse Inventory by Store Location">
<meta property="og:description" content="Browse used gear at any Guitar Center location. Filter by store, city, condition, and price across 300+ stores nationwide. Free watch list and want list.">
<meta property="og:image" content="https://gcgeartracker.com/static/og-image.png">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="Guitar Center Used Gear Tracker — Browse Inventory by Store Location">
<meta name="twitter:description" content="Browse used gear at any Guitar Center location. Filter by store, city, condition, and price. Free watch list and want list alerts.">
<meta name="twitter:image" content="https://gcgeartracker.com/static/og-image.png">
<link rel="icon" href="/static/favicon.svg" type="image/svg+xml">
<script type="application/ld+json">
{"@context":"https://schema.org","@type":"WebSite","name":"GC Used Inventory Tracker","url":"https://gcgeartracker.com/","description":"Browse used gear at any Guitar Center location across 300+ stores nationwide.","potentialAction":{"@type":"SearchAction","target":{"@type":"EntryPoint","urlTemplate":"https://gcgeartracker.com/?q={search_term_string}"},"query-input":"required name=search_term_string"}}
</script>
<link rel="stylesheet" href="/static/gc.css">
<!-- __GA__ -->
</head>
<body>

<!-- Image thumbnail tooltip -->
<div id="img-tooltip"><img src="" alt=""></div>

<!-- Password modal -->
<!-- Validate stores modal -->
<div id="vs-modal" style="display:none;position:fixed;inset:0;z-index:100;align-items:center;justify-content:center">
  <div style="position:absolute;inset:0;background:rgba(0,0,0,.7)" id="vs-backdrop"></div>
  <div style="position:relative;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:10px;padding:30px 28px;width:360px;z-index:1">
    <h2 style="color:#fff;font-size:1.05rem;margin-bottom:8px">✓ Validate Stores</h2>
    <p style="color:#777;font-size:.82rem;margin-bottom:18px;line-height:1.6">Clear the invalid-stores blocklist before validating?<br><br>
    <b style="color:#ccc">Yes (recommended)</b> — re-checks all stores including any previously removed ones.<br><br>
    <b style="color:#ccc">No</b> — only checks stores currently in your list.</p>
    <div style="display:flex;gap:8px">
      <button id="vs-cancel-btn" style="flex:1;padding:9px;border-radius:5px;font-size:.88rem;font-weight:600;cursor:pointer;border:1px solid #3a3a3a;background:#2a2a2a;color:#aaa">Cancel</button>
      <button id="vs-no-btn" style="flex:1;padding:9px;border-radius:5px;font-size:.88rem;font-weight:600;cursor:pointer;border:none;background:#444;color:#eee">No</button>
      <button id="vs-yes-btn" style="flex:1;padding:9px;border-radius:5px;font-size:.88rem;font-weight:600;cursor:pointer;border:none;background:#c00;color:#fff">Yes</button>
    </div>
  </div>
</div>


<!-- ── Welcome / auth modal (shown on first visit when not logged in) ── -->
<div id="first-run-modal" style="display:none;position:fixed;inset:0;z-index:200;align-items:center;justify-content:center">
  <div style="position:absolute;inset:0;background:rgba(0,0,0,.75)" id="first-run-backdrop"></div>
  <div style="position:relative;background:#1a1a1a;border:1px solid #2e2e2e;border-radius:12px;padding:32px 36px;width:360px;max-width:92vw;z-index:1">
    <h2 style="color:#fff;font-size:1.15rem;margin-bottom:6px">Welcome to GC Used Inventory Tracker</h2>
    <p style="color:#aaa;font-size:.82rem;margin-bottom:20px;line-height:1.5">Track Guitar Center used inventory. Create an account to save your watch list, want list, and favorites across all your devices.</p>
    <!-- Auth tabs -->
    <div style="display:flex;border-bottom:1px solid #2e2e2e;margin-bottom:20px">
      <button id="welcome-tab-login" style="flex:1;padding:8px;background:none;border:none;border-bottom:2px solid #c00;color:#ff5555;font-size:.85rem;font-weight:600;cursor:pointer;margin-bottom:-1px">Sign In</button>
      <button id="welcome-tab-register" style="flex:1;padding:8px;background:none;border:none;border-bottom:2px solid transparent;color:#666;font-size:.85rem;font-weight:600;cursor:pointer;margin-bottom:-1px">Create Account</button>
    </div>
    <!-- Login form -->
    <div id="welcome-form-login">
      <div id="welcome-google-wrap" style="display:none">
        <button class="auth-google-btn">
          <svg width="18" height="18" viewBox="0 0 18 18"><path fill="#4285F4" d="M17.64 9.2c0-.637-.057-1.251-.164-1.84H9v3.481h4.844c-.209 1.125-.843 2.078-1.796 2.717v2.258h2.908c1.702-1.566 2.684-3.875 2.684-6.615z"/><path fill="#34A853" d="M9 18c2.43 0 4.467-.806 5.956-2.18l-2.908-2.259c-.806.54-1.837.86-3.048.86-2.344 0-4.328-1.584-5.036-3.711H.957v2.332A8.997 8.997 0 0 0 9 18z"/><path fill="#FBBC05" d="M3.964 10.71A5.41 5.41 0 0 1 3.682 9c0-.593.102-1.17.282-1.71V4.958H.957A8.996 8.996 0 0 0 0 9c0 1.452.348 2.827.957 4.042l3.007-2.332z"/><path fill="#EA4335" d="M9 3.58c1.321 0 2.508.454 3.44 1.345l2.582-2.58C13.463.891 11.426 0 9 0A8.997 8.997 0 0 0 .957 4.958L3.964 7.29C4.672 5.163 6.656 3.58 9 3.58z"/></svg>
          Sign in with Google
        </button>
        <div class="auth-divider"><span>or sign in with username</span></div>
      </div>
      <input class="auth-field" type="text" id="welcome-login-user" placeholder="Username" autocomplete="username">
      <input class="auth-field" type="password" id="welcome-login-pw" placeholder="Password" autocomplete="current-password">
      <div class="auth-err" id="welcome-login-err"></div>
      <button class="auth-submit" id="welcome-login-submit">Sign In</button>
    </div>
    <!-- Register form -->
    <div id="welcome-form-register" style="display:none">
      <div id="welcome-google-wrap-reg" style="display:none">
        <button class="auth-google-btn">
          <svg width="18" height="18" viewBox="0 0 18 18"><path fill="#4285F4" d="M17.64 9.2c0-.637-.057-1.251-.164-1.84H9v3.481h4.844c-.209 1.125-.843 2.078-1.796 2.717v2.258h2.908c1.702-1.566 2.684-3.875 2.684-6.615z"/><path fill="#34A853" d="M9 18c2.43 0 4.467-.806 5.956-2.18l-2.908-2.259c-.806.54-1.837.86-3.048.86-2.344 0-4.328-1.584-5.036-3.711H.957v2.332A8.997 8.997 0 0 0 9 18z"/><path fill="#FBBC05" d="M3.964 10.71A5.41 5.41 0 0 1 3.682 9c0-.593.102-1.17.282-1.71V4.958H.957A8.996 8.996 0 0 0 0 9c0 1.452.348 2.827.957 4.042l3.007-2.332z"/><path fill="#EA4335" d="M9 3.58c1.321 0 2.508.454 3.44 1.345l2.582-2.58C13.463.891 11.426 0 9 0A8.997 8.997 0 0 0 .957 4.958L3.964 7.29C4.672 5.163 6.656 3.58 9 3.58z"/></svg>
          Continue with Google
        </button>
        <div class="auth-divider"><span>or create a username account</span></div>
      </div>
      <input class="auth-field" type="text" id="welcome-reg-user" placeholder="Choose a username" autocomplete="username" maxlength="30">
      <input class="auth-field" type="password" id="welcome-reg-pw" placeholder="Password (8+ characters)" autocomplete="new-password">
      <input class="auth-field" type="password" id="welcome-reg-pw2" placeholder="Confirm password" autocomplete="new-password">
      <input class="auth-field" type="email" id="welcome-reg-email" placeholder="Email (optional)" autocomplete="email" style="margin-bottom:4px">
      <div style="color:#aaa;font-size:.72rem;margin-bottom:12px;line-height:1.4">Optional — helps identify your account if you ever need support. Never shared or used for marketing.</div>
      <div class="auth-err" id="welcome-reg-err"></div>
      <button class="auth-submit" id="welcome-register-submit">Create Account &amp; Start Scanning</button>
    </div>
    <!-- Guest option -->
    <div style="text-align:center;margin-top:16px">
      <button id="first-run-guest-btn" style="background:none;border:none;color:#aaa;font-size:.78rem;cursor:pointer;text-decoration:underline">Use as guest</button>
    </div>
  </div>
</div>

<div id="kw-modal" style="display:none;position:fixed;inset:0;z-index:100;align-items:center;justify-content:center">
  <div style="position:absolute;inset:0;background:rgba(0,0,0,.7)" id="kw-modal-backdrop"></div>
  <div style="position:relative;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:10px;width:420px;max-width:calc(100vw - 32px);max-height:80vh;display:flex;flex-direction:column;overflow:hidden;z-index:1">
    <!-- pinned header: title, instructions, add input -->
    <div style="padding:16px 20px 0;flex-shrink:0">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:8px;position:relative;margin-bottom:4px">
        <h2 style="color:#fff;font-size:1.05rem;margin:0">🎯 Want List</h2>
        <button id="kw-info-btn" title="Keyword syntax help" style="background:none;border:1.5px solid #4ade80;border-radius:5px;color:#4ade80;font-weight:700;font-size:1rem;cursor:pointer;padding:4px 11px;line-height:1.4;flex-shrink:0">ⓘ</button>
        <div id="kw-info-popover">
          <b>Keyword syntax</b><br>
          <code>Allen</code> — whole word (not Allentown or McAllen)<br>
          <code>"Jam Pedals"</code> — exact phrase<br>
          <code>Thorpy, Dane</code> — comma = AND (both required)<br>
          <code>OD*</code> — wildcard (OD808, OD-1…) &nbsp;·&nbsp; <code>*drive*</code> — contains "drive"<br>
          <code>Mesa, -combo</code> — minus = NOT (also <code>-"combo amp"</code>)<br>
          <code>Mesa, Mark*; Heartbreaker</code> — semicolon = OR (either side)<br>
          <code>Mesa, -combo: Angel; Blues; Trem</code> — colon = apply the prefix to every OR branch
        </div>
      </div>
      <p style="color:#aaa;font-size:.82rem;margin-bottom:12px;line-height:1.45">
        Highlights matches across all results. New matches sort to top after a scan.
      </p>
      <div style="display:flex;gap:6px;margin-bottom:12px">
        <input id="kw-input" type="text" placeholder="Add an item to your want list…"
               style="flex:1;padding:8px 12px;background:#252525;border:1px solid #3a3a3a;border-radius:5px;color:#eee;font-size:.9rem;outline:none"
               >
        <button id="kw-add-btn" style="padding:8px 16px;background:#0a5c2a;border:1px solid #2d6a2d;border-radius:5px;color:#4ade80;font-size:.85rem;cursor:pointer;white-space:nowrap">+ Add</button>
      </div>
    </div>
    <!-- scrollable keyword chips -->
    <div id="kw-list" style="overflow-y:auto;flex:1;min-height:0;padding:0 24px 16px"></div>
    <!-- pinned footer: always visible -->
    <div style="display:flex;gap:10px;justify-content:space-between;border-top:1px solid #2e2e2e;padding:14px 24px 20px;flex-shrink:0">
      <button id="kw-clear-btn" style="padding:6px 14px;background:#1a1a1a;border:1px solid #5a2a2a;border-radius:5px;color:#a05050;font-size:.78rem;cursor:pointer">Clear Want List</button>
      <button id="kw-done-btn" style="padding:6px 18px;background:#252525;border:1px solid #3a3a3a;border-radius:5px;color:#aaa;font-size:.85rem;cursor:pointer">Done</button>
    </div>
  </div>
</div>

<header>
  <h1>GC Used Inventory Tracker <span style="font-size:.65rem;font-weight:400;opacity:.6"><!-- __VER__ --></span></h1>
  <button id="stop-btn">⏹ Stop Running</button>
  <span id="hdr-status">Loading…</span>
  <div id="auth-widget">
    <span id="auth-sync-dot" title="Synced to account"></span>
    <div id="auth-user-info">
      <span id="auth-email"></span>
      <button id="auth-logout-btn">Sign out</button>
    </div>
    <button id="auth-login-btn">Sign in</button>
  </div>
</header>

<!-- ── Auth modal ── -->
<div id="auth-modal">
  <div class="auth-box">
    <button class="auth-close">✕</button>
    <div class="auth-tabs">
      <button class="auth-tab active" id="auth-tab-login">Sign In</button>
      <button class="auth-tab" id="auth-tab-register">Create Account</button>
    </div>
    <!-- Login form -->
    <div id="auth-form-login">
      <div id="auth-google-wrap" style="display:none">
        <button class="auth-google-btn">
          <svg width="18" height="18" viewBox="0 0 18 18"><path fill="#4285F4" d="M17.64 9.2c0-.637-.057-1.251-.164-1.84H9v3.481h4.844c-.209 1.125-.843 2.078-1.796 2.717v2.258h2.908c1.702-1.566 2.684-3.875 2.684-6.615z"/><path fill="#34A853" d="M9 18c2.43 0 4.467-.806 5.956-2.18l-2.908-2.259c-.806.54-1.837.86-3.048.86-2.344 0-4.328-1.584-5.036-3.711H.957v2.332A8.997 8.997 0 0 0 9 18z"/><path fill="#FBBC05" d="M3.964 10.71A5.41 5.41 0 0 1 3.682 9c0-.593.102-1.17.282-1.71V4.958H.957A8.996 8.996 0 0 0 0 9c0 1.452.348 2.827.957 4.042l3.007-2.332z"/><path fill="#EA4335" d="M9 3.58c1.321 0 2.508.454 3.44 1.345l2.582-2.58C13.463.891 11.426 0 9 0A8.997 8.997 0 0 0 .957 4.958L3.964 7.29C4.672 5.163 6.656 3.58 9 3.58z"/></svg>
          Sign in with Google
        </button>
        <div class="auth-divider"><span>or sign in with username</span></div>
      </div>
      <input class="auth-field" type="text" id="auth-login-user" placeholder="Username" autocomplete="username">
      <input class="auth-field" type="password" id="auth-login-pw" placeholder="Password" autocomplete="current-password">
      <div class="auth-err" id="auth-login-err"></div>
      <button class="auth-submit" id="auth-login-submit">Sign In</button>
      <div class="auth-note">Your watch list, want list &amp; favorites sync across all your devices.</div>
    </div>
    <!-- Register form -->
    <div id="auth-form-register" style="display:none">
      <div id="auth-google-wrap-reg" style="display:none">
        <button class="auth-google-btn">
          <svg width="18" height="18" viewBox="0 0 18 18"><path fill="#4285F4" d="M17.64 9.2c0-.637-.057-1.251-.164-1.84H9v3.481h4.844c-.209 1.125-.843 2.078-1.796 2.717v2.258h2.908c1.702-1.566 2.684-3.875 2.684-6.615z"/><path fill="#34A853" d="M9 18c2.43 0 4.467-.806 5.956-2.18l-2.908-2.259c-.806.54-1.837.86-3.048.86-2.344 0-4.328-1.584-5.036-3.711H.957v2.332A8.997 8.997 0 0 0 9 18z"/><path fill="#FBBC05" d="M3.964 10.71A5.41 5.41 0 0 1 3.682 9c0-.593.102-1.17.282-1.71V4.958H.957A8.996 8.996 0 0 0 0 9c0 1.452.348 2.827.957 4.042l3.007-2.332z"/><path fill="#EA4335" d="M9 3.58c1.321 0 2.508.454 3.44 1.345l2.582-2.58C13.463.891 11.426 0 9 0A8.997 8.997 0 0 0 .957 4.958L3.964 7.29C4.672 5.163 6.656 3.58 9 3.58z"/></svg>
          Continue with Google
        </button>
        <div class="auth-divider"><span>or create a username account</span></div>
      </div>
      <input class="auth-field" type="text" id="auth-reg-username" placeholder="Choose a username" autocomplete="username" maxlength="30">
      <input class="auth-field" type="password" id="auth-reg-pw" placeholder="Password (8+ characters)" autocomplete="new-password">
      <input class="auth-field" type="password" id="auth-reg-pw2" placeholder="Confirm password" autocomplete="new-password">
      <input class="auth-field" type="email" id="auth-reg-email" placeholder="Email (optional)" autocomplete="email" style="margin-bottom:4px">
      <div class="auth-note" style="margin-bottom:12px;margin-top:0;color:#aaa">Optional — helps identify your account if you ever need support. Never shared or used for marketing.</div>
      <div class="auth-err" id="auth-reg-err"></div>
      <button class="auth-submit" id="auth-register-submit">Create Account</button>
    </div>
  </div>
</div>

<!-- ── Google new-user welcome modal ── -->
<div id="google-welcome-modal">
  <div class="gw-backdrop"></div>
  <div class="gw-box">
    <h2>Welcome to GC Tracker! 👋</h2>
    <p>You're signed in with Google. Choose a username for your account — it's how you'll appear and sign in if you ever use a password instead.</p>
    <label class="gw-label" for="gw-username">Username</label>
    <input class="auth-field" type="text" id="gw-username" placeholder="Choose a username" autocomplete="username" maxlength="30">
    <div class="gw-msg" id="gw-msg"></div>
    <hr class="gw-divider">
    <div style="color:#aaa;font-size:.82rem;margin-bottom:10px">Already have a GC Tracker account? Import your watch list, want list, and favorites.</div>
    <button class="gw-import-toggle" id="gw-import-toggle">+ Import existing account</button>
    <div class="gw-import-section" id="gw-import-section">
      <div style="color:#888;font-size:.78rem;margin-bottom:10px;line-height:1.5">Enter your existing username above and the password for that account. Your saved data will be moved over and the old account will be removed.</div>
      <label class="gw-label" for="gw-import-pw">Password for existing account</label>
      <input class="auth-field" type="password" id="gw-import-pw" placeholder="Password" autocomplete="current-password">
    </div>
    <button class="auth-submit" id="gw-submit" style="margin-top:6px">Save &amp; Continue</button>
    <div style="text-align:center;margin-top:12px">
      <button id="gw-skip-btn" style="background:none;border:none;color:#555;font-size:.78rem;cursor:pointer;text-decoration:underline">Skip for now</button>
    </div>
  </div>
</div>

<!-- ── Google link nudge banner (for existing password users) ── -->
<div id="google-link-banner">
  🔒 <span>Link Google Sign-In to your account for added security — password-only login will be retired in a future update.</span>
  <button class="glib-link">Link Google Account</button>
  <button class="glib-dismiss">✕ Dismiss</button>
</div>

</div>

<!-- ══ GC PANEL ══ -->
<div class="mobile-title-bar"><button class="mtb-about">About</button><span class="mtb-title">GC Used Inventory Tracker</span><span class="mtb-ver"><!-- __VER__ --></span></div>
<div class="layout">

  <div class="left" id="gc-left">
    <button id="sidebar-collapse-btn" title="Collapse store panel">«</button>
    <div class="sheet-handle"></div>
    <button class="mobile-sidebar-toggle" id="gc-sidebar-toggle">
      <span class="toggle-arrow" id="gc-toggle-arrow">▶</span>
      Stores
      <span class="toggle-count" id="gc-toggle-count"></span>
    </button>
    <div class="search-wrap" id="search-wrap">
      <input id="search" type="text" placeholder="Filter by location name…" autocomplete="off">
      <div class="sel-btns">
        <button class="sel-btn" id="favs-btn">★ Favorites</button>
        <button class="sel-btn" id="sel-all-btn">Select All</button>
      </div>
      <div class="zip-sort-row">
        <button id="zip-sort-btn" title="Sort stores by distance from ZIP">📍 ZIP Sort</button>
        <input id="zip-input" type="text" maxlength="5" placeholder="ZIP code…"
          autocomplete="postal-code" inputmode="numeric" enterkeyhint="go"
>
      </div>
      <div class="zip-radius-row" id="zip-radius-row" style="display:none">
        <label for="zip-radius-select">Within</label>
        <select id="zip-radius-select" title="Only show and search stores within this distance of your ZIP. Stores without a map location — the (?) rows — are excluded by any distance limit.">
          <option value="">Any distance</option>
          <option value="5">5 mi</option>
          <option value="10">10 mi</option>
          <option value="25">25 mi</option>
          <option value="50">50 mi</option>
          <option value="100">100 mi</option>
        </select>
      </div>
    </div>

    <div id="store-list"></div>

    <div class="left-footer">
      <div id="sel-count">0 stores selected</div>
    </div>
  </div>

  <div class="right">
    <div class="status-bar">
      <span id="s-last-wrap">Last checked for new gear: <b id="s-last">—</b> <button id="check-now-btn" style="padding:2px 10px;background:#c00;color:#fff;border:none;border-radius:4px;font-size:.72rem;font-weight:700;cursor:pointer;margin-left:4px;display:none">Scan For New</button> <button id="view-toggle-btn" class="view-toggle-btn" title="Switch card / list view"><span id="view-toggle-icon">⊞</span></button></span>
      <span>Items: <b id="s-known">—</b></span>
      <span>Stores: <b id="s-stores">—</b></span>
      <!-- global-search moved into filter sheet -->
      <span id="s-want-match" style="display:none;color:#4caf50;font-weight:600;font-size:.82rem;cursor:pointer" title="Click to view want list matches"></span>
    </div>
    <div id="log"><span class="log-dim">Ready</span></div>
    <div class="results" id="res-panel" style="display:none">
      <!-- ── Persistent view-toggle chips (always visible, not in filter sheet) ── -->
      <div id="results-top-bar">
      <div class="quick-filter-bar">
        <button id="view-toggle-chip"       class="qf-chip view-toggle-chip-btn" title="Switch list / card view">☰</button>
        <button id="desktop-thumb-toggle" class="qf-chip" title="Show thumbnail grid view">⊞</button>
        <button id="price-drop-toggle" class="qf-chip">↓ Price Drops</button>
        <button id="vintage-toggle" class="qf-chip" title="Show only genuine vintage gear (GC's own classification)">🎸 Vintage</button>
        <div id="ss-wrap" style="display:none;position:relative">
          <button id="saved-searches-btn" class="qf-chip" title="Your saved filter combinations">🔖 Saved Searches</button>
        </div>
        <button id="watchlist-toggle"      class="qf-chip">★ Watch List</button>
        <button id="want-list-toggle"         class="qf-chip">🎯 Want List</button>
        <a id="search-wl-link" class="qf-edit-link" style="display:none;font-size:.75rem">✏︎ Edit Want List</a>
      </div>
      <div class="results-hdr">
        <span id="res-title" style="display:none"></span>
        <span class="badge" id="res-badge" style="display:none!important"></span>
        <button class="mobile-filter-toggle" id="gc-filter-toggle">
          <span class="toggle-arrow" id="gc-filter-arrow">▶</span> Filters
          <span class="filter-active-dot" id="gc-filter-dot"></span>
        </button>
        <div id="gc-filter-collapsible" class="filter-collapsible">
          <!-- ── Mobile sheet header (hidden on desktop) ── -->
          <div class="filter-sheet-header">
            <div class="filter-sheet-handle"></div>
            <div class="filter-sheet-hdr-row">
              <span class="filter-sheet-title">Filters</span>
              <button class="filter-clear-all-btn" id="filter-clear-all-btn">Clear All</button>
            </div>
          </div>
          <!-- ── Scrollable filter content ── -->
          <div class="filter-scroll-body">
            <span id="filter-item-count" style="color:#888;font-size:.78rem;white-space:nowrap;margin-right:6px"></span>
            <!-- ── Mobile sort row (hidden on desktop) ── -->
            <div class="mobile-sort-row" id="mobile-sort-row">
              <span class="mobile-sort-label">Sort:</span>
              <button class="mobile-sort-btn active" data-sort-field="date" data-sort-dir="desc">Newest</button>
              <button class="mobile-sort-btn" data-sort-field="date" data-sort-dir="asc">Oldest</button>
              <button class="mobile-sort-btn" data-sort-field="price" data-sort-dir="asc">Price ↑</button>
              <button class="mobile-sort-btn" data-sort-field="price" data-sort-dir="desc">Price ↓</button>
            </div>
            <!-- Keyword search — at top of sheet, searches all stores globally -->
            <div id="res-search-wrap">
              <span class="res-search-icon">🔍</span>
              <input id="res-search" type="text" placeholder="Search all stores…" autocomplete="off">
              <button id="res-search-clear" title="Clear search" style="display:none;background:none;border:none;color:#888;font-size:.85rem;cursor:pointer;padding:0 4px;line-height:1">✕</button>
              <button id="search-info-btn" title="Search syntax help" style="background:none;border:1px solid #3a3a3a;border-radius:4px;color:#555;font-size:.78rem;cursor:pointer;padding:2px 6px;line-height:1.4;flex-shrink:0">ⓘ</button>
              <div id="search-info-popover">
                <b>Search syntax</b><br>
                <code>Allen</code> — exact word match<br>
                <code>"Jam Pedals"</code> — phrase match<br>
                <code>Thorpy, Dane</code> — must contain both<br>
                <code>OD*</code> — wildcard (OD808, OD-1…)<br>
                <code>*drive*</code> — contains "drive"<br>
                <code>-combo</code> — NOT (exclude; also <code>-"combo amp"</code>)<br>
                <code>fuzz; octave</code> — OR (either matches)<br>
                <code>Mesa -combo: Angel; Trem</code> — colon = apply prefix to every OR branch
              </div>
              <span id="res-search-count"></span>
            </div>
            <!-- ── Price range (mobile: always visible; desktop: see #price-dropdown below) ── -->
            <div class="price-range-mobile">
              <span class="price-range-label">Price</span>
              <div class="price-inputs-row">
                <span class="price-sym">$</span>
                <input id="price-min" type="number" min="0" step="0.01" placeholder="Min"
                  class="price-inp" autocomplete="off" inputmode="decimal">
                <span class="price-sep">–</span>
                <span class="price-sym">$</span>
                <input id="price-max" type="number" min="0" step="0.01" placeholder="Max"
                  class="price-inp" autocomplete="off" inputmode="decimal">
              </div>
            </div>
            <!-- ── Mobile accordion sections (hidden on desktop) ── -->
            <div class="filter-accordion" id="acc-brand">
              <button class="acc-header" data-acc="brand">
                <span class="acc-title">Brand</span>
                <span class="acc-summary" id="acc-brand-summary"></span>
                <span class="acc-arrow" id="acc-brand-arrow">▾</span>
              </button>
              <div class="acc-body" id="acc-brand-body">
                <div class="acc-search-wrap">
                  <input id="acc-brand-search" type="text" placeholder="Search brands…" autocomplete="off">
                </div>
                <div class="acc-list" id="acc-brand-list"></div>
              </div>
            </div>
            <div class="filter-accordion" id="acc-cond">
              <button class="acc-header" data-acc="cond">
                <span class="acc-title">Condition</span>
                <span class="acc-summary" id="acc-cond-summary"></span>
                <span class="acc-arrow" id="acc-cond-arrow">▾</span>
              </button>
              <div class="acc-body" id="acc-cond-body">
                <div class="acc-list" id="acc-cond-list"></div>
              </div>
            </div>
            <div class="filter-accordion" id="acc-cat" style="display:none">
              <button class="acc-header" data-acc="cat">
                <span class="acc-title">Category</span>
                <span class="acc-summary" id="acc-cat-summary"></span>
                <span class="acc-arrow" id="acc-cat-arrow">▾</span>
              </button>
              <div class="acc-body" id="acc-cat-body">
                <div class="acc-list" id="acc-cat-list"></div>
              </div>
            </div>
            <div class="filter-accordion" id="acc-sub" style="display:none">
              <button class="acc-header" data-acc="sub">
                <span class="acc-title">Subcategory</span>
                <span class="acc-summary" id="acc-sub-summary"></span>
                <span class="acc-arrow" id="acc-sub-arrow">▾</span>
              </button>
              <div class="acc-body" id="acc-sub-body">
                <div class="acc-list" id="acc-sub-list"></div>
              </div>
            </div>
            <!-- ── Desktop dropdown filters (hidden on mobile) ── -->
            <div id="brand-dropdown" class="brand-dd" style="display:none;position:relative">
              <button id="brand-dd-btn" class="cat-sel" style="cursor:pointer;white-space:nowrap">All Brands ▾</button>
              <div id="brand-dd-panel" style="display:none;position:fixed;z-index:500;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:6px;width:260px;max-height:320px;overflow:hidden;box-shadow:0 8px 24px rgba(0,0,0,.5)">
                <div style="padding:6px">
                  <input id="brand-dd-search" type="text" placeholder="Search brands…"
                    style="width:100%;padding:6px 10px;background:#252525;border:1px solid #3a3a3a;border-radius:4px;color:#eee;font-size:.82rem;outline:none;box-sizing:border-box"
                    autocomplete="off">
                </div>
                <div id="brand-dd-list" style="overflow-y:auto;max-height:260px"></div>
              </div>
            </div>
            <div id="cond-dropdown" class="cond-dd" style="display:none;position:relative">
              <button id="cond-dd-btn" class="cat-sel" style="cursor:pointer;white-space:nowrap">All Conditions ▾</button>
              <div id="cond-dd-panel" style="display:none;position:fixed;z-index:500;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:6px;width:220px;max-height:300px;overflow:hidden;box-shadow:0 8px 24px rgba(0,0,0,.5)">
                <div style="overflow-y:auto;max-height:260px;padding:4px 0" id="cond-dd-inner"></div>
              </div>
            </div>
            <div id="cat-dropdown" class="cond-dd" style="display:none;position:relative">
              <button id="cat-dd-btn" class="cat-sel" style="cursor:pointer;white-space:nowrap">All Categories ▾</button>
              <div id="cat-dd-panel" style="display:none;position:fixed;z-index:500;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:6px;width:240px;max-height:300px;overflow:hidden;box-shadow:0 8px 24px rgba(0,0,0,.5)">
                <div style="overflow-y:auto;max-height:260px;padding:4px 0" id="cat-dd-inner"></div>
              </div>
            </div>
            <div id="subcat-dropdown" class="cond-dd" style="display:none;position:relative">
              <button id="subcat-dd-btn" class="cat-sel" style="cursor:pointer;white-space:nowrap">All Subcategories ▾</button>
              <div id="subcat-dd-panel" style="display:none;position:fixed;z-index:500;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:6px;width:240px;max-height:300px;overflow:hidden;box-shadow:0 8px 24px rgba(0,0,0,.5)">
                <div style="overflow-y:auto;max-height:260px;padding:4px 0" id="subcat-dd-inner"></div>
              </div>
            </div>
            <!-- ── Price range — desktop dropdown (hidden on mobile via CSS) ── -->
            <div id="price-dropdown" style="display:none;position:relative">
              <button id="price-dd-btn" class="cat-sel" style="cursor:pointer;white-space:nowrap">Price ▾</button>
              <div id="price-dd-panel" style="display:none;position:fixed;z-index:500;background:#1a1a1a;border:1px solid #3a3a3a;border-radius:6px;padding:14px 14px 10px;width:236px;box-shadow:0 8px 24px rgba(0,0,0,.5)">
                <div style="font-size:.72rem;color:#aaa;margin-bottom:9px;text-transform:uppercase;letter-spacing:.05em">Price Range</div>
                <div style="display:flex;align-items:center;gap:6px">
                  <span style="color:#bbb;font-size:.82rem">$</span>
                  <input id="price-min-dd" type="number" min="0" step="0.01" placeholder="Min"
                    style="width:78px;padding:6px 8px;background:#252525;border:1px solid #3a3a3a;border-radius:4px;color:#eee;font-size:.85rem;outline:none;box-sizing:border-box"
                    autocomplete="off" inputmode="decimal">
                  <span style="color:#999;font-size:.85rem">–</span>
                  <span style="color:#bbb;font-size:.82rem">$</span>
                  <input id="price-max-dd" type="number" min="0" step="0.01" placeholder="Max"
                    style="width:78px;padding:6px 8px;background:#252525;border:1px solid #3a3a3a;border-radius:4px;color:#eee;font-size:.85rem;outline:none;box-sizing:border-box"
                    autocomplete="off" inputmode="decimal">
                </div>
                <button id="price-dd-clear" style="display:none;margin-top:10px;background:none;border:none;color:#f88;font-size:.78rem;cursor:pointer;padding:0;line-height:1.4">✕ Clear price filter</button>
              </div>
            </div>
            <!-- Action buttons row (side-by-side on mobile, inline on desktop) -->
            <div id="filter-action-btns" style="display:none;gap:8px">
              <button id="save-search-btn" title="Save current search + filters"
                style="padding:7px 10px;border-radius:4px;background:#1e2e1e;border:1px solid #4ade80;color:#4ade80;font-size:.78rem;cursor:pointer;white-space:nowrap">
                💾 Save Search
              </button>
              <button id="clear-filters-btn"
                style="padding:7px 10px;border-radius:4px;background:#1e1e1e;border:1px solid #c00;color:#f88;font-size:.78rem;cursor:pointer;white-space:nowrap">
                ✕ Clear All
              </button>
            </div>
          </div>
          <!-- ── Pinned Show Results (mobile only) ── -->
          <button class="filter-done-btn">Show Results</button>
        </div>
      </div>
      </div><!-- /results-top-bar -->
      <!-- ss-dropdown lives here (outside overflow-x:auto chip bar) so position:fixed works on iOS -->
      <div id="ss-dropdown" class="ss-dropdown"></div>
      <!-- Shared condition_note tooltip — one node reused for every row, positioned via JS
           (position:fixed escapes the table cell's overflow:hidden, same reason ss-dropdown lives here) -->
      <div id="cond-tooltip" class="cond-tooltip"></div>
      <div id="res-body"></div>
    </div>
  </div>

</div>

<!-- ══ CL PANEL (moved to /cl route) ══ -->
<!-- placeholder: cl-left, cl-sidebar-toggle etc kept for JS refs -->
<div id="cl-panel" style="display:none">
  <div class="cl-left" id="cl-left">
    <button class="mobile-sidebar-toggle" id="cl-sidebar-toggle" style="display:none">
      <span class="toggle-arrow" id="cl-toggle-arrow"></span>
      Cities
      <span class="toggle-count" id="cl-toggle-count"></span>
    </button>
    <div class="search-wrap cl-left">
      <input id="cl-city-search" type="text" placeholder="Search cities…" autocomplete="off">
      <div class="cl-sel-btns">
        <button class="cl-sel-btn" id="cl-favs-btn">★ Favorites</button>
        <button class="cl-sel-btn" id="cl-select-all-btn">Select All</button>
        <button class="cl-sel-btn" id="cl-clear-all-btn">Clear All</button>
      </div>
    </div>
    <div id="cl-city-list"></div>
  </div>

  <!-- Right content: search bar + results -->
  <div class="cl-right">
    <div class="cl-search-bar">
      <input id="cl-query" type="text" placeholder="e.g. telecaster, les paul, fender twin…" autocomplete="off"
>
      <span id="cl-status"></span>
      <button id="cl-search-btn">Search</button>
    </div>
    <div class="cl-results-hdr" id="cl-toolbar" style="display:flex;align-items:center;gap:8px">
      <button id="cl-watchlist-toggle"
        class="cat-sel" style="border-color:#3a3a3a;color:#aaa;cursor:pointer;white-space:nowrap;font-size:.78rem;padding:5px 10px">
        ★ Watch List
      </button>
      <button id="cl-stub-open-kw-btn"
        class="cat-sel" style="border-color:#2d6a2d;color:#4ade80;cursor:pointer;white-space:nowrap;font-size:.78rem;padding:5px 10px">
        🎯 Want List
      </button>
      <a id="cl-search-wl-link" style="color:#4ade80;cursor:pointer;white-space:nowrap;font-size:.78rem;text-decoration:none;margin-left:2px">Search Want List</a>
    </div>
    <div class="cl-results-hdr" id="cl-results-hdr" style="display:none">
      <span id="cl-count"></span>
      <input id="cl-res-search" type="text" placeholder="Filter results…" autocomplete="off">
    </div>
    <div id="cl-body"><div class="cl-empty">Select cities on the left, enter a search term, and click Search.</div></div>
  </div>

</div>

<!-- ── Store sheet backdrop (mobile only) ── -->
<div class="store-sheet-backdrop" id="store-sheet-backdrop"></div>

<!-- ── Mobile bottom action bar (hidden on desktop via CSS) ── -->
<div class="mobile-bottom-bar" id="mobile-bottom-bar">
  <button class="mbb-btn mbb-check" id="mbb-check">
    <span class="mbb-icon" id="mbb-check-icon">▶</span>
    <span class="mbb-label" id="mbb-check-label">Scan For New</span>
  </button>
  <button class="mbb-btn" id="mbb-filters">
    <span class="mbb-icon">🔍</span>
    <span class="mbb-label">Filter & Sort</span>
    <span class="mbb-dot" id="mbb-filter-dot"></span>
  </button>
  <button class="mbb-btn" id="mbb-stores">
    <span class="mbb-icon">🏪</span>
    <span class="mbb-label">Stores</span>
  </button>
  <button class="mbb-btn" id="mbb-auth">
    <span class="mbb-icon" id="mbb-auth-icon">👤</span>
    <span class="mbb-label" id="mbb-auth-label">Sign In</span>
  </button>
</div>

<script src="/static/gc.js"></script>

<div id="dev-footer">
  <span>Buy the developer a pack of strings</span>
  <a href="https://paypal.me/smurfco" target="_blank" rel="noopener" title="PayPal">
    <svg width="18" height="18" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
      <path d="M19.5 8.5c.3-2-1.2-3.5-3.5-3.5H9.5L7 20h3l.7-4.5h2.3c3.5 0 6-2 6.5-5.5l.5-1.5z" fill="#009cde"/>
      <path d="M16 10.5c.2-1.5-.8-2.5-2.5-2.5H9l-1.5 9h2.5l.5-3h2c2.5 0 4-1.5 4.3-3.5l.2-.5z" fill="#003087"/>
    </svg>
  </a>
  <a href="https://account.venmo.com/u/charles-boehmig" target="_blank" rel="noopener" title="Venmo">
    <svg width="18" height="18" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
      <rect width="24" height="24" rx="4" fill="#3D95CE"/>
      <path d="M17 5.5c.5 1 .7 2 .7 3.3 0 4-3.4 9.2-6.2 12.7H7.3L5 6.3l4-.4 1.3 10.2C11.6 14 13 10.8 13 8.3c0-1.3-.2-2.3-.6-3L17 5.5z" fill="#fff"/>
    </svg>
  </a>
  <span style="margin-left:4px">·</span>
  <a href="https://animalsintrees.com" target="_blank" rel="noopener" title="Animals in Trees" style="gap:5px">
    My music
    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
      <path d="M9 18V5l12-2v13" stroke="#aaa" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/>
      <circle cx="6" cy="18" r="3" stroke="#aaa" stroke-width="1.5"/>
      <circle cx="18" cy="16" r="3" stroke="#aaa" stroke-width="1.5"/>
    </svg>
  </a>
  <span style="margin-left:4px">·</span>
  <a href="#" data-action="open-about">About</a>
  <span style="margin-left:4px">·</span>
  <a href="/privacy">Privacy Policy</a>
  <span id="admin-footer-sep" style="display:none;margin-left:4px">·</span>
  <a id="admin-footer-link" href="/admin/users" style="display:none;margin-left:0;color:#888;font-size:11px">Admin</a>
</div>

<!-- ── About modal ── -->
<div id="about-modal">
  <div id="about-box">
    <h3>GC Used Inventory Tracker</h3>
    <div class="about-sub">Developed by CKB</div>
    <p style="font-size:.82rem;color:#aaa;line-height:1.55;margin:12px 0 4px;text-align:center">A free tool for tracking Guitar Center's used instrument inventory. Scan for new listings, build a watch list, set up a want list, and save searches — all synced across your devices.</p>
    <p style="font-size:.75rem;color:#666;line-height:1.45;margin:0 0 10px;text-align:center;font-style:italic">Independent tool — not affiliated with or endorsed by Guitar Center, Inc.</p>
    <div class="about-donate-row">
      <span class="about-donate-label">Donate</span>
      <a href="https://paypal.me/smurfco" target="_blank" rel="noopener" title="PayPal">
        <svg width="22" height="22" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
          <path d="M19.5 8.5c.3-2-1.2-3.5-3.5-3.5H9.5L7 20h3l.7-4.5h2.3c3.5 0 6-2 6.5-5.5l.5-1.5z" fill="#009cde"/>
          <path d="M16 10.5c.2-1.5-.8-2.5-2.5-2.5H9l-1.5 9h2.5l.5-3h2c2.5 0 4-1.5 4.3-3.5l.2-.5z" fill="#003087"/>
        </svg>
      </a>
      <a href="https://account.venmo.com/u/charles-boehmig" target="_blank" rel="noopener" title="Venmo">
        <svg width="22" height="22" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
          <rect width="24" height="24" rx="4" fill="#3D95CE"/>
          <path d="M17 5.5c.5 1 .7 2 .7 3.3 0 4-3.4 9.2-6.2 12.7H7.3L5 6.3l4-.4 1.3 10.2C11.6 14 13 10.8 13 8.3c0-1.3-.2-2.3-.6-3L17 5.5z" fill="#fff"/>
        </svg>
      </a>
    </div>
    <a href="https://animalsintrees.com" target="_blank" rel="noopener" class="about-music-link">
      My music
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
        <path d="M9 18V5l12-2v13" stroke="#aaa" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/>
        <circle cx="6" cy="18" r="3" stroke="#aaa" stroke-width="1.5"/>
        <circle cx="18" cy="16" r="3" stroke="#aaa" stroke-width="1.5"/>
      </svg>
    </a>
    <a href="/privacy" target="_blank" rel="noopener" style="display:block;margin-top:8px;font-size:.78rem;color:#666;text-align:center;text-decoration:none">Privacy Policy</a>
    <button class="about-close-btn">Close</button>
  </div>
</div>

<!-- __STORES_NOSCRIPT__ -->
<footer class="seo-footer">
  <a href="/privacy">Privacy Policy</a> &middot; Not affiliated with Guitar Center, Inc.
</footer>

</body>
</html>"""

PRIVACY_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Privacy Policy — GC Used Inventory Tracker</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#111;color:#ccc;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;font-size:15px;line-height:1.7;padding:40px 20px 80px}
.wrap{max-width:700px;margin:0 auto}
a{color:#f88;text-decoration:none}
a:hover{text-decoration:underline}
h1{color:#fff;font-size:1.5rem;margin-bottom:6px}
.subtitle{color:#888;font-size:.85rem;margin-bottom:36px}
h2{color:#eee;font-size:1rem;font-weight:700;margin:32px 0 10px;padding-bottom:6px;border-bottom:1px solid #2a2a2a}
p{margin-bottom:14px}
ul{margin:0 0 14px 20px}
ul li{margin-bottom:6px}
.back{display:inline-block;margin-bottom:28px;color:#888;font-size:.85rem}
.back:hover{color:#ccc}
footer{margin-top:48px;padding-top:16px;border-top:1px solid #222;color:#555;font-size:.8rem}
</style>
</head>
<body>
<div class="wrap">
  <a href="/" class="back">← Back to GC Used Inventory Tracker</a>
  <h1>Privacy Policy</h1>
  <p class="subtitle">Last updated: May 2026</p>

  <p>GC Used Inventory Tracker ("the site", "we", "us") is an independent personal project that
  helps musicians track used gear listings at Guitar Center. It is not affiliated with, sponsored
  by, or endorsed by Guitar Center, Inc. This policy explains what information we collect, how
  we use it, and your rights regarding that information.</p>

  <h2>Information We Collect</h2>

  <p><strong style="color:#eee">Account information.</strong> If you create an account, we store
  your chosen username, an optional email address, and a hashed (never plain-text) version of your
  password. If you sign in with Google, we store your Google account ID and the display name Google
  provides. Your email address is never required and is used only for account recovery if you
  choose to provide it.</p>

  <p><strong style="color:#eee">Preferences and scan history.</strong> To sync your data across
  devices, we store your watch list, want list keywords, favorited stores, saved searches, and the
  timestamp and item IDs from your most recent scan. This data lives on our server and is tied to
  your account.</p>

  <p><strong style="color:#eee">Technical data.</strong> When you use the site, our server
  receives your IP address. We use it only for rate-limiting (to prevent abuse) and do not log or
  store it persistently. We also set an anonymous device ID cookie (<code style="color:#aaa;font-size:.85em">gt_device_id</code>)
  to count unique devices in aggregate — it contains no personal information.</p>

  <p><strong style="color:#eee">Analytics.</strong> We use Google Analytics 4 to understand how
  visitors use the site in aggregate (page views, session counts, general geography). Google
  Analytics may set its own cookies in your browser. You can opt out using the
  <a href="https://tools.google.com/dlpage/gaoptout" target="_blank" rel="noopener">Google Analytics
  Opt-out Browser Add-on</a>.</p>

  <h2>How We Use Your Information</h2>
  <ul>
    <li>To provide the core tracker functionality (scan results, watch list, want list)</li>
    <li>To sync your preferences across your own devices when you are logged in</li>
    <li>To prevent abuse via rate limiting on scan and login endpoints</li>
    <li>To understand aggregate site usage through analytics</li>
  </ul>
  <p>We do not sell, rent, or share your personal information with third parties for their
  marketing purposes.</p>

  <h2>Cookies</h2>
  <ul>
    <li><strong style="color:#eee">Session cookie</strong> — keeps you logged in across browser
    sessions. Set by Flask, signed with a server secret, HttpOnly and Secure.</li>
    <li><strong style="color:#eee">gt_device_id</strong> — anonymous device identifier for
    internal usage counting. Contains no personal information.</li>
    <li><strong style="color:#eee">Google Analytics cookies</strong> (_ga, _gid, and related)
    — set by Google's analytics script to measure aggregate traffic.</li>
  </ul>

  <h2>Third-Party Services</h2>
  <ul>
    <li><strong style="color:#eee">Google Analytics</strong> — aggregate usage data.
    <a href="https://policies.google.com/privacy" target="_blank" rel="noopener">Google Privacy Policy</a>.</li>
    <li><strong style="color:#eee">Google Sign-In (OAuth)</strong> — optional login method.
    We receive your Google ID and display name only. We do not receive your Google contacts,
    Drive files, or any other Google data.</li>
    <li><strong style="color:#eee">Railway</strong> — the cloud platform that hosts the site.
    Your data is stored on Railway's infrastructure in the United States.</li>
  </ul>

  <h2>Data Retention</h2>
  <p>Your account and associated data are retained until you request deletion. You can request
  that your account be deleted at any time by contacting us at the address below. Guest users
  (no account) have no data stored on our servers beyond the anonymous device ID cookie.</p>

  <h2>Children's Privacy</h2>
  <p>This site is not directed at children under 13. We do not knowingly collect personal
  information from children.</p>

  <h2>Changes to This Policy</h2>
  <p>If we make material changes to this policy, we will update the "Last updated" date at the
  top of this page. Continued use of the site after changes are posted constitutes acceptance
  of the updated policy.</p>

  <h2>Contact</h2>
  <p>Questions about this privacy policy or your data can be sent to:
  <a href="mailto:cboehmig@gmail.com">cboehmig@gmail.com</a></p>

  <footer>GC Used Inventory Tracker is an independent tool and is not affiliated with or
  endorsed by Guitar Center, Inc.</footer>
</div>
</body>
</html>"""

# ── Google Analytics ──────────────────────────────────────────────────────────
if GA_MEASUREMENT_ID:
    # Only the async src= loader — no inline script block.
    # The gtag('config', ...) init lives in static/gc.js and static/cl.js,
    # which read the GA ID from the <meta name="ga-id"> tag below.
    _ga_snippet = (
        f'<!-- Google tag (gtag.js) -->\n'
        f'<script async src="https://www.googletagmanager.com/gtag/js?id={GA_MEASUREMENT_ID}"></script>\n'
        f'<meta name="ga-id" content="{GA_MEASUREMENT_ID}">'
    )
else:
    _ga_snippet = ''
APP_VERSION = "2.16.21"
HTML_TEMPLATE    = HTML_TEMPLATE.replace('<!-- __GA__ -->', _ga_snippet)
HTML_TEMPLATE    = HTML_TEMPLATE.replace('<!-- __VER__ -->', f'v{APP_VERSION}')
CL_TEMPLATE      = CL_TEMPLATE.replace('<!-- __GA__ -->', _ga_snippet)
NEWDEALS_TEMPLATE = NEWDEALS_TEMPLATE.replace('<!-- __GA__ -->', _ga_snippet)

# Cache-busting: version every local static JS/CSS reference so the 1-year
# SEND_FILE_MAX_AGE_DEFAULT above can never serve a stale asset across deploys.
# (2026-07 audit S7)
def _version_static(tpl: str) -> str:
    return re.sub(r'(/static/[a-zA-Z0-9._-]+\.(?:js|css))', rf'\1?v={APP_VERSION}', tpl)

HTML_TEMPLATE     = _version_static(HTML_TEMPLATE)
CL_TEMPLATE       = _version_static(CL_TEMPLATE)
NEWDEALS_TEMPLATE = _version_static(NEWDEALS_TEMPLATE)

# __STORES_NOSCRIPT__ is replaced at request time in index() so it always reflects
# the live store cache — see the index() route handler below.




# ── Startup bootstrap ─────────────────────────────────────────────────────────
# Runs unconditionally at import time (not just under `if __name__ == "__main__"`)
# so it executes the same way whether the app is launched directly (`python
# gc_tracker_app.py`, dev server) or imported by a WSGI server like gunicorn
# (`gunicorn gc_tracker_app:app`, v2.16.10) — gunicorn never runs the module as
# __main__, so anything load-bearing has to live out here or it silently never
# runs in production. _load_cat_cache() also self-heals via lazy-load calls
# sprinkled through the route handlers, but _load_cookies() has no other call
# site, so it MUST run here.
_load_cat_cache()
_load_cookies()
if not STORES_CACHE.exists():
    print("Building store list…")
    refresh_store_list()

# ── Periodic malloc_trim ──────────────────────────────────────────────────────
# Root cause (investigated 2026-08-31): Railway production memory climbs from
# ~2GB to 7-8GB+ between deploys and never comes back down. Confirmed via local
# load testing that this is NOT a Python-level leak — it's glibc's allocator
# holding onto pages at their high-water mark. Under concurrent /api/browse
# traffic (large per-request temporary lists built from the ~92K-item catalog,
# --threads=8), glibc grabs extra heap to serve simultaneous large allocations
# and, once freed, keeps those pages reserved rather than returning them to the
# OS — classic RSS high-water-mark behavior, not a growing live working set.
# Confirmed locally that calling libc's malloc_trim(0) reliably reclaims it
# (reproduced twice: ~450MB baseline → ~1.2GB after load → ~410MB after trim,
# both rounds). This runs a trim on a timer so RSS gets reclaimed periodically
# in production without needing a deploy to reset it. Linux-only (Railway) —
# no-ops harmlessly if libc.so.6 isn't available (e.g. local dev on macOS).
_MALLOC_TRIM_INTERVAL_SECS = 300  # 5 minutes

def _malloc_trim_loop():
    try:
        _libc = _ctypes.CDLL("libc.so.6")
    except OSError:
        return  # not on glibc/Linux (e.g. local macOS dev) — nothing to do
    while True:
        time.sleep(_MALLOC_TRIM_INTERVAL_SECS)
        try:
            _gc.collect()
            _libc.malloc_trim(0)
        except Exception:
            pass  # best-effort — never let this background thread take the app down

threading.Thread(target=_malloc_trim_loop, daemon=True).start()

# Nightly scan removed — "Check for New" is manual only

if __name__ == "__main__":
    url = f"http://localhost:{PORT}"
    print(f"\n  Guitar Center Tracker v{APP_VERSION} is running!")
    print(f"  Open: {url}")
    print(f"  Press Ctrl+C to stop.\n")
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    app.run(host="0.0.0.0", port=PORT, threaded=True, debug=False)
