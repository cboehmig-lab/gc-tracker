#!/usr/bin/env python3
"""
Guitar Center Used Inventory Tracker
-------------------------------------
Tracks new used gear listings at the Austin & South Austin Guitar Center stores.

Usage:
    python3 gc_inventory_tracker.py

First run:  saves current inventory as your baseline (nothing written to Excel yet).
Later runs: new items are inserted at the TOP of gc_new_inventory.xlsx with a "New" label.

Files created in the same folder as this script:
    gc_state.json          — remembers which items you've already seen
    gc_new_inventory.xlsx  — cumulative log of new finds
"""

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing 'requests'. Run:  pip3 install requests openpyxl")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing 'openpyxl'. Run:  pip3 install openpyxl")

# ── Configuration ─────────────────────────────────────────────────────────────

SCRIPT_DIR  = Path(__file__).parent
STATE_FILE  = SCRIPT_DIR / "gc_state.json"
OUTPUT_FILE = SCRIPT_DIR / "gc_new_inventory.xlsx"

BASE_URL      = "https://www.guitarcenter.com/Used/"
PAGE_SIZE     = 24
REQUEST_DELAY = 2.0
MAX_PAGES     = 50

STORES = [
    ("Austin",       "filters=stores:Austin&Ns=cD"),
    ("South Austin", "filters=stores:South%20Austin&Ns=cD"),
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── HTTP ──────────────────────────────────────────────────────────────────────

_session = requests.Session()
_session.headers.update(HEADERS)

def fetch_page(query: str, page: int = 1) -> str:
    url = f"{BASE_URL}?{query}&page={page}"
    r = _session.get(url, timeout=20)
    r.raise_for_status()
    return r.text


# ── Parsing ───────────────────────────────────────────────────────────────────

def parse_products(html: str, store_name: str) -> list[dict]:
    blocks = re.findall(
        r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>',
        html, re.DOTALL
    )

    for block in blocks:
        try:
            data = json.loads(block)
        except json.JSONDecodeError:
            continue

        if data.get("@type") != "CollectionPage":
            continue

        items = data.get("mainEntity", {}).get("itemListElement", [])
        if not items:
            continue

        products = []
        for entry in items:
            item  = entry.get("item", {})
            name  = item.get("name", "").strip()
            url   = item.get("url", "").strip()
            sku   = item.get("sku", "").strip()
            raw_price = item.get("offers", {}).get("price", "")
            try:
                price = float(raw_price) if raw_price else None
            except ValueError:
                price = None

            if not name or not sku:
                continue

            products.append({
                "id":    sku,
                "name":  name,
                "price": price,   # stored as float for Excel currency formatting
                "store": store_name,
                "url":   url,
            })
        return products

    debug_path = SCRIPT_DIR / "gc_debug_page.html"
    debug_path.write_text(html, encoding="utf-8")
    print(f"\n  Could not find product data. Saved HTML to: {debug_path}")
    return []


# ── Fetching (per store) ──────────────────────────────────────────────────────

def fetch_all_for_store(store_name: str, query: str, seen_ids: set) -> list[dict]:
    all_products: list[dict] = []
    ids_this_store: set = set()
    page = 1

    while page <= MAX_PAGES:
        print(f"    Page {page}...", end=" ", flush=True)
        try:
            html = fetch_page(query, page)
        except requests.HTTPError as e:
            print(f"\n  HTTP error {e.response.status_code}: {e}")
            sys.exit(1)
        except requests.RequestException as e:
            print(f"\n  Network error: {e}")
            sys.exit(1)

        products = parse_products(html, store_name)
        print(f"{len(products)} items")

        if not products:
            break

        # Detect looping (site returning same page repeatedly)
        if all(p["id"] in ids_this_store for p in products):
            print("  (Duplicate page detected — stopping.)")
            break

        new_this_page = [p for p in products if p["id"] not in ids_this_store]
        all_products.extend(new_this_page)
        for p in products:
            ids_this_store.add(p["id"])

        # Stop early if everything on this page was already known
        if seen_ids and all(p["id"] in seen_ids for p in products):
            print("    All items on this page already seen. Stopping early.")
            break

        if len(products) < PAGE_SIZE:
            break

        page += 1
        time.sleep(REQUEST_DELAY)

    return all_products


# ── State ─────────────────────────────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"last_run": None, "seen_ids": []}


def save_state(seen_ids: list[str], run_time: str):
    STATE_FILE.write_text(
        json.dumps({"last_run": run_time, "seen_ids": seen_ids}, indent=2),
        encoding="utf-8",
    )


# ── Excel output ──────────────────────────────────────────────────────────────

# Column layout: Status | Date Found | Item Name | Price | Store | Link
_COLS   = ["Status", "Date Found", "Item Name", "Price", "Store", "Link"]
_WIDTHS = [8, 18, 58, 12, 16, 70]
_NUM_COLS = len(_COLS)

_HDR_FILL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
_HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_ROW_FONT  = Font(name="Arial", size=10)
_NEW_FONT  = Font(name="Arial", bold=True, size=10)
_ALT_FILL  = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
_CURRENCY  = '$#,##0.00'


def _apply_row_formatting(ws, row_idx: int):
    """Apply alternating fill and base font to a data row."""
    fill = _ALT_FILL if row_idx % 2 == 0 else None
    for col in range(1, _NUM_COLS + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = _ROW_FONT
        if fill:
            c.fill = fill


def write_excel(new_items: list[dict]):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    n = len(new_items)

    if OUTPUT_FILE.exists():
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active

        # Insert blank rows right after the header to make room for new items
        ws.insert_rows(2, amount=n)

        # Write new items into the freshly inserted rows
        for i, item in enumerate(new_items):
            row_idx = 2 + i
            ws.cell(row=row_idx, column=1, value="New").font = _NEW_FONT
            ws.cell(row=row_idx, column=2, value=timestamp)
            ws.cell(row=row_idx, column=3, value=item["name"])
            price_cell = ws.cell(row=row_idx, column=4, value=item["price"])
            price_cell.number_format = _CURRENCY
            ws.cell(row=row_idx, column=5, value=item["store"])
            url_cell = ws.cell(row=row_idx, column=6, value=item["url"] or "")
            if item["url"]:
                url_cell.hyperlink = item["url"]
                url_cell.style = "Hyperlink"
            _apply_row_formatting(ws, row_idx)
            # Re-apply bold "New" after _apply_row_formatting resets the font
            ws.cell(row=row_idx, column=1).font = _NEW_FONT

        # Re-stripe all existing (now shifted) data rows for consistent formatting
        for row_idx in range(2 + n, ws.max_row + 1):
            _apply_row_formatting(ws, row_idx)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "New Inventory"
        ws.freeze_panes = "A2"

        # Header row
        ws.append(_COLS)
        for col_idx in range(1, _NUM_COLS + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = _HDR_FILL
            cell.font      = _HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        for col_idx, width in enumerate(_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for i, item in enumerate(new_items):
            row_idx = 2 + i
            ws.cell(row=row_idx, column=1, value="New")
            ws.cell(row=row_idx, column=2, value=timestamp)
            ws.cell(row=row_idx, column=3, value=item["name"])
            price_cell = ws.cell(row=row_idx, column=4, value=item["price"])
            price_cell.number_format = _CURRENCY
            ws.cell(row=row_idx, column=5, value=item["store"])
            url_cell = ws.cell(row=row_idx, column=6, value=item["url"] or "")
            if item["url"]:
                url_cell.hyperlink = item["url"]
                url_cell.style = "Hyperlink"
            _apply_row_formatting(ws, row_idx)
            ws.cell(row=row_idx, column=1).font = _NEW_FONT

    wb.save(OUTPUT_FILE)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    state    = load_state()
    seen_ids = set(state["seen_ids"])
    last_run = state.get("last_run")
    run_time = datetime.now().isoformat()

    print("=" * 52)
    print("  Guitar Center Used Inventory Tracker")
    print("=" * 52)
    print(f"  Last run: {last_run or 'never (this is your first run)'}")
    print(f"  Known items: {len(seen_ids)}")
    print()

    all_products: list[dict] = []
    ids_this_run: set = set()

    for store_name, query in STORES:
        print(f"  [{store_name}]")
        store_products = fetch_all_for_store(store_name, query, seen_ids)
        # Deduplicate across stores (shouldn't happen with used gear, but just in case)
        for p in store_products:
            if p["id"] not in ids_this_run:
                all_products.append(p)
                ids_this_run.add(p["id"])
        print()

    new_items = [p for p in all_products if p["id"] not in seen_ids]

    all_ids = list(seen_ids | ids_this_run)
    save_state(all_ids, run_time)

    print(f"  Items scanned : {len(all_products)}")
    print(f"  New items     : {len(new_items)}")
    print()

    if new_items:
        write_excel(new_items)
        print(f"  Excel updated -> {OUTPUT_FILE.name}")
        print()
        print("  New items:")
        for item in new_items:
            price_str = f"${item['price']:,.2f}" if item["price"] else ""
            print(f"    - [{item['store']}] {item['name']}  {price_str}")
    else:
        if last_run:
            print("  No new items since last run.")
        else:
            print(
                "  First run complete -- baseline saved.\n"
                "  Run again any time to see what's new!"
            )
    print()


if __name__ == "__main__":
    main()
